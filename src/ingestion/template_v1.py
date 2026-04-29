"""Excel template generator for the V1 payroll spreadsheet.

The workbook prioritizes human-friendly entry in ``LANCAMENTOS_FACEIS`` while
keeping technical tabs ready for the canonical pipeline.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_V1_FILENAME = "planilha_padrao_folha_v1.xlsx"
MAX_DATA_ROWS = 1000

SHEET_ORDER = (
    "PARAMETROS",
    "FUNCIONARIOS",
    "LANCAMENTOS_FACEIS",
    "MOVIMENTOS_CANONICOS",
    "PENDENCIAS",
    "LISTAS",
)

PARAMETROS_HEADERS = ("campo", "valor", "obrigatorio", "descricao", "exemplo")
PARAMETROS_ROWS = (
    ("empresa_codigo", "", "sim", "Codigo da empresa ou filial do arquivo", "72"),
    ("empresa_nome", "", "sim", "Nome legivel da empresa", "Dela More"),
    ("competencia", "", "sim", "Competencia da folha no formato MM/AAAA", "03/2024"),
    ("tipo_folha", "mensal", "sim", "Tipo de folha suportado nesta versao", "mensal"),
    ("processo_padrao", "", "sim", "Codigo de processo usado na competencia", "11"),
    ("versao_layout", "v1", "sim", "Versao do template e do layout alvo", "v1"),
    (
        "responsavel_preenchimento",
        "",
        "sim",
        "Nome de quem consolidou o preenchimento",
        "Maria Silva",
    ),
    ("data_referencia", "", "sim", "Data base do fechamento", "31/03/2024"),
    (
        "observacoes_gerais",
        "",
        "sim",
        "Observacoes gerais do arquivo e do mes",
        "Lancamentos revisados antes da importacao",
    ),
)

FUNCIONARIOS_HEADERS = (
    "chave_colaborador",
    "nome_colaborador",
    "cpf",
    "codigo_interno",
    "matricula_dominio",
    "departamento",
    "centro_custo",
    "status_colaborador",
    "admite_lancamento",
    "observacao_identificacao",
)

LANCAMENTOS_FACEIS_HEADERS = (
    "linha_status",
    "chave_colaborador",
    "nome_colaborador",
    "matricula_dominio",
    "departamento",
    "observacao_geral",
    "horas_extras_50",
    "gratificacao",
    "bonus",
    "bonus_vendas",
    "pontualidade",
    "ajuda_custo",
    "reembolso_plano_saude",
    "vale_transporte",
    "farmacia",
    "mercadoria",
    "plano_saude",
    "faltas_dias",
    "atrasos_horas",
    "desconto_adiantamento",
    "observacao_eventos",
    "horas_extras_70",
    "horas_extras_100",
    "hora_extra_noturna",
    "faltas_dsr",
)

LANCAMENTOS_PROFILE_EXTENSION_HEADERS = (
    "horas_extras_70",
    "horas_extras_100",
    "hora_extra_noturna",
    "faltas_dsr",
)

LANCAMENTOS_FACEIS_REQUIRED_HEADERS = tuple(
    header for header in LANCAMENTOS_FACEIS_HEADERS if header not in LANCAMENTOS_PROFILE_EXTENSION_HEADERS
)

MOVIMENTOS_CANONICOS_HEADERS = (
    "id_movimento",
    "empresa_codigo",
    "competencia",
    "tipo_folha",
    "processo_padrao",
    "chave_colaborador",
    "nome_colaborador",
    "matricula_dominio",
    "evento_negocio",
    "rubrica_informada",
    "rubrica_saida",
    "natureza",
    "tipo_valor",
    "quantidade",
    "valor",
    "unidade_serializacao",
    "origem_aba",
    "origem_celula",
    "origem_coluna",
    "pendencia",
    "codigo_pendencia",
    "mensagem_pendencia",
    "observacao",
)

PENDENCIAS_HEADERS = (
    "id_pendencia",
    "severidade",
    "empresa_codigo",
    "competencia",
    "chave_colaborador",
    "nome_colaborador",
    "matricula_dominio",
    "evento_negocio",
    "origem_celula",
    "tipo_pendencia",
    "descricao",
    "acao_recomendada",
    "status_tratamento",
    "resolucao_manual",
    "resolvido_por",
    "resolvido_em",
)

LIST_BLOCKS: dict[str, tuple[str, ...]] = {
    "tipos_folha": ("mensal",),
    "status_colaborador": ("ativo", "afastado", "rescindido", "ferias", "ignorar"),
    "sim_nao_revisar": ("sim", "nao", "revisar"),
    "severidade_pendencia": ("baixa", "media", "alta", "bloqueante"),
    "status_tratamento": ("aberta", "em_analise", "resolvida", "ignorada"),
    "eventos_negocio_v1": (
        "horas_extras_50",
        "gratificacao",
        "bonus",
        "bonus_vendas",
        "pontualidade",
        "ajuda_custo",
        "reembolso_plano_saude",
        "vale_transporte",
        "farmacia",
        "mercadoria",
        "plano_saude",
        "faltas_dias",
        "atrasos_horas",
        "desconto_adiantamento",
        "horas_extras_70",
        "horas_extras_100",
        "hora_extra_noturna",
        "faltas_dsr",
    ),
    # Preparacao tecnica para a aba canonica; ainda nao dirigem regra automatica.
    "tipos_valor": ("monetario", "horas", "dias", "texto"),
    "natureza_evento": ("provento", "desconto", "informativo"),
}

HOUR_COLUMNS = ("G", "S", "V", "W", "X")
MONETARY_COLUMNS = ("H", "I", "J", "K", "L", "M", "O", "P", "Q", "T")
NUMBER_COLUMNS = ("R", "Y")

THIN_SIDE = Side(style="thin", color="D9E2F3")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ID_FILL = PatternFill("solid", fgColor="D9EAF7")
EVENT_FILL = PatternFill("solid", fgColor="E2F0D9")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
TECH_FILL = PatternFill("solid", fgColor="E7E6E6")
REQUIRED_FILL = PatternFill("solid", fgColor="FCE4D6")
WARNING_FILL = PatternFill("solid", fgColor="F8CBAD")
LIST_FILL = PatternFill("solid", fgColor="DDEBF7")


def create_planilha_padrao_folha_v1(max_data_rows: int = MAX_DATA_ROWS) -> Workbook:
    """Build the V1 workbook template in memory."""

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    workbook.properties.title = "Planilha padrao folha V1"
    workbook.properties.subject = "Entrada humana e base canonica para o motor TXT"
    workbook.properties.creator = "motor-txt-dominio-folha"

    parametros = workbook.create_sheet("PARAMETROS")
    funcionarios = workbook.create_sheet("FUNCIONARIOS")
    lancamentos = workbook.create_sheet("LANCAMENTOS_FACEIS")
    movimentos = workbook.create_sheet("MOVIMENTOS_CANONICOS")
    pendencias = workbook.create_sheet("PENDENCIAS")
    listas = workbook.create_sheet("LISTAS")

    _build_parametros_sheet(parametros)
    _build_funcionarios_sheet(funcionarios, max_data_rows=max_data_rows)
    _build_lancamentos_faceis_sheet(lancamentos, max_data_rows=max_data_rows)
    _build_movimentos_canonicos_sheet(movimentos, max_data_rows=max_data_rows)
    _build_pendencias_sheet(pendencias, max_data_rows=max_data_rows)
    _build_listas_sheet(workbook, listas)

    workbook.active = workbook.sheetnames.index("LANCAMENTOS_FACEIS")
    return workbook


def save_planilha_padrao_folha_v1(path: str | Path, max_data_rows: int = MAX_DATA_ROWS) -> Path:
    """Persist the workbook template to disk."""

    target_path = Path(path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = create_planilha_padrao_folha_v1(max_data_rows=max_data_rows)
    workbook.save(target_path)
    return target_path


def _build_parametros_sheet(worksheet) -> None:
    _write_headers(worksheet, PARAMETROS_HEADERS)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = True
    worksheet.sheet_properties.tabColor = "1F4E78"

    widths = {
        "A": 28,
        "B": 26,
        "C": 14,
        "D": 56,
        "E": 28,
    }
    _set_column_widths(worksheet, widths)

    for row_index, row_values in enumerate(PARAMETROS_ROWS, start=2):
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.cell(row=row_index, column=1).fill = ID_FILL
        worksheet.cell(row=row_index, column=3).fill = REQUIRED_FILL

    worksheet["B2"].comment = Comment("Informe o codigo da empresa ou filial.", "motor")
    worksheet["B4"].comment = Comment("Use sempre MM/AAAA, por exemplo 03/2024.", "motor")
    worksheet["B5"].comment = Comment("V1 suporta apenas folha mensal.", "motor")
    worksheet["B6"].comment = Comment("Campo obrigatorio para o processo da competencia.", "motor")
    worksheet["B9"].number_format = "DD/MM/YYYY"

    mandatory_rule = FormulaRule(
        formula=['AND($C2="sim",LEN(TRIM($B2))=0)'],
        stopIfTrue=False,
        fill=WARNING_FILL,
    )
    worksheet.conditional_formatting.add(f"B2:B{len(PARAMETROS_ROWS) + 1}", mandatory_rule)

    competencia_validation = DataValidation(
        type="custom",
        formula1='OR(B4="",AND(LEN(B4)=7,MID(B4,3,1)="/",ISNUMBER(--LEFT(B4,2)),ISNUMBER(--RIGHT(B4,4)),--LEFT(B4,2)>=1,--LEFT(B4,2)<=12,--RIGHT(B4,4)>=2000))',
        allow_blank=True,
    )
    competencia_validation.promptTitle = "Competencia"
    competencia_validation.prompt = "Digite no formato MM/AAAA."
    competencia_validation.errorTitle = "Competencia invalida"
    competencia_validation.error = "Use o formato MM/AAAA, por exemplo 03/2024."
    worksheet.add_data_validation(competencia_validation)
    competencia_validation.add("B4")

    tipo_folha_validation = DataValidation(type="list", formula1="=tipos_folha", allow_blank=False)
    tipo_folha_validation.promptTitle = "Tipo de folha"
    tipo_folha_validation.prompt = "A versao V1 aceita apenas mensal."
    worksheet.add_data_validation(tipo_folha_validation)
    tipo_folha_validation.add("B5")


def _build_funcionarios_sheet(worksheet, max_data_rows: int) -> None:
    _write_headers(worksheet, FUNCIONARIOS_HEADERS)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:J{max_data_rows}"
    worksheet.sheet_properties.tabColor = "5B9BD5"

    widths = {
        "A": 24,
        "B": 28,
        "C": 16,
        "D": 18,
        "E": 20,
        "F": 18,
        "G": 18,
        "H": 20,
        "I": 18,
        "J": 30,
    }
    _set_column_widths(worksheet, widths)
    _paint_header_groups(worksheet, id_columns=range(1, 8), note_columns=range(10, 11))

    status_validation = DataValidation(
        type="list",
        formula1="=status_colaborador",
        allow_blank=True,
    )
    status_validation.promptTitle = "Status do colaborador"
    status_validation.prompt = "Escolha um status padrao."
    worksheet.add_data_validation(status_validation)
    status_validation.add(f"H2:H{max_data_rows}")

    admite_validation = DataValidation(
        type="list",
        formula1='"sim,nao"',
        allow_blank=True,
    )
    admite_validation.promptTitle = "Admite lancamento"
    admite_validation.prompt = "Use sim ou nao."
    worksheet.add_data_validation(admite_validation)
    admite_validation.add(f"I2:I{max_data_rows}")

    duplicate_rule = FormulaRule(
        formula=['AND($E2<>"",COUNTIF($E:$E,$E2)>1)'],
        stopIfTrue=False,
        fill=WARNING_FILL,
    )
    worksheet.conditional_formatting.add(f"E2:E{max_data_rows}", duplicate_rule)


def _build_lancamentos_faceis_sheet(worksheet, max_data_rows: int) -> None:
    _write_headers(worksheet, LANCAMENTOS_FACEIS_HEADERS)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:Y{max_data_rows}"
    worksheet.sheet_properties.tabColor = "70AD47"

    widths = {
        "A": 16,
        "B": 24,
        "C": 28,
        "D": 20,
        "E": 18,
        "F": 26,
        "G": 16,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 15,
        "L": 15,
        "M": 22,
        "N": 18,
        "O": 15,
        "P": 15,
        "Q": 18,
        "R": 14,
        "S": 16,
        "T": 20,
        "U": 30,
        "V": 18,
        "W": 18,
        "X": 20,
        "Y": 14,
    }
    _set_column_widths(worksheet, widths)
    _paint_lancamentos_header_groups(worksheet)

    worksheet["A1"].comment = Comment("Uso tecnico futuro para status da linha.", "motor")
    worksheet["G1"].comment = Comment("Use HH:MM para horas extras.", "motor")
    worksheet["N1"].comment = Comment("Selecione a situacao do vale transporte.", "motor")
    worksheet["S1"].comment = Comment("Use HH:MM para atrasos.", "motor")
    worksheet["U1"].comment = Comment("Descreva apenas o que precisa de revisao humana.", "motor")
    worksheet["V1"].comment = Comment("Use HH:MM para horas extras 70%.", "motor")
    worksheet["W1"].comment = Comment("Use HH:MM para horas extras 100%.", "motor")
    worksheet["X1"].comment = Comment("Use HH:MM para hora extra noturna.", "motor")

    _add_hour_validation(worksheet, "G", max_data_rows)
    _add_hour_validation(worksheet, "S", max_data_rows)
    _add_non_negative_decimal_validation(worksheet, MONETARY_COLUMNS, max_data_rows)
    _add_non_negative_decimal_validation(worksheet, NUMBER_COLUMNS, max_data_rows)

    vale_transporte_validation = DataValidation(
        type="list",
        formula1='"sim,nao,parcial,revisar"',
        allow_blank=True,
    )
    vale_transporte_validation.promptTitle = "Vale transporte"
    vale_transporte_validation.prompt = "Use uma das opcoes padrao."
    worksheet.add_data_validation(vale_transporte_validation)
    vale_transporte_validation.add(f"N2:N{max_data_rows}")

    for column_letter in HOUR_COLUMNS:
        for row_index in range(2, max_data_rows + 1):
            worksheet[f"{column_letter}{row_index}"].number_format = "@"

    for column_letter in MONETARY_COLUMNS:
        for row_index in range(2, max_data_rows + 1):
            worksheet[f"{column_letter}{row_index}"].number_format = "#,##0.00"

    for row_index in range(2, max_data_rows + 1):
        worksheet[f"R{row_index}"].number_format = "0.00"


def _build_movimentos_canonicos_sheet(worksheet, max_data_rows: int) -> None:
    _write_headers(worksheet, MOVIMENTOS_CANONICOS_HEADERS)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:W{max_data_rows}"
    worksheet.sheet_properties.tabColor = "A5A5A5"
    _paint_header_groups(worksheet, id_columns=range(1, 9), event_columns=range(9, 17), note_columns=range(17, 24))
    _set_column_widths(
        worksheet,
        {
            "A": 16,
            "B": 16,
            "C": 14,
            "D": 14,
            "E": 16,
            "F": 22,
            "G": 28,
            "H": 20,
            "I": 22,
            "J": 18,
            "K": 16,
            "L": 14,
            "M": 14,
            "N": 14,
            "O": 14,
            "P": 20,
            "Q": 16,
            "R": 16,
            "S": 16,
            "T": 14,
            "U": 18,
            "V": 28,
            "W": 28,
        },
    )


def _build_pendencias_sheet(worksheet, max_data_rows: int) -> None:
    _write_headers(worksheet, PENDENCIAS_HEADERS)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:P{max_data_rows}"
    worksheet.sheet_properties.tabColor = "C55A11"
    _paint_header_groups(worksheet, id_columns=range(1, 10), note_columns=range(10, 17))
    _set_column_widths(
        worksheet,
        {
            "A": 16,
            "B": 14,
            "C": 16,
            "D": 14,
            "E": 22,
            "F": 28,
            "G": 20,
            "H": 22,
            "I": 16,
            "J": 20,
            "K": 36,
            "L": 28,
            "M": 18,
            "N": 32,
            "O": 18,
            "P": 18,
        },
    )

    severidade_validation = DataValidation(
        type="list",
        formula1="=severidade_pendencia",
        allow_blank=True,
    )
    severidade_validation.promptTitle = "Severidade"
    severidade_validation.prompt = "Escolha a severidade da pendencia."
    worksheet.add_data_validation(severidade_validation)
    severidade_validation.add(f"B2:B{max_data_rows}")

    status_validation = DataValidation(
        type="list",
        formula1="=status_tratamento",
        allow_blank=True,
    )
    status_validation.promptTitle = "Status do tratamento"
    status_validation.prompt = "Informe o status atual da pendencia."
    worksheet.add_data_validation(status_validation)
    status_validation.add(f"M2:M{max_data_rows}")

    for row_index in range(2, max_data_rows + 1):
        worksheet[f"P{row_index}"].number_format = "DD/MM/YYYY HH:MM"


def _build_listas_sheet(workbook: Workbook, worksheet) -> None:
    worksheet.sheet_properties.tabColor = "4472C4"
    worksheet.freeze_panes = "A2"

    for column_index, (block_name, values) in enumerate(LIST_BLOCKS.items(), start=1):
        header_cell = worksheet.cell(row=1, column=column_index, value=block_name)
        header_cell.fill = LIST_FILL
        header_cell.font = Font(bold=True)
        header_cell.border = THIN_BORDER
        header_cell.alignment = Alignment(horizontal="center")

        worksheet.column_dimensions[get_column_letter(column_index)].width = max(
            20,
            len(block_name) + 4,
        )

        for row_offset, value in enumerate(values, start=2):
            cell = worksheet.cell(row=row_offset, column=column_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")

        start_row = 2
        end_row = len(values) + 1
        reference = f"'LISTAS'!${get_column_letter(column_index)}${start_row}:${get_column_letter(column_index)}${end_row}"
        workbook.defined_names.add(DefinedName(block_name, attr_text=reference))


def _write_headers(worksheet, headers: tuple[str, ...]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 24


def _set_column_widths(worksheet, widths: dict[str, int]) -> None:
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width


def _paint_header_groups(worksheet, id_columns=(), event_columns=(), note_columns=()) -> None:
    for column_index in id_columns:
        worksheet.cell(row=1, column=column_index).fill = ID_FILL
        worksheet.cell(row=1, column=column_index).font = Font(bold=True, color="1F1F1F")
    for column_index in event_columns:
        worksheet.cell(row=1, column=column_index).fill = EVENT_FILL
        worksheet.cell(row=1, column=column_index).font = Font(bold=True, color="1F1F1F")
    for column_index in note_columns:
        worksheet.cell(row=1, column=column_index).fill = NOTE_FILL
        worksheet.cell(row=1, column=column_index).font = Font(bold=True, color="1F1F1F")


def _paint_lancamentos_header_groups(worksheet) -> None:
    worksheet["A1"].fill = TECH_FILL
    worksheet["A1"].font = Font(bold=True, color="1F1F1F")

    for column_letter in ("B", "C", "D", "E", "F"):
        worksheet[f"{column_letter}1"].fill = ID_FILL
        worksheet[f"{column_letter}1"].font = Font(bold=True, color="1F1F1F")

    for column_letter in ("G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"):
        worksheet[f"{column_letter}1"].fill = EVENT_FILL
        worksheet[f"{column_letter}1"].font = Font(bold=True, color="1F1F1F")

    worksheet["U1"].fill = NOTE_FILL
    worksheet["U1"].font = Font(bold=True, color="1F1F1F")


def _add_non_negative_decimal_validation(
    worksheet,
    columns: tuple[str, ...],
    max_data_rows: int,
) -> None:
    validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    validation.errorTitle = "Valor invalido"
    validation.error = "Informe um numero maior ou igual a zero."
    validation.promptTitle = "Valor numerico"
    validation.prompt = "Use numero maior ou igual a zero."
    worksheet.add_data_validation(validation)

    for column_letter in columns:
        validation.add(f"{column_letter}2:{column_letter}{max_data_rows}")


def _add_hour_validation(worksheet, column_letter: str, max_data_rows: int) -> None:
    formula = (
        f'OR({column_letter}2="",'
        f'AND(LEN({column_letter}2)=5,'
        f'MID({column_letter}2,3,1)=":",'
        f'ISNUMBER(--LEFT({column_letter}2,2)),'
        f'ISNUMBER(--RIGHT({column_letter}2,2)),'
        f'--LEFT({column_letter}2,2)>=0,'
        f'--LEFT({column_letter}2,2)<=99,'
        f'--RIGHT({column_letter}2,2)>=0,'
        f'--RIGHT({column_letter}2,2)<=59))'
    )
    validation = DataValidation(type="custom", formula1=formula, allow_blank=True)
    validation.errorTitle = "Hora invalida"
    validation.error = "Use o formato HH:MM, por exemplo 02:16."
    validation.promptTitle = "Horas"
    validation.prompt = "Digite horas no formato HH:MM."
    worksheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{max_data_rows}")
