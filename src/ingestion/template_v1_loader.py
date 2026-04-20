"""Loader and writer for the V1 human payroll spreadsheet."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain import (
    CanonicalMovement,
    IngestionResult,
    PayrollFileParameters,
    PendingCode,
    PendingItem,
    RegistrationSource,
    ResolvedEmployee,
    SourceRef,
    ValueType,
)

from .errors import NormalizationError, TemplateV1IngestionError
from .normalization import (
    is_empty_value,
    normalize_hours_hhmm,
    normalize_money_brl,
    normalize_quantity,
    normalized_optional_text,
    validate_competence,
)
from .taxonomy import FatalIngestionCode, render_fatal_error_message, render_pending_definition
from .template_v1 import (
    FUNCIONARIOS_HEADERS,
    LANCAMENTOS_FACEIS_HEADERS,
    MOVIMENTOS_CANONICOS_HEADERS,
    PARAMETROS_HEADERS,
    PENDENCIAS_HEADERS,
)

REQUIRED_SHEETS = ("PARAMETROS", "FUNCIONARIOS", "LANCAMENTOS_FACEIS")
TECHNICAL_SHEETS = ("MOVIMENTOS_CANONICOS", "PENDENCIAS")
SUPPORTED_LAYOUT_VERSION = "v1"
SUPPORTED_PAYROLL_TYPE = "mensal"
LANCAMENTOS_SHEET_NAME = "LANCAMENTOS_FACEIS"
FUNCIONARIOS_SHEET_NAME = "FUNCIONARIOS"

REQUIRED_PARAMETERS = (
    "empresa_codigo",
    "empresa_nome",
    "competencia",
    "tipo_folha",
    "processo_padrao",
    "versao_layout",
)


@dataclass(frozen=True, slots=True)
class HumanEventSpec:
    column_name: str
    value_type: ValueType | None
    allows_automatic_movement: bool = True


EVENT_SPECS: tuple[HumanEventSpec, ...] = (
    HumanEventSpec("horas_extras_50", ValueType.HOURS),
    HumanEventSpec("gratificacao", ValueType.MONETARY),
    HumanEventSpec("bonus", ValueType.MONETARY),
    HumanEventSpec("bonus_vendas", ValueType.MONETARY),
    HumanEventSpec("pontualidade", ValueType.MONETARY),
    HumanEventSpec("ajuda_custo", ValueType.MONETARY),
    HumanEventSpec("reembolso_plano_saude", ValueType.MONETARY),
    HumanEventSpec("vale_transporte", None, allows_automatic_movement=False),
    HumanEventSpec("farmacia", ValueType.MONETARY),
    HumanEventSpec("mercadoria", ValueType.MONETARY),
    HumanEventSpec("plano_saude", ValueType.MONETARY),
    HumanEventSpec("faltas_dias", ValueType.DAYS),
    HumanEventSpec("atrasos_horas", ValueType.HOURS),
    HumanEventSpec("desconto_adiantamento", ValueType.MONETARY),
)

AUTOMATED_EVENT_SPECS = tuple(spec for spec in EVENT_SPECS if spec.allows_automatic_movement)
SPECIAL_EVENT_SPECS = tuple(spec for spec in EVENT_SPECS if not spec.allows_automatic_movement)


@dataclass(frozen=True, slots=True)
class _EmployeeRegistryRecord:
    employee_key: str
    employee_name: str | None
    domain_registration: str | None
    status: str | None
    allows_entries: bool | None
    source: SourceRef
    duplicate_registration: bool = False
    duplicate_key: bool = False

    @property
    def is_consistent(self) -> bool:
        return not self.duplicate_key and not self.duplicate_registration


def load_planilha_padrao_folha_v1(path: str | Path) -> IngestionResult:
    """Load a V1 workbook into canonical in-memory models."""

    workbook = load_workbook(path)
    return ingest_template_v1_workbook(workbook)


def ingest_template_v1_workbook(workbook: Workbook) -> IngestionResult:
    """Ingest an already loaded workbook."""

    _require_sheets(workbook, REQUIRED_SHEETS)

    parameters = _load_parameters(workbook["PARAMETROS"])
    employee_registry = _load_employee_registry(workbook[FUNCIONARIOS_SHEET_NAME])
    pendings: list[PendingItem] = []
    movements: list[CanonicalMovement] = []
    movement_counter = 1
    pending_counter = 1

    pending_counter = _append_registry_pendings(
        parameters=parameters,
        employee_registry=employee_registry,
        pending_items=pendings,
        pending_counter=pending_counter,
    )

    human_sheet = workbook[LANCAMENTOS_SHEET_NAME]
    human_headers = _read_header_map(human_sheet, LANCAMENTOS_FACEIS_HEADERS)

    for row_number in range(2, human_sheet.max_row + 1):
        row_values = {
            header: human_sheet.cell(row=row_number, column=column_index).value
            for header, column_index in human_headers.items()
        }

        if _row_is_completely_empty(row_values):
            continue

        employee = _resolve_employee_for_row(
            row_values=row_values,
            row_number=row_number,
            employee_registry=employee_registry,
        )
        row_pending_items: list[PendingItem] = []
        movement_candidates: list[tuple[HumanEventSpec, object, SourceRef]] = []

        event_cells_filled = _row_has_any_event_cell(row_values)
        observacao_eventos = normalized_optional_text(row_values["observacao_eventos"])

        if employee.allows_entries is False and event_cells_filled:
            row_pending_items.append(
                _make_pending(
                    pending_counter,
                    parameters,
                    employee=employee,
                    event_name=None,
                    source=SourceRef(
                        LANCAMENTOS_SHEET_NAME,
                        row_number,
                        f"A{row_number}",
                        "linha_status",
                    ),
                    code=PendingCode.LINE_BLOCKED_BY_STATUS,
                )
            )
            pending_counter += 1
            pendings.extend(row_pending_items)
            continue

        if employee.employee_key is None and event_cells_filled:
            row_pending_items.append(
                _make_pending(
                    pending_counter,
                    parameters,
                    employee=employee,
                    event_name=None,
                    source=SourceRef(
                        LANCAMENTOS_SHEET_NAME,
                        row_number,
                        f"B{row_number}",
                        "chave_colaborador",
                    ),
                    code=PendingCode.EMPLOYEE_NOT_FOUND,
                )
            )
            pending_counter += 1

        if event_cells_filled and employee.resolved_from_registry and not employee.registry_consistent:
            row_pending_items.append(
                _make_pending(
                    pending_counter,
                    parameters,
                    employee=employee,
                    event_name=None,
                    source=SourceRef(
                        LANCAMENTOS_SHEET_NAME,
                        row_number,
                        f"B{row_number}",
                        "chave_colaborador",
                    ),
                    code=PendingCode.EMPLOYEE_REGISTRY_INCONSISTENT,
                )
            )
            pending_counter += 1

        if observacao_eventos:
            row_pending_items.append(
                _make_pending(
                    pending_counter,
                    parameters,
                    employee=employee,
                    event_name=None,
                    source=SourceRef(
                        LANCAMENTOS_SHEET_NAME,
                        row_number,
                        f"U{row_number}",
                        "observacao_eventos",
                    ),
                    code=PendingCode.AMBIGUOUS_EVENT_NOTE,
                )
            )
            pending_counter += 1

        if event_cells_filled and employee.resolved_from_registry:
            if employee.registration_source == RegistrationSource.LINE:
                row_pending_items.append(
                    _make_pending(
                        pending_counter,
                        parameters,
                        employee=employee,
                        event_name=None,
                        source=SourceRef(
                            LANCAMENTOS_SHEET_NAME,
                            row_number,
                            f"D{row_number}",
                            "matricula_dominio",
                        ),
                        code=PendingCode.DOMAIN_REGISTRATION_LINE_ONLY,
                    )
                )
                pending_counter += 1
            elif employee.registration_source == RegistrationSource.CONFLICT:
                row_pending_items.append(
                    _make_pending(
                        pending_counter,
                        parameters,
                        employee=employee,
                        event_name=None,
                        source=SourceRef(
                            LANCAMENTOS_SHEET_NAME,
                            row_number,
                            f"D{row_number}",
                            "matricula_dominio",
                        ),
                        code=PendingCode.DOMAIN_REGISTRATION_CONFLICT,
                    )
                )
                pending_counter += 1

        if (
            event_cells_filled
            and employee.domain_registration is None
            and employee.registration_source != RegistrationSource.CONFLICT
        ):
            row_pending_items.append(
                _make_pending(
                    pending_counter,
                    parameters,
                    employee=employee,
                    event_name=None,
                    source=SourceRef(
                        LANCAMENTOS_SHEET_NAME,
                        row_number,
                        f"D{row_number}",
                        "matricula_dominio",
                    ),
                    code=PendingCode.DOMAIN_REGISTRATION_MISSING,
                )
            )
            pending_counter += 1

        for spec in SPECIAL_EVENT_SPECS:
            raw_value = row_values[spec.column_name]
            if is_empty_value(raw_value):
                continue

            source = SourceRef(
                LANCAMENTOS_SHEET_NAME,
                row_number,
                _column_letter_from_header(human_headers, spec.column_name) + str(row_number),
                spec.column_name,
            )
            row_pending_items.append(
                _make_pending(
                    pending_counter,
                    parameters,
                    employee=employee,
                    event_name=spec.column_name,
                    source=source,
                    code=PendingCode.NON_AUTOMATABLE_EVENT,
                )
            )
            pending_counter += 1

        for spec in AUTOMATED_EVENT_SPECS:
            raw_value = row_values[spec.column_name]
            if is_empty_value(raw_value):
                continue

            source = SourceRef(
                LANCAMENTOS_SHEET_NAME,
                row_number,
                _column_letter_from_header(human_headers, spec.column_name) + str(row_number),
                spec.column_name,
            )

            try:
                _normalize_event_value(spec, raw_value)
            except NormalizationError as exc:
                row_pending_items.append(
                    _make_pending(
                        pending_counter,
                        parameters,
                        employee=employee,
                        event_name=spec.column_name,
                        source=source,
                        code=exc.code,
                        description_override=str(exc),
                    )
                )
                pending_counter += 1
                continue

            movement_candidates.append((spec, raw_value, source))

        pendings.extend(row_pending_items)
        blocking_row_pendings = tuple(row_pending_items)

        for spec, raw_value, source in movement_candidates:
            movement = _build_movement(
                movement_counter=movement_counter,
                parameters=parameters,
                employee=employee,
                event_spec=spec,
                raw_value=raw_value,
                source=source,
                row_pending_items=blocking_row_pendings,
                row_values=row_values,
            )
            movements.append(movement)
            movement_counter += 1

    return IngestionResult(
        parameters=parameters,
        employees=tuple(_registry_record_to_employee(record) for record in employee_registry.values()),
        movements=tuple(movements),
        pendings=tuple(pendings),
    )


def write_ingestion_result_to_workbook(
    workbook_or_path: Workbook | str | Path,
    result: IngestionResult,
    output_path: str | Path | None = None,
) -> Path | Workbook:
    """Write canonical movements and pendings into the technical tabs."""

    workbook, opened_from_path = _coerce_workbook(workbook_or_path)
    _require_sheets(workbook, TECHNICAL_SHEETS)

    _write_movements_sheet(workbook["MOVIMENTOS_CANONICOS"], result.movements)
    _write_pendings_sheet(workbook["PENDENCIAS"], result.pendings)

    if opened_from_path is None and output_path is None:
        return workbook

    target_path = Path(output_path or opened_from_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_path)
    return target_path


def ingest_and_fill_planilha_padrao_v1(
    input_path: str | Path,
    output_path: str | Path | None = None,
) -> IngestionResult:
    """Ingest the workbook and update its technical tabs."""

    result = load_planilha_padrao_folha_v1(input_path)
    write_ingestion_result_to_workbook(input_path, result, output_path=output_path)
    return result


def _load_parameters(worksheet: Worksheet) -> PayrollFileParameters:
    _read_header_map(worksheet, PARAMETROS_HEADERS)

    field_rows: dict[str, tuple[int, object]] = {}
    for row_number in range(2, worksheet.max_row + 1):
        field_name = normalized_optional_text(worksheet.cell(row=row_number, column=1).value)
        if field_name is None:
            continue
        field_rows[field_name] = (row_number, worksheet.cell(row=row_number, column=2).value)

    missing_parameters = [name for name in REQUIRED_PARAMETERS if name not in field_rows]
    if missing_parameters:
        raise TemplateV1IngestionError(
            FatalIngestionCode.PARAMETER_REQUIRED_MISSING,
            render_fatal_error_message(
                FatalIngestionCode.PARAMETER_REQUIRED_MISSING,
                field_name=", ".join(missing_parameters),
            ),
            source="PARAMETROS",
        )

    values: dict[str, str] = {}
    source_cells: dict[str, str] = {}

    for parameter_name in REQUIRED_PARAMETERS:
        row_number, raw_value = field_rows[parameter_name]
        text_value = normalized_optional_text(raw_value)
        if text_value is None:
            raise TemplateV1IngestionError(
                FatalIngestionCode.PARAMETER_REQUIRED_MISSING,
                render_fatal_error_message(
                    FatalIngestionCode.PARAMETER_REQUIRED_MISSING,
                    field_name=parameter_name,
                ),
                source=f"PARAMETROS!B{row_number}",
            )
        values[parameter_name] = text_value
        source_cells[parameter_name] = f"B{row_number}"

    try:
        competence = validate_competence(values["competencia"])
    except NormalizationError as exc:
        raise TemplateV1IngestionError(
            FatalIngestionCode.COMPETENCE_INVALID,
            render_fatal_error_message(FatalIngestionCode.COMPETENCE_INVALID),
            source=f"PARAMETROS!{source_cells['competencia']}",
        ) from exc

    payroll_type = values["tipo_folha"].lower()
    if payroll_type != SUPPORTED_PAYROLL_TYPE:
        raise TemplateV1IngestionError(
            FatalIngestionCode.PAYROLL_TYPE_NOT_SUPPORTED,
            render_fatal_error_message(
                FatalIngestionCode.PAYROLL_TYPE_NOT_SUPPORTED,
                payroll_type=values["tipo_folha"],
            ),
            source=f"PARAMETROS!{source_cells['tipo_folha']}",
        )

    layout_version = values["versao_layout"].lower()
    if layout_version != SUPPORTED_LAYOUT_VERSION:
        raise TemplateV1IngestionError(
            FatalIngestionCode.LAYOUT_VERSION_NOT_SUPPORTED,
            render_fatal_error_message(
                FatalIngestionCode.LAYOUT_VERSION_NOT_SUPPORTED,
                layout_version=values["versao_layout"],
                supported_layout_version=SUPPORTED_LAYOUT_VERSION,
            ),
            source=f"PARAMETROS!{source_cells['versao_layout']}",
        )

    return PayrollFileParameters(
        company_code=values["empresa_codigo"],
        company_name=values["empresa_nome"],
        competence=competence,
        payroll_type=payroll_type,
        default_process=values["processo_padrao"],
        layout_version=layout_version,
        source_cells=source_cells,
    )


def _load_employee_registry(worksheet: Worksheet) -> dict[str, _EmployeeRegistryRecord]:
    headers = _read_header_map(worksheet, FUNCIONARIOS_HEADERS)
    registry: dict[str, _EmployeeRegistryRecord] = {}
    registrations: dict[str, list[str]] = {}
    duplicate_keys: set[str] = set()

    for row_number in range(2, worksheet.max_row + 1):
        row = {
            header: worksheet.cell(row=row_number, column=column_index).value
            for header, column_index in headers.items()
        }
        if _row_is_completely_empty(row):
            continue

        employee_key = normalized_optional_text(row["chave_colaborador"])
        if employee_key is None:
            continue

        if employee_key in registry:
            duplicate_keys.add(employee_key)
            continue

        registration = normalized_optional_text(row["matricula_dominio"])
        if registration is not None:
            registrations.setdefault(registration, []).append(employee_key)

        registry[employee_key] = _EmployeeRegistryRecord(
            employee_key=employee_key,
            employee_name=normalized_optional_text(row["nome_colaborador"]),
            domain_registration=registration,
            status=normalized_optional_text(row["status_colaborador"]),
            allows_entries=_normalize_allows_entries(row["admite_lancamento"], row["status_colaborador"]),
            source=SourceRef(FUNCIONARIOS_SHEET_NAME, row_number, f"A{row_number}", "chave_colaborador"),
        )

    duplicate_registrations = {
        registration for registration, keys in registrations.items() if len(keys) > 1
    }

    updated_registry: dict[str, _EmployeeRegistryRecord] = {}
    for employee_key, record in registry.items():
        updated_registry[employee_key] = _EmployeeRegistryRecord(
            employee_key=record.employee_key,
            employee_name=record.employee_name,
            domain_registration=record.domain_registration,
            status=record.status,
            allows_entries=record.allows_entries,
            source=record.source,
            duplicate_registration=(
                record.domain_registration in duplicate_registrations
                if record.domain_registration is not None
                else False
            ),
            duplicate_key=employee_key in duplicate_keys,
        )
    return updated_registry


def _resolve_employee_for_row(
    row_values: dict[str, object],
    row_number: int,
    employee_registry: dict[str, _EmployeeRegistryRecord],
) -> ResolvedEmployee:
    employee_key = normalized_optional_text(row_values["chave_colaborador"])
    human_name = normalized_optional_text(row_values["nome_colaborador"])
    human_registration = normalized_optional_text(row_values["matricula_dominio"])

    record = employee_registry.get(employee_key) if employee_key is not None else None
    if record is not None:
        resolved_registration = None
        registration_source = RegistrationSource.UNRESOLVED

        if human_registration and record.domain_registration:
            if human_registration == record.domain_registration:
                resolved_registration = human_registration
                registration_source = RegistrationSource.BOTH_MATCH
            else:
                registration_source = RegistrationSource.CONFLICT
        elif human_registration and not record.domain_registration:
            resolved_registration = human_registration
            registration_source = RegistrationSource.LINE
        elif not human_registration and record.domain_registration and record.is_consistent:
            resolved_registration = record.domain_registration
            registration_source = RegistrationSource.REGISTRY

        return ResolvedEmployee(
            employee_key=record.employee_key,
            employee_name=record.employee_name or human_name,
            domain_registration=resolved_registration,
            status=record.status,
            allows_entries=record.allows_entries,
            source=record.source,
            resolved_from_registry=True,
            registration_source=registration_source,
            registry_consistent=record.is_consistent,
        )

    return ResolvedEmployee(
        employee_key=employee_key,
        employee_name=human_name,
        domain_registration=human_registration,
        status=None,
        allows_entries=None,
        source=SourceRef(LANCAMENTOS_SHEET_NAME, row_number, f"B{row_number}", "chave_colaborador"),
        resolved_from_registry=False,
        registration_source=(
            RegistrationSource.LINE if human_registration is not None else RegistrationSource.UNRESOLVED
        ),
        registry_consistent=True,
    )


def _append_registry_pendings(
    parameters: PayrollFileParameters,
    employee_registry: dict[str, _EmployeeRegistryRecord],
    pending_items: list[PendingItem],
    pending_counter: int,
) -> int:
    for record in employee_registry.values():
        registry_employee = _registry_record_to_employee(record)
        if record.duplicate_registration:
            pending_items.append(
                _make_pending(
                    pending_counter,
                    parameters,
                    employee=registry_employee,
                    event_name=None,
                    source=record.source,
                    code=PendingCode.DOMAIN_REGISTRATION_DUPLICATED,
                )
            )
            pending_counter += 1

        if record.duplicate_key:
            pending_items.append(
                _make_pending(
                    pending_counter,
                    parameters,
                    employee=registry_employee,
                    event_name=None,
                    source=record.source,
                    code=PendingCode.EMPLOYEE_REGISTRY_INCONSISTENT,
                )
            )
            pending_counter += 1

    return pending_counter


def _build_movement(
    movement_counter: int,
    parameters: PayrollFileParameters,
    employee: ResolvedEmployee,
    event_spec: HumanEventSpec,
    raw_value: object,
    source: SourceRef,
    row_pending_items: tuple[PendingItem, ...],
    row_values: dict[str, object],
) -> CanonicalMovement:
    quantity, hours, amount, serialization_unit = _normalize_event_value(event_spec, raw_value)

    observation = _compose_observation(row_values)
    return CanonicalMovement(
        movement_id=f"mov-{movement_counter:05d}",
        company_code=parameters.company_code,
        competence=parameters.competence,
        payroll_type=parameters.payroll_type,
        default_process=parameters.default_process,
        employee_key=employee.employee_key,
        employee_name=employee.employee_name,
        domain_registration=employee.domain_registration,
        event_name=event_spec.column_name,
        value_type=event_spec.value_type,
        quantity=quantity,
        hours=hours,
        amount=amount,
        source=source,
        blocked=bool(row_pending_items),
        pending_codes=tuple(pending.pending_code for pending in row_pending_items),
        pending_messages=tuple(pending.description for pending in row_pending_items),
        observation=observation,
        serialization_unit=serialization_unit,
    )


def _compose_observation(row_values: dict[str, object]) -> str | None:
    parts: list[str] = []
    general = normalized_optional_text(row_values.get("observacao_geral"))
    events = normalized_optional_text(row_values.get("observacao_eventos"))
    if general:
        parts.append(f"observacao_geral={general}")
    if events:
        parts.append(f"observacao_eventos={events}")
    return " | ".join(parts) if parts else None


def _make_pending(
    pending_counter: int,
    parameters: PayrollFileParameters,
    employee: ResolvedEmployee,
    event_name: str | None,
    source: SourceRef,
    code: str,
    description_override: str | None = None,
    recommended_action_override: str | None = None,
) -> PendingItem:
    severity, description, recommended_action = render_pending_definition(
        code,
        event_name=event_name or "",
        field_name=event_name or "",
    )
    return PendingItem(
        pending_id=f"pend-{pending_counter:05d}",
        severity=severity,
        company_code=parameters.company_code,
        competence=parameters.competence,
        employee_key=employee.employee_key,
        employee_name=employee.employee_name,
        domain_registration=employee.domain_registration,
        event_name=event_name,
        source=source,
        pending_code=code,
        description=description_override or description,
        recommended_action=recommended_action_override or recommended_action,
    )


def _normalize_event_value(event_spec: HumanEventSpec, raw_value: object):
    quantity = None
    hours = None
    amount = None
    serialization_unit = None

    if event_spec.value_type == ValueType.MONETARY:
        amount = normalize_money_brl(raw_value)
        serialization_unit = "BRL"
    elif event_spec.value_type == ValueType.HOURS:
        hours = normalize_hours_hhmm(raw_value)
        serialization_unit = "HH:MM"
    elif event_spec.value_type == ValueType.DAYS:
        quantity = normalize_quantity(raw_value)
        serialization_unit = "DIAS"
    else:
        raise AssertionError(f"Unsupported event spec for automatic movement: {event_spec.column_name}")

    return quantity, hours, amount, serialization_unit


def _registry_record_to_employee(record: _EmployeeRegistryRecord) -> ResolvedEmployee:
    return ResolvedEmployee(
        employee_key=record.employee_key,
        employee_name=record.employee_name,
        domain_registration=record.domain_registration if record.is_consistent else None,
        status=record.status,
        allows_entries=record.allows_entries,
        source=record.source,
        resolved_from_registry=True,
        registration_source=(
            RegistrationSource.REGISTRY
            if record.domain_registration is not None and record.is_consistent
            else RegistrationSource.UNRESOLVED
        ),
        registry_consistent=record.is_consistent,
    )


def _normalize_allows_entries(admite_lancamento: object, status_colaborador: object) -> bool | None:
    text = normalized_optional_text(admite_lancamento)
    status_text = normalized_optional_text(status_colaborador)
    if status_text == "ignorar":
        return False
    if text is None:
        return None
    lowered = text.lower()
    if lowered == "sim":
        return True
    if lowered == "nao":
        return False
    return None


def _row_is_completely_empty(row_values: dict[str, object]) -> bool:
    return all(is_empty_value(value) for value in row_values.values())


def _row_has_any_event_cell(row_values: dict[str, object]) -> bool:
    interesting_columns = {spec.column_name for spec in EVENT_SPECS} | {"observacao_eventos"}
    return any(not is_empty_value(row_values[column_name]) for column_name in interesting_columns)


def _require_sheets(workbook: Workbook, required_sheets: tuple[str, ...]) -> None:
    missing = [sheet_name for sheet_name in required_sheets if sheet_name not in workbook.sheetnames]
    if missing:
        raise TemplateV1IngestionError(
            FatalIngestionCode.MISSING_REQUIRED_SHEET,
            render_fatal_error_message(
                FatalIngestionCode.MISSING_REQUIRED_SHEET,
                missing=", ".join(missing),
            ),
        )


def _read_header_map(worksheet: Worksheet, expected_headers: tuple[str, ...]) -> dict[str, int]:
    header_values = {
        normalized_optional_text(worksheet.cell(row=1, column=column_index).value): column_index
        for column_index in range(1, worksheet.max_column + 1)
    }
    missing_headers = [header for header in expected_headers if header not in header_values]
    if missing_headers:
        raise TemplateV1IngestionError(
            FatalIngestionCode.MISSING_REQUIRED_HEADER,
            render_fatal_error_message(
                FatalIngestionCode.MISSING_REQUIRED_HEADER,
                sheet_name=worksheet.title,
                missing=", ".join(missing_headers),
            ),
            source=worksheet.title,
        )

    return {header: header_values[header] for header in expected_headers}


def _coerce_workbook(workbook_or_path: Workbook | str | Path) -> tuple[Workbook, Path | None]:
    if isinstance(workbook_or_path, Workbook):
        return workbook_or_path, None
    input_path = Path(workbook_or_path)
    return load_workbook(input_path), input_path


def _write_movements_sheet(worksheet: Worksheet, movements: tuple[CanonicalMovement, ...]) -> None:
    _read_header_map(worksheet, MOVIMENTOS_CANONICOS_HEADERS)
    _clear_data_rows(worksheet, total_columns=len(MOVIMENTOS_CANONICOS_HEADERS))

    for row_offset, movement in enumerate(movements, start=2):
        row_values = (
            movement.movement_id,
            movement.company_code,
            movement.competence,
            movement.payroll_type,
            movement.default_process,
            movement.employee_key,
            movement.employee_name,
            movement.domain_registration,
            movement.event_name,
            movement.informed_rubric,
            movement.output_rubric,
            movement.event_nature,
            movement.value_type.value,
            movement.quantity_for_sheet(),
            movement.amount_for_sheet(),
            movement.serialization_unit,
            movement.source.sheet_name,
            movement.source.cell,
            movement.source.column_name,
            "sim" if movement.has_pending else "nao",
            movement.pending_codes_for_sheet(),
            movement.pending_messages_for_sheet(),
            movement.observation,
        )
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_offset, column=column_index, value=value)


def _write_pendings_sheet(worksheet: Worksheet, pendings: tuple[PendingItem, ...]) -> None:
    _read_header_map(worksheet, PENDENCIAS_HEADERS)
    _clear_data_rows(worksheet, total_columns=len(PENDENCIAS_HEADERS))

    for row_offset, pending in enumerate(pendings, start=2):
        row_values = (
            pending.pending_id,
            pending.severity.value,
            pending.company_code,
            pending.competence,
            pending.employee_key,
            pending.employee_name,
            pending.domain_registration,
            pending.event_name,
            pending.source.full_cell_ref,
            pending.pending_code,
            pending.description,
            pending.recommended_action,
            pending.treatment_status,
            pending.manual_resolution,
            pending.resolved_by,
            pending.resolved_at,
        )
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_offset, column=column_index, value=value)


def _clear_data_rows(worksheet: Worksheet, total_columns: int) -> None:
    for row_number in range(2, worksheet.max_row + 1):
        for column_index in range(1, total_columns + 1):
            worksheet.cell(row=row_number, column=column_index, value=None)


def _column_letter_from_header(header_map: dict[str, int], header_name: str) -> str:
    column_index = header_map[header_name]
    quotient = column_index
    letters = ""
    while quotient:
        quotient, remainder = divmod(quotient - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
