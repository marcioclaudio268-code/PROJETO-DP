"""Deterministic detection and normalization for uploaded payroll workbooks."""

from __future__ import annotations

import calendar
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config.master_data import CompanyMasterDataStore, normalize_text

from .errors import InputLayoutNormalizationError
from .normalization import is_empty_value, normalize_hours_hhmm, normalize_quantity, normalized_optional_text
from .template_v1 import (
    FUNCIONARIOS_HEADERS,
    LANCAMENTOS_FACEIS_HEADERS,
    MAX_DATA_ROWS,
    MOVIMENTOS_CANONICOS_HEADERS,
    PARAMETROS_HEADERS,
    PENDENCIAS_HEADERS,
    create_planilha_padrao_folha_v1,
)


CANONICAL_LAYOUT_ID = "template_v1_canonico"
MONTHLY_LAYOUT_ID = "resumo_mensal_por_abas"
DEFAULT_CANONICAL_DEFAULT_PROCESS = "11"
DEFAULT_CANONICAL_LAYOUT_VERSION = "v1"
DEFAULT_CANONICAL_PAYROLL_TYPE = "mensal"
DEFAULT_LAYOUT_NOTE_PREFIX = "layout_normalizado"

MONTH_NAME_TO_NUMBER = {
    "jan": 1,
    "janeiro": 1,
    "fev": 2,
    "fevereiro": 2,
    "mar": 3,
    "marco": 3,
    "marcoo": 3,
    "marcoo ": 3,
    "marco ": 3,
    "marco": 3,
    "março": 3,
    "abr": 4,
    "abril": 4,
    "mai": 5,
    "maio": 5,
    "jun": 6,
    "junho": 6,
    "jul": 7,
    "julho": 7,
    "ago": 8,
    "agosto": 8,
    "set": 9,
    "setembro": 9,
    "out": 10,
    "outubro": 10,
    "nov": 11,
    "novembro": 11,
    "dez": 12,
    "dezembro": 12,
}

CANONICAL_REQUIRED_SHEETS = ("PARAMETROS", "FUNCIONARIOS", "LANCAMENTOS_FACEIS")
MONTHLY_HEADER_REQUIREMENTS = {
    1: ("cod",),
    2: ("nome",),
    3: ("h extra 50",),
    4: ("h extra 100",),
    5: ("ad noturno",),
    6: ("gorjeta",),
    7: ("adiantamento",),
    8: ("faltas",),
    9: ("dsr de falta",),
    10: ("atrasos",),
    11: ("consumo",),
    12: ("vale transporte",),
    13: ("plano odont",),
    14: ("observa",),
}

MONTHLY_SUPPORTED_SOURCE_COLUMNS = {
    3: ("horas_extras_50", "horas"),
    8: ("faltas_dias", "quantidade"),
    10: ("atrasos_horas", "horas"),
    12: ("vale_transporte", "special"),
}

MONTHLY_STATUS_TOKENS = {
    "ferias": "ferias",
    "férias": "ferias",
    "afastado": "afastado",
    "licenca": "afastado",
    "licença": "afastado",
    "rescisao": "rescindido",
    "rescisão": "rescindido",
    "ignorar": "ignorar",
}


@dataclass(frozen=True, slots=True)
class InputLayoutDetection:
    layout_id: str
    active_sheet_name: str | None
    selected_sheet_name: str
    selected_sheet_reason: str
    company_code: str
    company_name: str
    competence: str
    source_company_name: str
    source_title_text: str | None
    source_sheet_names: tuple[str, ...]
    rules_applied: tuple[str, ...]
    warnings: tuple[str, ...] = ()

    def as_dict(self) -> dict[str, Any]:
        return {
            "layout_id": self.layout_id,
            "active_sheet_name": self.active_sheet_name,
            "selected_sheet_name": self.selected_sheet_name,
            "selected_sheet_reason": self.selected_sheet_reason,
            "company_code": self.company_code,
            "company_name": self.company_name,
            "competence": self.competence,
            "source_company_name": self.source_company_name,
            "source_title_text": self.source_title_text,
            "source_sheet_names": list(self.source_sheet_names),
            "rules_applied": list(self.rules_applied),
            "warnings": list(self.warnings),
        }


@dataclass(frozen=True, slots=True)
class InputNormalizationResult:
    source_path: Path
    workbook_path: Path
    report_path: Path | None
    layout: InputLayoutDetection
    canonical_rows_written: int
    employee_rows_written: int
    supported_cells_written: int
    unsupported_cells_preserved: int
    orphan_note_rows: int
    manifest: dict[str, Any]

    def as_dict(self) -> dict[str, Any]:
        return {
            "source_path": str(self.source_path),
            "workbook_path": str(self.workbook_path),
            "report_path": str(self.report_path) if self.report_path is not None else None,
            "layout": self.layout.as_dict(),
            "canonical_rows_written": self.canonical_rows_written,
            "employee_rows_written": self.employee_rows_written,
            "supported_cells_written": self.supported_cells_written,
            "unsupported_cells_preserved": self.unsupported_cells_preserved,
            "orphan_note_rows": self.orphan_note_rows,
            "manifest": self.manifest,
        }


def detect_input_layout(
    workbook: Workbook,
    *,
    registry_root: str | Path | None = None,
) -> InputLayoutDetection:
    canonical_detection = _detect_canonical_v1_layout(workbook)
    if canonical_detection is not None:
        return canonical_detection

    monthly_candidate = _detect_monthly_layout(workbook, registry_root=registry_root)
    if monthly_candidate is not None:
        return monthly_candidate

    raise InputLayoutNormalizationError(
        "layout_desconhecido",
        "Nao foi possivel reconhecer com seguranca o layout do workbook enviado.",
        details={"sheetnames": list(workbook.sheetnames)},
    )


def build_canonical_v1_workbook(
    workbook: Workbook,
    detection: InputLayoutDetection,
    *,
    registry_root: str | Path | None = None,
    max_data_rows: int = MAX_DATA_ROWS,
) -> tuple[Workbook, dict[str, Any]]:
    if detection.layout_id == CANONICAL_LAYOUT_ID:
        report = {
            "layout_id": detection.layout_id,
            "selected_sheet_name": detection.selected_sheet_name,
            "selected_sheet_reason": detection.selected_sheet_reason,
            "company_code": detection.company_code,
            "company_name": detection.company_name,
            "competence": detection.competence,
            "rules_applied": list(detection.rules_applied),
            "warnings": list(detection.warnings),
            "source_sheet_names": list(detection.source_sheet_names),
            "status": "no-op",
        }
        return workbook, report

    if detection.layout_id != MONTHLY_LAYOUT_ID:
        raise InputLayoutNormalizationError(
            "layout_nao_suportado",
            f"Layout nao suportado para construcao canonica: {detection.layout_id}.",
            source=detection.selected_sheet_name,
        )

    selected_sheet = workbook[detection.selected_sheet_name]
    normalized_workbook = create_planilha_padrao_folha_v1(max_data_rows=max_data_rows)
    _populate_canonical_parameters(normalized_workbook, detection)

    employee_rows_written = 0
    canonical_rows_written = 0
    supported_cells_written = 0
    unsupported_cells_preserved = 0
    orphan_note_rows = 0
    row_warnings: list[str] = []

    funcionarios = normalized_workbook["FUNCIONARIOS"]
    lancamentos = normalized_workbook["LANCAMENTOS_FACEIS"]

    for source_row in range(6, selected_sheet.max_row + 1):
        row_values = {
            column_index: selected_sheet.cell(row=source_row, column=column_index).value
            for column_index in range(1, selected_sheet.max_column + 1)
        }
        if _row_is_empty(row_values):
            continue

        employee_key = normalized_optional_text(row_values.get(1))
        employee_name = normalized_optional_text(row_values.get(2))
        if employee_key is None and employee_name is None:
            supported_event_present = any(
                not is_empty_value(row_values.get(column_index))
                and _parse_monthly_supported_cell(column_index, row_values.get(column_index))[0] is not None
                for column_index in MONTHLY_SUPPORTED_SOURCE_COLUMNS
            )
            if supported_event_present:
                raise InputLayoutNormalizationError(
                    "linha_orfa_com_dado_critico",
                    "Uma linha da planilha mensal trouxe dados de lancamento sem identidade de colaborador.",
                    source=f"{selected_sheet.title}!A{source_row}:N{source_row}",
                )

            note_only = _build_row_note(selected_sheet.title, source_row, row_values, supported_only=False)
            if note_only is not None:
                orphan_note_rows += 1
                row_warnings.append(note_only)
            continue

        employee_status, employee_note = _infer_employee_status(row_values)
        employee_rows_written += 1
        employee_target_row = _write_employee_row(
            funcionarios,
            employee_rows_written + 1,
            employee_key=employee_key,
            employee_name=employee_name,
            status_colaborador=employee_status,
            note=employee_note,
            source_sheet=selected_sheet.title,
            source_row=source_row,
        )

        row_support_notes: list[str] = []
        launch_row_has_supported_cell = False

        for column_index, (event_name, kind) in MONTHLY_SUPPORTED_SOURCE_COLUMNS.items():
            raw_value = row_values.get(column_index)
            if is_empty_value(raw_value):
                continue

            parsed_value, cell_note = _parse_monthly_supported_cell(column_index, raw_value)
            if parsed_value is None:
                if cell_note is not None:
                    row_support_notes.append(cell_note)
                    unsupported_cells_preserved += 1
                continue

            launch_row_has_supported_cell = True
            if event_name == "horas_extras_50":
                lancamentos.cell(row=canonical_rows_written + 2, column=7, value=parsed_value)
            elif event_name == "faltas_dias":
                lancamentos.cell(row=canonical_rows_written + 2, column=18, value=parsed_value)
            elif event_name == "atrasos_horas":
                lancamentos.cell(row=canonical_rows_written + 2, column=19, value=parsed_value)
            elif event_name == "vale_transporte":
                lancamentos.cell(row=canonical_rows_written + 2, column=14, value=parsed_value)
            supported_cells_written += 1

        row_note = _build_row_note(
            selected_sheet.title,
            source_row,
            row_values,
            supported_only=True,
            extra_notes=row_support_notes,
        )
        if row_note is not None:
            launch_row_has_supported_cell = True
            canonical_rows_written += 1
            launch_target_row = canonical_rows_written + 1
            lancamentos.cell(row=launch_target_row, column=2, value=employee_key)
            lancamentos.cell(row=launch_target_row, column=3, value=employee_name)
            lancamentos.cell(row=launch_target_row, column=4, value=None)
            lancamentos.cell(row=launch_target_row, column=5, value=None)
            lancamentos.cell(row=launch_target_row, column=6, value=row_note)
        elif launch_row_has_supported_cell:
            canonical_rows_written += 1
            launch_target_row = canonical_rows_written + 1
            lancamentos.cell(row=launch_target_row, column=2, value=employee_key)
            lancamentos.cell(row=launch_target_row, column=3, value=employee_name)

        if not launch_row_has_supported_cell and row_note is None:
            # Employee-only row, keep the cadastro populated but do not emit a launch row.
            continue

    report = {
        "layout_id": detection.layout_id,
        "selected_sheet_name": detection.selected_sheet_name,
        "selected_sheet_reason": detection.selected_sheet_reason,
        "company_code": detection.company_code,
        "company_name": detection.company_name,
        "competence": detection.competence,
        "source_company_name": detection.source_company_name,
        "source_title_text": detection.source_title_text,
        "source_sheet_names": list(detection.source_sheet_names),
        "rules_applied": list(detection.rules_applied),
        "warnings": list(detection.warnings),
        "counts": {
            "employee_rows_written": employee_rows_written,
            "canonical_rows_written": canonical_rows_written,
            "supported_cells_written": supported_cells_written,
            "unsupported_cells_preserved": unsupported_cells_preserved,
            "orphan_note_rows": orphan_note_rows,
        },
        "row_warnings": row_warnings[:20],
    }
    return normalized_workbook, report


def normalize_input_workbook(
    input_path: str | Path,
    *,
    output_path: str | Path | None = None,
    report_path: str | Path | None = None,
    registry_root: str | Path | None = None,
) -> InputNormalizationResult:
    source_path = Path(input_path)
    workbook = load_workbook(source_path)
    detection = detect_input_layout(workbook, registry_root=registry_root)

    target_path = Path(output_path) if output_path is not None else source_path
    normalized_workbook, manifest = build_canonical_v1_workbook(
        workbook,
        detection,
        registry_root=registry_root,
    )

    target_path.parent.mkdir(parents=True, exist_ok=True)
    normalized_workbook.save(target_path)

    report_target = Path(report_path) if report_path is not None else None
    if report_target is not None:
        report_target.parent.mkdir(parents=True, exist_ok=True)
        result = InputNormalizationResult(
            source_path=source_path,
            workbook_path=target_path,
            report_path=report_target,
            layout=detection,
            canonical_rows_written=int(manifest.get("counts", {}).get("canonical_rows_written", 0)),
            employee_rows_written=int(manifest.get("counts", {}).get("employee_rows_written", 0)),
            supported_cells_written=int(manifest.get("counts", {}).get("supported_cells_written", 0)),
            unsupported_cells_preserved=int(manifest.get("counts", {}).get("unsupported_cells_preserved", 0)),
            orphan_note_rows=int(manifest.get("counts", {}).get("orphan_note_rows", 0)),
            manifest=manifest,
        )
        report_target.write_text(
            json.dumps(result.as_dict(), ensure_ascii=True, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        return result

    return InputNormalizationResult(
        source_path=source_path,
        workbook_path=target_path,
        report_path=report_target,
        layout=detection,
        canonical_rows_written=int(manifest.get("counts", {}).get("canonical_rows_written", 0)),
        employee_rows_written=int(manifest.get("counts", {}).get("employee_rows_written", 0)),
        supported_cells_written=int(manifest.get("counts", {}).get("supported_cells_written", 0)),
        unsupported_cells_preserved=int(manifest.get("counts", {}).get("unsupported_cells_preserved", 0)),
        orphan_note_rows=int(manifest.get("counts", {}).get("orphan_note_rows", 0)),
        manifest=manifest,
    )


def _detect_canonical_v1_layout(workbook: Workbook) -> InputLayoutDetection | None:
    if any(sheet_name not in workbook.sheetnames for sheet_name in CANONICAL_REQUIRED_SHEETS):
        return None

    header_errors: list[str] = []
    for sheet_name, expected_headers in (
        ("PARAMETROS", PARAMETROS_HEADERS),
        ("FUNCIONARIOS", FUNCIONARIOS_HEADERS),
        ("LANCAMENTOS_FACEIS", LANCAMENTOS_FACEIS_HEADERS),
    ):
        worksheet = workbook[sheet_name]
        actual_headers = tuple(
            normalized_optional_text(worksheet.cell(row=1, column=column_index).value)
            for column_index in range(1, len(expected_headers) + 1)
        )
        expected_normalized = tuple(normalize_text(header) for header in expected_headers)
        if actual_headers != expected_normalized:
            header_errors.append(sheet_name)

    if header_errors:
        raise InputLayoutNormalizationError(
            "layout_canonico_invalido",
            "O workbook contem as abas canonicas, mas o contrato da planilha V1 nao esta consistente.",
            details={"sheetnames": list(workbook.sheetnames), "invalid_sheets": header_errors},
        )

    return InputLayoutDetection(
        layout_id=CANONICAL_LAYOUT_ID,
        active_sheet_name=workbook.active.title if workbook.active is not None else None,
        selected_sheet_name="LANCAMENTOS_FACEIS",
        selected_sheet_reason="canonical_required_sheets_present",
        company_code=_read_parameter_value(workbook["PARAMETROS"], "empresa_codigo"),
        company_name=_read_parameter_value(workbook["PARAMETROS"], "empresa_nome"),
        competence=_read_parameter_value(workbook["PARAMETROS"], "competencia"),
        source_company_name=_read_parameter_value(workbook["PARAMETROS"], "empresa_nome"),
        source_title_text=_read_parameter_value(workbook["PARAMETROS"], "observacoes_gerais", allow_blank=True),
        source_sheet_names=tuple(workbook.sheetnames),
        rules_applied=("required_sheet_contract", "required_header_contract"),
        warnings=(),
    )


def _detect_monthly_layout(
    workbook: Workbook,
    *,
    registry_root: str | Path | None = None,
) -> InputLayoutDetection | None:
    valid_candidates: list[tuple[Worksheet, str, str, tuple[str, ...]]] = []
    invalid_monthly_sheets: list[str] = []
    warnings: list[str] = []

    for worksheet in workbook.worksheets:
        competence = _parse_competence_from_sheet_name(worksheet.title)
        if competence is None:
            continue
        try:
            _validate_monthly_sheet_headers(worksheet)
        except InputLayoutNormalizationError:
            invalid_monthly_sheets.append(worksheet.title)
            continue
        sheet_warnings = _sheet_title_warnings(worksheet, competence)
        warnings.extend(sheet_warnings)
        valid_candidates.append((worksheet, competence, "monthly_sheet_header_contract", tuple(sheet_warnings)))

    if not valid_candidates and invalid_monthly_sheets:
        raise InputLayoutNormalizationError(
            "layout_mensal_invalido",
            "O workbook contem abas mensais, mas elas nao seguem o contrato esperado.",
            details={"invalid_monthly_sheets": invalid_monthly_sheets, "sheetnames": list(workbook.sheetnames)},
        )

    if not valid_candidates:
        return None

    active_title = workbook.active.title if workbook.active is not None else None
    active_candidate = next((item for item in valid_candidates if item[0].title == active_title), None)
    if active_candidate is not None:
        selected_sheet, competence, rule, candidate_warnings = active_candidate
        company_name = normalized_optional_text(selected_sheet["A1"].value)
        company = _resolve_company_from_master(company_name, registry_root=registry_root, source_sheet=selected_sheet.title)
        return InputLayoutDetection(
            layout_id=MONTHLY_LAYOUT_ID,
            active_sheet_name=active_title,
            selected_sheet_name=selected_sheet.title,
            selected_sheet_reason="active_sheet",
            company_code=company["company_code"],
            company_name=company["company_name"],
            competence=competence,
            source_company_name=company_name or "",
            source_title_text=normalized_optional_text(selected_sheet["A2"].value),
            source_sheet_names=tuple(workbook.sheetnames),
            rules_applied=("active_sheet_selected", rule, company["match_rule"], "competence_from_sheet_name"),
            warnings=tuple(_deduplicate_strings((*warnings, *candidate_warnings))),
        )

    if len(valid_candidates) > 1:
        raise InputLayoutNormalizationError(
            "layout_mensal_ambiguous",
            "O workbook contem mais de uma aba mensal valida e a aba ativa nao identifica a competencia com seguranca.",
            details={"valid_monthly_sheets": [item[0].title for item in valid_candidates], "active_sheet": active_title},
        )

    selected_sheet, competence, rule, candidate_warnings = valid_candidates[0]
    company_name = normalized_optional_text(selected_sheet["A1"].value)
    company = _resolve_company_from_master(company_name, registry_root=registry_root, source_sheet=selected_sheet.title)
    return InputLayoutDetection(
        layout_id=MONTHLY_LAYOUT_ID,
        active_sheet_name=active_title,
        selected_sheet_name=selected_sheet.title,
        selected_sheet_reason="unique_monthly_sheet",
        company_code=company["company_code"],
        company_name=company["company_name"],
        competence=competence,
        source_company_name=company_name or "",
        source_title_text=normalized_optional_text(selected_sheet["A2"].value),
        source_sheet_names=tuple(workbook.sheetnames),
        rules_applied=("unique_monthly_sheet_selected", rule, company["match_rule"], "competence_from_sheet_name"),
        warnings=tuple(_deduplicate_strings((*warnings, *candidate_warnings))),
    )


def _validate_monthly_sheet_headers(worksheet: Worksheet) -> None:
    missing: list[str] = []
    for column_index, expected_fragments in MONTHLY_HEADER_REQUIREMENTS.items():
        actual = _normalize_text_for_match(worksheet.cell(row=4, column=column_index).value)
        if not actual:
            missing.append(f"{worksheet.title}!{column_index}")
            continue
        if not any(fragment in actual for fragment in expected_fragments):
            missing.append(f"{worksheet.title}!{column_index}")

    if missing:
        raise InputLayoutNormalizationError(
            "cabecalho_mensal_invalido",
            f"A aba mensal '{worksheet.title}' nao contem o cabecalho esperado no contrato suportado.",
            source=worksheet.title,
            details={"missing_or_invalid_columns": missing},
        )


def _sheet_title_warnings(worksheet: Worksheet, competence: str) -> list[str]:
    warnings: list[str] = []
    title_text = normalized_optional_text(worksheet["A2"].value)
    if title_text is None:
        return warnings

    if _normalize_text_for_match(title_text).find(_competence_to_month_label(competence)) == -1:
        warnings.append(
            f"title_row_mismatch:{worksheet.title}:'{title_text}' ignorado em favor do nome da aba."
        )
    return warnings


def _resolve_company_from_master(
    company_name: str | None,
    *,
    registry_root: str | Path | None,
    source_sheet: str,
) -> dict[str, str]:
    normalized_target = _normalize_text_for_match(company_name)
    if not normalized_target:
        raise InputLayoutNormalizationError(
            "empresa_nao_identificada",
            "Nao foi possivel identificar a empresa no layout mensal.",
            source=source_sheet,
        )

    store = CompanyMasterDataStore(registry_root)
    entries = store.load_registry_entries()

    def _entry_names(entry) -> tuple[str, ...]:
        return tuple(
            name
            for name in (
                _normalize_text_for_match(entry.razao_social),
                _normalize_text_for_match(entry.nome_fantasia),
            )
            if name is not None
        )

    exact_matches = [
        entry
        for entry in entries
        if normalized_target == _normalize_text_for_match(entry.company_code)
        or normalized_target in {
            name
            for name in (
                _normalize_text_for_match(entry.razao_social),
                _normalize_text_for_match(entry.nome_fantasia),
            )
            if name is not None
        }
    ]
    if len(exact_matches) == 1:
        entry = exact_matches[0]
        return {
            "company_code": entry.company_code,
            "company_name": entry.razao_social or entry.nome_fantasia or company_name or entry.company_code,
            "match_rule": "exact_company_code_or_name",
        }

    if len(exact_matches) > 1:
        raise InputLayoutNormalizationError(
            "empresa_ambigua",
            "A empresa do layout mensal corresponde a mais de uma entrada do cadastro mestre.",
            source=source_sheet,
            details={
                "company_name": company_name,
                "candidate_company_codes": [entry.company_code for entry in exact_matches],
            },
        )

    substring_matches = []
    for entry in entries:
        entry_names = _entry_names(entry)
        if any(normalized_target in name or name in normalized_target for name in entry_names):
            substring_matches.append(entry)

    if len(substring_matches) == 1:
        entry = substring_matches[0]
        return {
            "company_code": entry.company_code,
            "company_name": entry.razao_social or entry.nome_fantasia or company_name or entry.company_code,
            "match_rule": "unique_substring_name_match",
        }

    if len(substring_matches) > 1:
        raise InputLayoutNormalizationError(
            "empresa_ambigua",
            "A empresa do layout mensal corresponde a mais de uma entrada do cadastro mestre.",
            source=source_sheet,
            details={
                "company_name": company_name,
                "candidate_company_codes": [entry.company_code for entry in substring_matches],
            },
        )

    raise InputLayoutNormalizationError(
        "empresa_nao_identificada",
        "Nao foi possivel identificar a empresa do layout mensal no cadastro mestre.",
        source=source_sheet,
        details={"company_name": company_name},
    )


def _populate_canonical_parameters(workbook: Workbook, detection: InputLayoutDetection) -> None:
    worksheet = workbook["PARAMETROS"]
    _set_parameter_value(worksheet, "empresa_codigo", detection.company_code)
    _set_parameter_value(worksheet, "empresa_nome", detection.company_name)
    _set_parameter_value(worksheet, "competencia", detection.competence)
    _set_parameter_value(worksheet, "tipo_folha", DEFAULT_CANONICAL_PAYROLL_TYPE)
    _set_parameter_value(worksheet, "processo_padrao", DEFAULT_CANONICAL_DEFAULT_PROCESS)
    _set_parameter_value(worksheet, "versao_layout", DEFAULT_CANONICAL_LAYOUT_VERSION)
    _set_parameter_value(worksheet, "responsavel_preenchimento", DEFAULT_LAYOUT_NOTE_PREFIX)
    _set_parameter_value(worksheet, "data_referencia", _competence_to_reference_date(detection.competence))
    _set_parameter_value(
        worksheet,
        "observacoes_gerais",
        f"normalizado_de={detection.layout_id}; aba={detection.selected_sheet_name}; origem={detection.source_company_name}",
    )


def _set_parameter_value(worksheet: Worksheet, field_name: str, value: object) -> None:
    for row_number in range(2, worksheet.max_row + 1):
        current = normalized_optional_text(worksheet.cell(row=row_number, column=1).value)
        if current == field_name:
            worksheet.cell(row=row_number, column=2, value=value)
            return
    raise InputLayoutNormalizationError(
        "parametro_interno_ausente",
        f"Nao foi possivel localizar o parametro interno '{field_name}' no workbook canonico.",
        source=worksheet.title,
    )


def _write_employee_row(
    worksheet: Worksheet,
    row_number: int,
    *,
    employee_key: str | None,
    employee_name: str | None,
    status_colaborador: str,
    note: str | None,
    source_sheet: str,
    source_row: int,
) -> int:
    values = {
        "A": employee_key,
        "B": employee_name,
        "H": status_colaborador,
        "I": "sim",
        "J": note or f"layout={MONTHLY_LAYOUT_ID}; aba={source_sheet}; linha_origem={source_row}",
    }
    for column_letter, value in values.items():
        worksheet[f"{column_letter}{row_number}"] = value
    return row_number


def _infer_employee_status(row_values: dict[int, object]) -> tuple[str, str | None]:
    tokens: list[str] = []
    for column_index, value in row_values.items():
        if column_index in {1, 2}:
            continue
        text = normalized_optional_text(value)
        if text is None:
            continue
        lowered = _normalize_text_for_match(text)
        for token, status in MONTHLY_STATUS_TOKENS.items():
            if token in lowered:
                tokens.append(status)
    if "afastado" in tokens:
        return "afastado", "status_origem=licenca"
    if "ferias" in tokens:
        return "ferias", "status_origem=ferias"
    if "rescindido" in tokens:
        return "rescindido", "status_origem=rescisao"
    if "ignorar" in tokens:
        return "ignorar", "status_origem=ignorar"
    return "ativo", None


def _parse_monthly_supported_cell(column_index: int, raw_value: object) -> tuple[str | None, str | None]:
    event_name, kind = MONTHLY_SUPPORTED_SOURCE_COLUMNS[column_index]
    if kind == "special":
        text = normalized_optional_text(raw_value)
        if text is None:
            return None, None
        return text, None

    if kind == "horas":
        parsed_hours = _parse_monthly_hours(raw_value)
        if parsed_hours is not None:
            return parsed_hours, None
        return None, f"{event_name}={normalized_optional_text(raw_value)}"

    if kind == "quantidade":
        parsed_quantity = _parse_monthly_quantity(raw_value)
        if parsed_quantity is not None:
            return parsed_quantity, None
        return None, f"{event_name}={normalized_optional_text(raw_value)}"

    return None, f"{event_name}={normalized_optional_text(raw_value)}"


def _parse_monthly_hours(value: object) -> str | None:
    if is_empty_value(value):
        return None
    if isinstance(value, (datetime, date)):
        return normalize_hours_hhmm(value).text

    if isinstance(value, (int, float)) and float(value).is_integer():
        return f"{int(value):02d}:00"

    text = normalized_optional_text(value)
    if text is None:
        return None

    normalized = _normalize_text_for_match(text)
    match = re.fullmatch(r"(?P<hours>\d{1,2})\s*(?:hora|horas|h)", normalized)
    if match:
        return f"{int(match.group('hours')):02d}:00"
    if re.fullmatch(r"\d{2}:\d{2}", normalized):
        return normalize_hours_hhmm(normalized).text
    return None


def _parse_monthly_quantity(value: object) -> str | None:
    if is_empty_value(value):
        return None
    try:
        return normalize_quantity(value).normalize().to_eng_string()
    except Exception:
        return None


def _build_row_note(
    sheet_name: str,
    source_row: int,
    row_values: dict[int, object],
    *,
    supported_only: bool,
    extra_notes: list[str] | None = None,
) -> str | None:
    notes: list[str] = []
    if not supported_only:
        notes.append(f"layout={MONTHLY_LAYOUT_ID}")
    notes.append(f"aba={sheet_name}")
    notes.append(f"linha_origem={source_row}")

    for column_index, value in row_values.items():
        if column_index in {1, 2}:
            continue
        if column_index in MONTHLY_SUPPORTED_SOURCE_COLUMNS:
            event_name, kind = MONTHLY_SUPPORTED_SOURCE_COLUMNS[column_index]
            parsed_value, cell_note = _parse_monthly_supported_cell(column_index, value)
            if parsed_value is None and cell_note is not None:
                notes.append(cell_note)
            continue
        text = normalized_optional_text(value)
        if text is not None:
            header = _monthly_header_for_column(column_index)
            notes.append(f"{header}={text}")

    if extra_notes:
        notes.extend(extra_notes)

    return "; ".join(note for note in notes if note)


def _monthly_header_for_column(column_index: int) -> str:
    return {
        1: "COD.",
        2: "NOME",
        3: "H. EXTRA 50% COD.150",
        4: "H. EXTRA 100% COD.200 FERIADO",
        5: "AD. NOTURNO COD.25",
        6: "GORJETA COD.237",
        7: "ADIANTAMENTO COD.981",
        8: "FALTAS COD.8792",
        9: "DSR DE FALTA COD.8794",
        10: "ATRASOS COD.8069",
        11: "CONSUMO COD.266",
        12: "VALE TRANSPORTE COD.48",
        13: "PLANO ODONT.",
        14: "OBSERVACOES",
    }.get(column_index, f"COLUNA_{column_index}")


def _read_parameter_value(worksheet: Worksheet, field_name: str, *, allow_blank: bool = False) -> str | None:
    for row_number in range(2, worksheet.max_row + 1):
        current = normalized_optional_text(worksheet.cell(row=row_number, column=1).value)
        if current == field_name:
            value = normalized_optional_text(worksheet.cell(row=row_number, column=2).value)
            if value is None:
                if allow_blank:
                    return None
                raise InputLayoutNormalizationError(
                    "parametro_interno_ausente",
                    f"O parametro '{field_name}' esta ausente no workbook canonico.",
                    source=worksheet.title,
                )
            return value
    raise InputLayoutNormalizationError(
        "parametro_interno_ausente",
        f"Nao foi possivel localizar o parametro '{field_name}' no workbook canonico.",
        source=worksheet.title,
    )


def _competence_to_reference_date(competence: str) -> str:
    month_text, year_text = competence.split("/")
    month = int(month_text)
    year = int(year_text)
    last_day = calendar.monthrange(year, month)[1]
    return f"{last_day:02d}/{month:02d}/{year:04d}"


def _parse_competence_from_sheet_name(sheet_name: str) -> str | None:
    normalized = _normalize_text_for_match(sheet_name)
    match = re.fullmatch(r"(?P<month>[a-z]+)\s+(?P<year>\d{2}|\d{4})", normalized)
    if not match:
        return None

    month = MONTH_NAME_TO_NUMBER.get(match.group("month"))
    if month is None:
        return None

    year_token = match.group("year")
    year = 2000 + int(year_token) if len(year_token) == 2 else int(year_token)
    return f"{month:02d}/{year:04d}"


def _competence_to_month_label(competence: str) -> str:
    month_text, year_text = competence.split("/")
    month = int(month_text)
    year = int(year_text)
    month_name = {
        1: "janeiro",
        2: "fevereiro",
        3: "marco",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro",
    }[month]
    return f"{month_name} {str(year)[2:]}"


def _normalize_text_for_match(value: object) -> str:
    text = normalized_optional_text(value) or ""
    normalized = unicodedata.normalize("NFKD", text)
    without_accents = "".join(char for char in normalized if not unicodedata.combining(char))
    lowered = without_accents.lower()
    cleaned = re.sub(r"[^a-z0-9]+", " ", lowered)
    return " ".join(cleaned.split())


def _row_is_empty(row_values: dict[int, object]) -> bool:
    return all(is_empty_value(value) for value in row_values.values())


def _detect_monthly_headers_text(worksheet: Worksheet) -> list[str]:
    return [
        _normalize_text_for_match(worksheet.cell(row=4, column=column_index).value)
        for column_index in range(1, 15)
    ]


def _deduplicate_strings(values: tuple[str, ...]) -> tuple[str, ...]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return tuple(result)
