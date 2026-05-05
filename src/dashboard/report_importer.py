"""Assisted import of payroll reports into company setup registries.

This module is intentionally deterministic. It only extracts values that are
explicitly present in the uploaded report and only persists selected
suggestions through explicit apply functions.
"""

from __future__ import annotations

import csv
import importlib
import io
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config.master_data import normalize_competence, normalize_text

from .company_employee_registry import (
    CompanyEmployeeRecord,
    EmployeeRegistryStatus,
    find_employee_by_domain_registration,
    load_company_employee_registry,
    save_company_employee_registry,
    upsert_employee_record,
)
from .company_rubric_catalog import (
    CompanyRubricRecord,
    RubricNature,
    RubricStatus,
    RubricValueKind,
    find_rubric_by_code,
    load_company_rubric_catalog,
    save_company_rubric_catalog,
    upsert_rubric_record,
)


EMPLOYEE_REVIEW_NEW = "novo"
EMPLOYEE_REVIEW_EXISTING = "existente"
EMPLOYEE_REVIEW_DIVERGENT = "divergente"
RUBRIC_REVIEW_NEW = "novo"
RUBRIC_REVIEW_EXISTING = "existente"
RUBRIC_REVIEW_DIVERGENT = "divergente"
RUBRIC_REVIEW_INCOMPLETE = "incompleto"
REPORT_IMPORTER_SOURCE = "dashboard_report_importer"

_EMPLOYEE_CODE_HEADERS = {
    "matricula",
    "matricula_dominio",
    "codigo_funcionario",
    "cod_funcionario",
    "codigo_colaborador",
    "cod_colaborador",
    "codigo_empregado",
    "cod_empregado",
    "registro_funcionario",
}
_EMPLOYEE_NAME_HEADERS = {
    "nome_funcionario",
    "funcionario",
    "nome_colaborador",
    "colaborador",
    "nome_empregado",
    "empregado",
}
_RUBRIC_CODE_HEADERS = {
    "rubrica",
    "codigo_rubrica",
    "cod_rubrica",
    "codigo_evento",
    "cod_evento",
    "evento",
}
_RUBRIC_DESCRIPTION_HEADERS = {
    "descricao_rubrica",
    "descricao",
    "nome_rubrica",
    "descricao_evento",
    "nome_evento",
    "evento_descricao",
}
_CANONICAL_EVENT_HEADERS = {
    "evento_canonico",
    "evento_negocio",
    "canonical_event",
}
_VALUE_KIND_HEADERS = {
    "tipo_valor",
    "tipo_do_valor",
    "value_kind",
}
_NATURE_HEADERS = {"natureza"}
_RUBRIC_TOTAL_HEADERS = {
    "total",
    "total_rubrica",
    "valor_total",
    "total_valor",
}
_COMPANY_CODE_HEADERS = {
    "empresa",
    "codigo_empresa",
    "cod_empresa",
}
_COMPANY_NAME_HEADERS = {
    "nome_empresa",
    "razao_social",
    "empresa_nome",
}
_COMPETENCE_HEADERS = {
    "competencia",
    "mes_ano",
    "referencia",
}


@dataclass(frozen=True, slots=True)
class ReportSuggestionOrigin:
    file_name: str
    competence: str | None
    detected_company_code: str | None
    detected_company_name: str | None
    parsed_at: datetime
    data_type: str


@dataclass(frozen=True, slots=True)
class ReportEmployeeSuggestion:
    domain_registration: str
    employee_name: str
    source_reference: str
    origin: ReportSuggestionOrigin


@dataclass(frozen=True, slots=True)
class ReportRubricSuggestion:
    rubric_code: str
    description: str
    source_reference: str
    origin: ReportSuggestionOrigin
    canonical_event: str | None = None
    value_kind: str | None = None
    nature: str | None = None


@dataclass(frozen=True, slots=True)
class ReportRubricTotal:
    rubric_code: str
    description: str | None
    total_value: str
    source_reference: str
    origin: ReportSuggestionOrigin


@dataclass(frozen=True, slots=True)
class ReportColumnProfileSuggestion:
    column_name: str
    generation_mode: str
    rubrica_target: str | None
    value_kind: str | None
    source_reference: str
    origin: ReportSuggestionOrigin
    notes: str | None = None


@dataclass(frozen=True, slots=True)
class ParsedPayrollReport:
    file_name: str
    parsed_at: datetime
    detected_company_code: str | None
    detected_company_name: str | None
    competence: str | None
    employees: tuple[ReportEmployeeSuggestion, ...]
    rubrics: tuple[ReportRubricSuggestion, ...]
    rubric_totals: tuple[ReportRubricTotal, ...]
    column_profiles: tuple[ReportColumnProfileSuggestion, ...]
    text_line_count: int


@dataclass(frozen=True, slots=True)
class ReportEmployeeReview:
    suggestion: ReportEmployeeSuggestion
    status: str
    message: str
    current_record: CompanyEmployeeRecord | None = None

    @property
    def can_apply(self) -> bool:
        return self.status in {EMPLOYEE_REVIEW_NEW, EMPLOYEE_REVIEW_DIVERGENT}


@dataclass(frozen=True, slots=True)
class ReportRubricReview:
    suggestion: ReportRubricSuggestion
    status: str
    message: str
    current_record: CompanyRubricRecord | None = None

    @property
    def can_apply(self) -> bool:
        has_required_fields = bool(self.resolved_canonical_event and self.resolved_value_kind)
        return self.status in {
            RUBRIC_REVIEW_NEW,
            RUBRIC_REVIEW_DIVERGENT,
            RUBRIC_REVIEW_EXISTING,
        } and has_required_fields

    @property
    def resolved_canonical_event(self) -> str | None:
        return self.suggestion.canonical_event or (
            self.current_record.canonical_event if self.current_record is not None else None
        )

    @property
    def resolved_value_kind(self) -> str | None:
        return self.suggestion.value_kind or (
            self.current_record.value_kind.value if self.current_record is not None else None
        )

    @property
    def resolved_nature(self) -> str:
        if self.suggestion.nature:
            return self.suggestion.nature
        if self.current_record is not None:
            return self.current_record.nature.value
        return RubricNature.UNKNOWN.value


@dataclass(frozen=True, slots=True)
class ReportImportAnalysis:
    selected_company_code: str | None
    selected_company_name: str | None
    report: ParsedPayrollReport
    employee_reviews: tuple[ReportEmployeeReview, ...]
    rubric_reviews: tuple[ReportRubricReview, ...]
    blocked_reason: str | None = None

    @property
    def is_blocked(self) -> bool:
        return self.blocked_reason is not None


@dataclass(frozen=True, slots=True)
class ReportApplyResult:
    applied: int
    skipped: int
    errors: tuple[str, ...] = ()
    target_path: str | None = None


def analyze_report_import(
    *,
    file_name: str,
    file_bytes: bytes,
    selected_company_code: str | None,
    selected_company_name: str | None = None,
    employee_registry_root: str | Path | None = None,
    rubric_catalog_root: str | Path | None = None,
) -> ReportImportAnalysis:
    """Parse an uploaded report and compare suggestions with current registries."""

    parsed = parse_report_file(file_name=file_name, file_bytes=file_bytes)
    normalized_selected_code = _clean_text(selected_company_code)
    blocked_reason = _report_block_reason(
        parsed,
        selected_company_code=normalized_selected_code,
    )
    employee_reviews = _build_employee_reviews(
        parsed,
        selected_company_code=normalized_selected_code,
        selected_company_name=selected_company_name,
        registry_root=employee_registry_root,
    )
    rubric_reviews = _build_rubric_reviews(
        parsed,
        selected_company_code=normalized_selected_code,
        selected_company_name=selected_company_name,
        catalog_root=rubric_catalog_root,
    )
    return ReportImportAnalysis(
        selected_company_code=normalized_selected_code,
        selected_company_name=_clean_text(selected_company_name),
        report=parsed,
        employee_reviews=employee_reviews,
        rubric_reviews=rubric_reviews,
        blocked_reason=blocked_reason,
    )


def parse_report_file(*, file_name: str, file_bytes: bytes) -> ParsedPayrollReport:
    parsed_at = datetime.now(timezone.utc)
    extension = Path(file_name).suffix.lower()
    rows, lines = _extract_rows_and_lines(file_bytes, extension=extension)
    detected_company_code, detected_company_name = _extract_company_metadata(rows, lines)
    competence = _extract_competence(rows, lines)
    base_origin = ReportSuggestionOrigin(
        file_name=file_name,
        competence=competence,
        detected_company_code=detected_company_code,
        detected_company_name=detected_company_name,
        parsed_at=parsed_at,
        data_type="relatorio",
    )
    employees = _dedupe_employees(
        (
            *_extract_employees_from_rows(rows, base_origin),
            *_extract_employees_from_text(lines, base_origin),
        )
    )
    rubrics = _dedupe_rubrics(
        (
            *_extract_rubrics_from_rows(rows, base_origin),
            *_extract_rubrics_from_text(lines, base_origin),
        )
    )
    rubric_totals = _dedupe_rubric_totals(
        (
            *_extract_rubric_totals_from_rows(rows, base_origin),
            *_extract_rubric_totals_from_text(lines, base_origin),
        )
    )
    column_profiles = _dedupe_column_profiles(
        _extract_column_profiles_from_rows(rows, base_origin)
    )
    return ParsedPayrollReport(
        file_name=file_name,
        parsed_at=parsed_at,
        detected_company_code=detected_company_code,
        detected_company_name=detected_company_name,
        competence=competence,
        employees=employees,
        rubrics=rubrics,
        rubric_totals=rubric_totals,
        column_profiles=column_profiles,
        text_line_count=len(lines),
    )


def apply_report_employee_suggestions(
    analysis: ReportImportAnalysis,
    *,
    selected_domain_registrations: tuple[str, ...] | list[str] | set[str],
    root: str | Path | None = None,
) -> ReportApplyResult:
    if analysis.is_blocked:
        return ReportApplyResult(
            applied=0,
            skipped=0,
            errors=(analysis.blocked_reason or "Importacao assistida bloqueada.",),
        )
    if not analysis.selected_company_code:
        return ReportApplyResult(
            applied=0,
            skipped=0,
            errors=("Selecione a empresa antes de aplicar funcionarios.",),
        )

    selected = {_normalize_lookup_token(item) for item in selected_domain_registrations}
    registry = load_company_employee_registry(
        analysis.selected_company_code,
        company_name=analysis.selected_company_name,
        root=root,
    )
    applied = 0
    skipped = 0
    errors: list[str] = []
    for review in analysis.employee_reviews:
        suggestion = review.suggestion
        if _normalize_lookup_token(suggestion.domain_registration) not in selected:
            continue
        if not review.can_apply:
            skipped += 1
            continue

        current = review.current_record
        try:
            updated = CompanyEmployeeRecord(
                employee_key=current.employee_key if current is not None else None,
                employee_name=suggestion.employee_name,
                domain_registration=suggestion.domain_registration,
                aliases=list(current.aliases) if current is not None else [],
                status=EmployeeRegistryStatus.ACTIVE,
                source=REPORT_IMPORTER_SOURCE,
                notes=_origin_note(suggestion.origin),
            )
            registry = upsert_employee_record(registry, updated)
            applied += 1
        except Exception as exc:  # pragma: no cover - pydantic guard
            errors.append(f"{suggestion.domain_registration}: {exc}")

    target_path = None
    if applied:
        target_path = str(save_company_employee_registry(registry, root=root))
    return ReportApplyResult(
        applied=applied,
        skipped=skipped,
        errors=tuple(errors),
        target_path=target_path,
    )


def apply_report_rubric_suggestions(
    analysis: ReportImportAnalysis,
    *,
    selected_rubric_codes: tuple[str, ...] | list[str] | set[str],
    review_overrides: dict[str, dict[str, str | None]] | None = None,
    root: str | Path | None = None,
) -> ReportApplyResult:
    if analysis.is_blocked:
        return ReportApplyResult(
            applied=0,
            skipped=0,
            errors=(analysis.blocked_reason or "Importacao assistida bloqueada.",),
        )
    if not analysis.selected_company_code:
        return ReportApplyResult(
            applied=0,
            skipped=0,
            errors=("Selecione a empresa antes de aplicar rubricas.",),
        )

    selected = {_normalize_lookup_token(item) for item in selected_rubric_codes}
    overrides = review_overrides or {}
    catalog = load_company_rubric_catalog(
        analysis.selected_company_code,
        company_name=analysis.selected_company_name,
        root=root,
    )
    applied = 0
    skipped = 0
    errors: list[str] = []
    for review in analysis.rubric_reviews:
        suggestion = review.suggestion
        if _normalize_lookup_token(suggestion.rubric_code) not in selected:
            continue

        override = overrides.get(suggestion.rubric_code, {})
        current = review.current_record
        canonical_event = _clean_text(override.get("canonical_event")) or review.resolved_canonical_event
        value_kind = _normalize_value_kind(override.get("value_kind")) or review.resolved_value_kind
        nature = _normalize_nature(override.get("nature")) or review.resolved_nature
        if not canonical_event or not value_kind:
            skipped += 1
            errors.append(
                f"{suggestion.rubric_code}: informe evento canonico e tipo do valor antes de aplicar."
            )
            continue

        try:
            updated = CompanyRubricRecord(
                rubric_code=suggestion.rubric_code,
                description=suggestion.description,
                canonical_event=canonical_event,
                value_kind=RubricValueKind(value_kind),
                nature=RubricNature(nature),
                aliases=list(current.aliases) if current is not None else [],
                status=RubricStatus.ACTIVE,
                source=REPORT_IMPORTER_SOURCE,
                notes=_origin_note(suggestion.origin),
            )
            catalog = upsert_rubric_record(catalog, updated)
            applied += 1
        except Exception as exc:  # pragma: no cover - enum/pydantic guard
            errors.append(f"{suggestion.rubric_code}: {exc}")

    target_path = None
    if applied:
        target_path = str(save_company_rubric_catalog(catalog, root=root))
    return ReportApplyResult(
        applied=applied,
        skipped=skipped,
        errors=tuple(errors),
        target_path=target_path,
    )


def _build_employee_reviews(
    parsed: ParsedPayrollReport,
    *,
    selected_company_code: str | None,
    selected_company_name: str | None,
    registry_root: str | Path | None,
) -> tuple[ReportEmployeeReview, ...]:
    if not selected_company_code:
        return tuple(
            ReportEmployeeReview(
                suggestion=suggestion,
                status=EMPLOYEE_REVIEW_NEW,
                message="Empresa nao selecionada; sugestao ainda nao comparada.",
            )
            for suggestion in parsed.employees
        )

    registry = load_company_employee_registry(
        selected_company_code,
        company_name=selected_company_name,
        root=registry_root,
    )
    reviews: list[ReportEmployeeReview] = []
    for suggestion in parsed.employees:
        matches = find_employee_by_domain_registration(
            registry,
            suggestion.domain_registration,
            active_only=True,
        )
        if not matches:
            reviews.append(
                ReportEmployeeReview(
                    suggestion=suggestion,
                    status=EMPLOYEE_REVIEW_NEW,
                    message="Funcionario novo no cadastro ativo da empresa.",
                )
            )
            continue

        current = matches[0]
        if _employee_name_matches(current, suggestion.employee_name):
            reviews.append(
                ReportEmployeeReview(
                    suggestion=suggestion,
                    status=EMPLOYEE_REVIEW_EXISTING,
                    message="Funcionario ja existe no cadastro ativo.",
                    current_record=current,
                )
            )
            continue

        reviews.append(
            ReportEmployeeReview(
                suggestion=suggestion,
                status=EMPLOYEE_REVIEW_DIVERGENT,
                message=(
                    "Matricula ja existe, mas o nome do relatorio diverge do cadastro atual."
                ),
                current_record=current,
            )
        )
    return tuple(reviews)


def _build_rubric_reviews(
    parsed: ParsedPayrollReport,
    *,
    selected_company_code: str | None,
    selected_company_name: str | None,
    catalog_root: str | Path | None,
) -> tuple[ReportRubricReview, ...]:
    if not selected_company_code:
        return tuple(
            ReportRubricReview(
                suggestion=suggestion,
                status=RUBRIC_REVIEW_NEW,
                message="Empresa nao selecionada; sugestao ainda nao comparada.",
            )
            for suggestion in parsed.rubrics
        )

    catalog = load_company_rubric_catalog(
        selected_company_code,
        company_name=selected_company_name,
        root=catalog_root,
    )
    reviews: list[ReportRubricReview] = []
    for suggestion in parsed.rubrics:
        matches = find_rubric_by_code(catalog, suggestion.rubric_code, active_only=True)
        if not matches:
            status = (
                RUBRIC_REVIEW_NEW
                if suggestion.canonical_event and suggestion.value_kind
                else RUBRIC_REVIEW_INCOMPLETE
            )
            message = (
                "Rubrica nova com dados suficientes para revisao e aplicacao."
                if status == RUBRIC_REVIEW_NEW
                else (
                    "Rubrica nova sem evento canonico ou tipo de valor no relatorio; "
                    "complete esses campos antes de aplicar."
                )
            )
            reviews.append(
                ReportRubricReview(
                    suggestion=suggestion,
                    status=status,
                    message=message,
                )
            )
            continue

        current = matches[0]
        if _same_lookup(current.description, suggestion.description):
            reviews.append(
                ReportRubricReview(
                    suggestion=suggestion,
                    status=RUBRIC_REVIEW_EXISTING,
                    message="Rubrica ja existe no catalogo ativo.",
                    current_record=current,
                )
            )
            continue

        reviews.append(
            ReportRubricReview(
                suggestion=suggestion,
                status=RUBRIC_REVIEW_DIVERGENT,
                message=(
                    "Codigo de rubrica ja existe, mas a descricao do relatorio diverge "
                    "do catalogo atual."
                ),
                current_record=current,
            )
        )
    return tuple(reviews)


def _report_block_reason(
    parsed: ParsedPayrollReport,
    *,
    selected_company_code: str | None,
) -> str | None:
    if not selected_company_code:
        return "Selecione a empresa antes de importar ou aplicar sugestoes do relatorio."

    detected = _clean_text(parsed.detected_company_code)
    if detected and detected != selected_company_code:
        return (
            "A empresa detectada no relatorio diverge da empresa selecionada. "
            f"selecionada={selected_company_code}; detectada={detected}."
        )
    return None


def _extract_rows_and_lines(file_bytes: bytes, *, extension: str) -> tuple[list[list[str]], list[str]]:
    if extension == ".xlsx":
        return _extract_xlsx_rows_and_lines(file_bytes)
    if extension == ".csv":
        text = _decode_text(file_bytes)
        rows = [[_clean_text(cell) or "" for cell in row] for row in csv.reader(io.StringIO(text))]
        return rows, _non_empty_lines(text)
    if extension == ".pdf":
        text = _extract_pdf_text(file_bytes) or _decode_text(file_bytes)
        return [], _non_empty_lines(text)

    text = _decode_text(file_bytes)
    rows = []
    for line in _non_empty_lines(text):
        delimiter = ";" if ";" in line else "|" if "|" in line else None
        if delimiter:
            rows.append([_clean_text(cell) or "" for cell in line.split(delimiter)])
    return rows, _non_empty_lines(text)


def _extract_xlsx_rows_and_lines(file_bytes: bytes) -> tuple[list[list[str]], list[str]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    rows: list[list[str]] = []
    lines: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(values_only=True):
            values = [_cell_text(value) for value in row]
            if not any(values):
                continue
            rows.append(values)
            lines.append(" | ".join(value for value in values if value))
    return rows, lines


def _extract_pdf_text(file_bytes: bytes) -> str | None:
    for module_name in ("pypdf", "PyPDF2"):
        try:
            module = importlib.import_module(module_name)
        except ModuleNotFoundError:
            continue
        try:
            reader = module.PdfReader(io.BytesIO(file_bytes))
            parts = [page.extract_text() or "" for page in reader.pages]
            text = "\n".join(part for part in parts if part.strip())
            if text.strip():
                return text
        except Exception:
            continue
    return None


def _extract_company_metadata(
    rows: list[list[str]],
    lines: list[str],
) -> tuple[str | None, str | None]:
    for row in rows:
        header_indexes = [_normalize_header(cell) for cell in row]
        code_index = _find_header_index(header_indexes, _COMPANY_CODE_HEADERS)
        name_index = _find_header_index(header_indexes, _COMPANY_NAME_HEADERS)
        if code_index is None and name_index is None:
            continue
        next_row = _next_data_row(rows, row)
        if next_row is None:
            continue
        code = _clean_text(_row_value(next_row, code_index)) if code_index is not None else None
        name = _clean_text(_row_value(next_row, name_index)) if name_index is not None else None
        if code or name:
            return code, name

    for line in lines:
        combined = re.search(
            r"\bempresa\b\s*(?:codigo|cod\.?)?\s*[:#-]?\s*(?P<code>\d{1,10})"
            r"(?:\s*[-|]\s*(?P<name>.+))?",
            line,
            flags=re.IGNORECASE,
        )
        if combined:
            return _clean_text(combined.group("code")), _clean_text(combined.group("name"))

    code = None
    name = None
    for line in lines:
        if code is None:
            code = _labeled_segment(
                line,
                labels=("codigo da empresa", "cod empresa", "empresa codigo"),
                stop_labels=("nome da empresa", "razao social", "competencia"),
            )
        if name is None:
            name = _labeled_segment(
                line,
                labels=("nome da empresa", "razao social"),
                stop_labels=("competencia",),
            )
        if code or name:
            return _clean_text(code), _clean_text(name)
    return None, None


def _extract_competence(rows: list[list[str]], lines: list[str]) -> str | None:
    for row_index, row in enumerate(rows):
        header_indexes = [_normalize_header(cell) for cell in row]
        competence_index = _find_header_index(header_indexes, _COMPETENCE_HEADERS)
        if competence_index is None:
            continue
        for data_row in rows[row_index + 1 :]:
            value = _clean_text(_row_value(data_row, competence_index))
            if value:
                return normalize_competence(value)

    for line in lines:
        match = re.search(
            r"\bcompet[eê]ncia\b\s*[:#-]?\s*(?P<competence>\d{1,2}[/-]\d{4}|\d{4}[/-]\d{1,2})",
            line,
            flags=re.IGNORECASE,
        )
        if match:
            return normalize_competence(match.group("competence"))
    return None


def _extract_employees_from_rows(
    rows: list[list[str]],
    origin: ReportSuggestionOrigin,
) -> tuple[ReportEmployeeSuggestion, ...]:
    suggestions: list[ReportEmployeeSuggestion] = []
    for row_index, row in enumerate(rows):
        headers = [_normalize_header(cell) for cell in row]
        code_index = _find_header_index(headers, _EMPLOYEE_CODE_HEADERS)
        name_index = _find_header_index(headers, _EMPLOYEE_NAME_HEADERS)
        if code_index is None or name_index is None:
            continue
        for data_index, data_row in enumerate(rows[row_index + 1 :], start=row_index + 2):
            if not any(data_row):
                break
            if _row_looks_like_header(data_row):
                break
            code = _clean_text(_row_value(data_row, code_index))
            name = _clean_text(_row_value(data_row, name_index))
            if not code or not name:
                continue
            suggestions.append(
                ReportEmployeeSuggestion(
                    domain_registration=code,
                    employee_name=name,
                    source_reference=f"linha {data_index}",
                    origin=_origin_for(origin, "funcionario"),
                )
            )
    return tuple(suggestions)


def _extract_employees_from_text(
    lines: list[str],
    origin: ReportSuggestionOrigin,
) -> tuple[ReportEmployeeSuggestion, ...]:
    suggestions: list[ReportEmployeeSuggestion] = []
    for line_number, line in enumerate(lines, start=1):
        if not re.search(r"\b(funcion[aá]rio|colaborador|empregado)\b", line, flags=re.IGNORECASE):
            continue
        registration = _labeled_segment(
            line,
            labels=("matricula", "matrícula", "codigo", "código"),
            stop_labels=("nome", "funcionario", "colaborador", "rubrica", "total"),
        )
        name = _labeled_segment(
            line,
            labels=("nome",),
            stop_labels=("matricula", "matrícula", "rubrica", "total", "valor"),
        )
        if registration and name:
            suggestions.append(
                ReportEmployeeSuggestion(
                    domain_registration=registration,
                    employee_name=name,
                    source_reference=f"linha {line_number}",
                    origin=_origin_for(origin, "funcionario"),
                )
            )
    return tuple(suggestions)


def _extract_rubrics_from_rows(
    rows: list[list[str]],
    origin: ReportSuggestionOrigin,
) -> tuple[ReportRubricSuggestion, ...]:
    suggestions: list[ReportRubricSuggestion] = []
    for row_index, row in enumerate(rows):
        headers = [_normalize_header(cell) for cell in row]
        code_index = _find_header_index(headers, _RUBRIC_CODE_HEADERS)
        description_index = _find_header_index(headers, _RUBRIC_DESCRIPTION_HEADERS)
        if code_index is None or description_index is None:
            continue
        canonical_index = _find_header_index(headers, _CANONICAL_EVENT_HEADERS)
        value_kind_index = _find_header_index(headers, _VALUE_KIND_HEADERS)
        nature_index = _find_header_index(headers, _NATURE_HEADERS)
        for data_index, data_row in enumerate(rows[row_index + 1 :], start=row_index + 2):
            if not any(data_row):
                break
            if _row_looks_like_header(data_row):
                break
            code = _clean_text(_row_value(data_row, code_index))
            description = _clean_text(_row_value(data_row, description_index))
            if not code or not description:
                continue
            suggestions.append(
                ReportRubricSuggestion(
                    rubric_code=code,
                    description=description,
                    canonical_event=(
                        _clean_text(_row_value(data_row, canonical_index))
                        if canonical_index is not None
                        else None
                    ),
                    value_kind=(
                        _normalize_value_kind(_row_value(data_row, value_kind_index))
                        if value_kind_index is not None
                        else None
                    ),
                    nature=(
                        _normalize_nature(_row_value(data_row, nature_index))
                        if nature_index is not None
                        else None
                    ),
                    source_reference=f"linha {data_index}",
                    origin=_origin_for(origin, "rubrica"),
                )
            )
    return tuple(suggestions)


def _extract_rubrics_from_text(
    lines: list[str],
    origin: ReportSuggestionOrigin,
) -> tuple[ReportRubricSuggestion, ...]:
    suggestions: list[ReportRubricSuggestion] = []
    for line_number, line in enumerate(lines, start=1):
        if not re.search(r"\b(rubrica|cod\.?\s*evento|codigo\s*evento)\b", line, flags=re.IGNORECASE):
            continue
        code = _first_code_token(
            _labeled_segment(
                line,
                labels=("rubrica", "codigo rubrica", "código rubrica", "codigo evento", "cod evento"),
                stop_labels=("descricao", "descrição", "nome", "evento canonico", "tipo", "natureza", "total"),
            )
        )
        description = _labeled_segment(
            line,
            labels=("descricao", "descrição", "nome"),
            stop_labels=("evento canonico", "evento canônico", "tipo", "natureza", "total", "valor"),
        )
        if code and not description:
            description = _description_after_code(line, code)
        if not code or not description:
            continue
        suggestions.append(
            ReportRubricSuggestion(
                rubric_code=code,
                description=description,
                canonical_event=_labeled_segment(
                    line,
                    labels=("evento canonico", "evento canônico", "canonical_event", "evento negocio"),
                    stop_labels=("tipo", "natureza", "total", "valor"),
                ),
                value_kind=_normalize_value_kind(
                    _labeled_segment(
                        line,
                        labels=("tipo do valor", "tipo valor", "tipo"),
                        stop_labels=("natureza", "total", "valor"),
                    )
                ),
                nature=_normalize_nature(
                    _labeled_segment(
                        line,
                        labels=("natureza",),
                        stop_labels=("total", "valor"),
                    )
                ),
                source_reference=f"linha {line_number}",
                origin=_origin_for(origin, "rubrica"),
            )
        )
    return tuple(suggestions)


def _extract_rubric_totals_from_rows(
    rows: list[list[str]],
    origin: ReportSuggestionOrigin,
) -> tuple[ReportRubricTotal, ...]:
    totals: list[ReportRubricTotal] = []
    for row_index, row in enumerate(rows):
        headers = [_normalize_header(cell) for cell in row]
        code_index = _find_header_index(headers, _RUBRIC_CODE_HEADERS)
        description_index = _find_header_index(headers, _RUBRIC_DESCRIPTION_HEADERS)
        total_index = _find_header_index(headers, _RUBRIC_TOTAL_HEADERS)
        if code_index is None or total_index is None:
            continue
        for data_index, data_row in enumerate(rows[row_index + 1 :], start=row_index + 2):
            if not any(data_row):
                break
            if _row_looks_like_header(data_row):
                break
            code = _clean_text(_row_value(data_row, code_index))
            total = _clean_text(_row_value(data_row, total_index))
            if not code or not total:
                continue
            totals.append(
                ReportRubricTotal(
                    rubric_code=code,
                    description=(
                        _clean_text(_row_value(data_row, description_index))
                        if description_index is not None
                        else None
                    ),
                    total_value=total,
                    source_reference=f"linha {data_index}",
                    origin=_origin_for(origin, "total_rubrica"),
                )
            )
    return tuple(totals)


def _extract_rubric_totals_from_text(
    lines: list[str],
    origin: ReportSuggestionOrigin,
) -> tuple[ReportRubricTotal, ...]:
    totals: list[ReportRubricTotal] = []
    for line_number, line in enumerate(lines, start=1):
        if not re.search(r"\brubrica\b", line, flags=re.IGNORECASE):
            continue
        code = _first_code_token(
            _labeled_segment(
                line,
                labels=("rubrica",),
                stop_labels=("descricao", "descrição", "nome", "total", "valor"),
            )
        )
        total = _labeled_segment(
            line,
            labels=("total", "valor total"),
            stop_labels=("rubrica", "descricao", "descrição"),
        )
        if not code or not total:
            continue
        totals.append(
            ReportRubricTotal(
                rubric_code=code,
                description=_description_after_code(line, code),
                total_value=total,
                source_reference=f"linha {line_number}",
                origin=_origin_for(origin, "total_rubrica"),
            )
        )
    return tuple(totals)


def _extract_column_profiles_from_rows(
    rows: list[list[str]],
    origin: ReportSuggestionOrigin,
) -> tuple[ReportColumnProfileSuggestion, ...]:
    suggestions: list[ReportColumnProfileSuggestion] = []
    for row_index, row in enumerate(rows):
        headers = [_clean_text(cell) for cell in row]
        normalized_headers = [_normalize_header(cell) for cell in row]
        if not _row_looks_like_header(row):
            continue
        for column_name, normalized in zip(headers, normalized_headers, strict=False):
            if not column_name or _is_identity_or_registry_header(normalized):
                continue
            match = re.search(r"\b(?:cod\.?|codigo|c[oó]digo|rubrica)\s*\.?\s*(\d{1,10})\b", column_name, re.IGNORECASE)
            if not match:
                continue
            suggestions.append(
                ReportColumnProfileSuggestion(
                    column_name=column_name,
                    generation_mode="single_line",
                    rubrica_target=match.group(1),
                    value_kind=None,
                    source_reference=f"linha {row_index + 1}",
                    origin=_origin_for(origin, "perfil_coluna"),
                    notes="Rubrica sugerida apenas porque o codigo aparece no nome da coluna.",
                )
            )
    return tuple(suggestions)


def _dedupe_employees(
    suggestions: tuple[ReportEmployeeSuggestion, ...],
) -> tuple[ReportEmployeeSuggestion, ...]:
    seen: set[str] = set()
    output: list[ReportEmployeeSuggestion] = []
    for suggestion in suggestions:
        key = _normalize_lookup_token(suggestion.domain_registration)
        if key in seen:
            continue
        seen.add(key)
        output.append(suggestion)
    return tuple(output)


def _dedupe_rubrics(
    suggestions: tuple[ReportRubricSuggestion, ...],
) -> tuple[ReportRubricSuggestion, ...]:
    seen: set[str] = set()
    output: list[ReportRubricSuggestion] = []
    for suggestion in suggestions:
        key = _normalize_lookup_token(suggestion.rubric_code)
        if key in seen:
            continue
        seen.add(key)
        output.append(suggestion)
    return tuple(output)


def _dedupe_rubric_totals(
    totals: tuple[ReportRubricTotal, ...],
) -> tuple[ReportRubricTotal, ...]:
    by_code: dict[str, ReportRubricTotal] = {}
    for total in totals:
        key = _normalize_lookup_token(total.rubric_code)
        existing = by_code.get(key)
        if existing is None:
            by_code[key] = total
            continue
        merged = _merge_total_values(existing.total_value, total.total_value)
        if merged is not None:
            by_code[key] = ReportRubricTotal(
                rubric_code=existing.rubric_code,
                description=existing.description or total.description,
                total_value=merged,
                source_reference=f"{existing.source_reference}; {total.source_reference}",
                origin=existing.origin,
            )
    return tuple(by_code.values())


def _dedupe_column_profiles(
    suggestions: tuple[ReportColumnProfileSuggestion, ...],
) -> tuple[ReportColumnProfileSuggestion, ...]:
    seen: set[str] = set()
    output: list[ReportColumnProfileSuggestion] = []
    for suggestion in suggestions:
        key = _normalize_lookup_token(suggestion.column_name)
        if key in seen:
            continue
        seen.add(key)
        output.append(suggestion)
    return tuple(output)


def _merge_total_values(left: str, right: str) -> str | None:
    left_decimal = _parse_decimal(left)
    right_decimal = _parse_decimal(right)
    if left_decimal is None or right_decimal is None:
        return None
    return _format_decimal(left_decimal + right_decimal)


def _parse_decimal(value: str) -> Decimal | None:
    text = _clean_text(value)
    if text is None:
        return None
    normalized = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return None


def _format_decimal(value: Decimal) -> str:
    return f"{value:.2f}".replace(".", ",")


def _employee_name_matches(current: CompanyEmployeeRecord, report_name: str) -> bool:
    report_token = _normalize_lookup_token(report_name)
    names = [current.employee_name, *current.aliases]
    return any(_normalize_lookup_token(name) == report_token for name in names)


def _same_lookup(left: Any, right: Any) -> bool:
    return _normalize_lookup_token(left) == _normalize_lookup_token(right)


def _origin_for(origin: ReportSuggestionOrigin, data_type: str) -> ReportSuggestionOrigin:
    return ReportSuggestionOrigin(
        file_name=origin.file_name,
        competence=origin.competence,
        detected_company_code=origin.detected_company_code,
        detected_company_name=origin.detected_company_name,
        parsed_at=origin.parsed_at,
        data_type=data_type,
    )


def _origin_note(origin: ReportSuggestionOrigin) -> str:
    return json.dumps(
        {
            "origem_sugestao": {
                "arquivo": origin.file_name,
                "competencia": origin.competence,
                "empresa_detectada": origin.detected_company_code,
                "nome_empresa_detectado": origin.detected_company_name,
                "timestamp": origin.parsed_at.isoformat(),
                "tipo_dado": origin.data_type,
            }
        },
        ensure_ascii=True,
        sort_keys=True,
    )


def _normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", "" if value is None else str(value))
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = re.sub(r"[^A-Za-z0-9]+", "_", text.strip().lower())
    return text.strip("_")


def _normalize_lookup_token(value: Any) -> str:
    text = unicodedata.normalize("NFKD", "" if value is None else str(value))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(text.strip().lower().split())


def _find_header_index(headers: list[str], aliases: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in aliases:
            return index
    return None


def _is_identity_or_registry_header(normalized_header: str) -> bool:
    known_headers = (
        _EMPLOYEE_CODE_HEADERS
        | _EMPLOYEE_NAME_HEADERS
        | _RUBRIC_CODE_HEADERS
        | _RUBRIC_DESCRIPTION_HEADERS
        | _CANONICAL_EVENT_HEADERS
        | _VALUE_KIND_HEADERS
        | _NATURE_HEADERS
        | _RUBRIC_TOTAL_HEADERS
        | _COMPANY_CODE_HEADERS
        | _COMPANY_NAME_HEADERS
        | _COMPETENCE_HEADERS
    )
    return normalized_header in known_headers


def _row_looks_like_header(row: list[str]) -> bool:
    normalized = {_normalize_header(cell) for cell in row if _clean_text(cell)}
    return bool(
        normalized
        & (
            _EMPLOYEE_CODE_HEADERS
            | _EMPLOYEE_NAME_HEADERS
            | _RUBRIC_CODE_HEADERS
            | _RUBRIC_DESCRIPTION_HEADERS
        )
    )


def _next_data_row(rows: list[list[str]], current_row: list[str]) -> list[str] | None:
    try:
        index = rows.index(current_row)
    except ValueError:
        return None
    for candidate in rows[index + 1 :]:
        if any(candidate):
            return candidate
    return None


def _row_value(row: list[str], index: int | None) -> str | None:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _labeled_segment(
    line: str,
    *,
    labels: tuple[str, ...],
    stop_labels: tuple[str, ...] = (),
) -> str | None:
    label_pattern = "|".join(re.escape(label) for label in sorted(labels, key=len, reverse=True))
    match = re.search(rf"(?:{label_pattern})\s*[:#-]?\s*(?P<value>.+)", line, flags=re.IGNORECASE)
    if not match:
        return None
    value = match.group("value")
    if stop_labels:
        stop_pattern = "|".join(re.escape(label) for label in sorted(stop_labels, key=len, reverse=True))
        stop = re.search(rf"\b(?:{stop_pattern})\b\s*[:#-]?", value, flags=re.IGNORECASE)
        if stop:
            value = value[: stop.start()]
    value = re.split(r"[;|]", value, maxsplit=1)[0]
    return _clean_text(value.strip(" -"))


def _description_after_code(line: str, code: str) -> str | None:
    pattern = re.escape(code)
    match = re.search(pattern + r"\s*[-:]\s*(?P<description>.+)", line)
    if not match:
        return None
    description = match.group("description")
    description = re.split(
        r"\b(evento\s+canonico|evento\s+can[oô]nico|tipo|natureza|total|valor)\b",
        description,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    return _clean_text(description.strip(" -"))


def _first_code_token(value: str | None) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    match = re.match(r"(?P<code>[A-Za-z0-9_.-]+)", text)
    return _clean_text(match.group("code")) if match else None


def _normalize_value_kind(value: Any) -> str | None:
    token = _normalize_lookup_token(value)
    if not token:
        return None
    if token in {"monetario", "monetario valor", "valor", "dinheiro", "moeda"}:
        return RubricValueKind.MONETARY.value
    if token in {"hora", "horas", "hh mm", "hh:mm"}:
        return RubricValueKind.HOURS.value
    if token in {"quantidade", "qtd", "qtde"}:
        return RubricValueKind.QUANTITY.value
    return None


def _normalize_nature(value: Any) -> str | None:
    token = _normalize_lookup_token(value)
    if not token:
        return None
    if token in {"provento", "vencimento", "credito"}:
        return RubricNature.PROVENTO.value
    if token in {"desconto", "debito"}:
        return RubricNature.DESCONTO.value
    if token in {"informativo", "base", "referencia"}:
        return RubricNature.INFORMATIVO.value
    if token == "unknown":
        return RubricNature.UNKNOWN.value
    return None


def _clean_text(value: Any) -> str | None:
    return normalize_text(value)


def _cell_text(value: Any) -> str:
    text = _clean_text(value)
    return text or ""


def _decode_text(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")


def _non_empty_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


__all__ = [
    "EMPLOYEE_REVIEW_DIVERGENT",
    "EMPLOYEE_REVIEW_EXISTING",
    "EMPLOYEE_REVIEW_NEW",
    "REPORT_IMPORTER_SOURCE",
    "RUBRIC_REVIEW_DIVERGENT",
    "RUBRIC_REVIEW_EXISTING",
    "RUBRIC_REVIEW_INCOMPLETE",
    "RUBRIC_REVIEW_NEW",
    "ParsedPayrollReport",
    "ReportApplyResult",
    "ReportColumnProfileSuggestion",
    "ReportEmployeeReview",
    "ReportEmployeeSuggestion",
    "ReportImportAnalysis",
    "ReportRubricReview",
    "ReportRubricSuggestion",
    "ReportRubricTotal",
    "ReportSuggestionOrigin",
    "analyze_report_import",
    "apply_report_employee_suggestions",
    "apply_report_rubric_suggestions",
    "parse_report_file",
]
