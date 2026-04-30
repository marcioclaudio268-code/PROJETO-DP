"""Guided overrides applied on top of the editable dashboard workspace."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Mapping
from uuid import uuid4

from openpyxl import load_workbook

from config import CompanyConfig
from ingestion.template_v1_loader import EVENT_SPECS

from .column_mapping_profiles import (
    ColumnGenerationMode,
    ColumnMappingProfileError,
    ColumnMappingRule,
    CompanyColumnMappingProfile,
    load_column_mapping_profile,
    save_column_mapping_profile,
    upsert_column_mapping_rule,
)
from .company_employee_registry import (
    CompanyEmployeeRecord,
    load_company_employee_registry,
    save_company_employee_registry,
    upsert_employee_record,
)
from .company_rubric_catalog import (
    CompanyRubricRecord,
    load_company_rubric_catalog,
    save_company_rubric_catalog,
    upsert_rubric_record,
)
from .errors import DashboardOperationError
from .models import (
    DashboardActionRecord,
    DashboardActionType,
    DashboardPaths,
    DashboardPendingItem,
    DashboardState,
)
from .storage import load_dashboard_state, write_dashboard_state


EVENT_COLUMN_NAMES = tuple(spec.column_name for spec in EVENT_SPECS)
COLUMN_MAPPING_PROFILE_CODES = {
    "column_mapping_profile_missing",
    "column_mapping_profile_incomplete",
}
IGNORABLE_NOTE_CODES = {"observacao_ambigua"}
IGNORABLE_EVENT_CODES = {
    "evento_nao_automatizavel",
    "hora_invalida",
    "valor_invalido",
    "quantidade_invalida",
    "mapeamento_evento_ausente",
    "mapeamento_evento_inativo",
    "evento_requer_revisao_por_politica",
}
IGNORABLE_LINE_CODES = {
    "funcionario_nao_encontrado",
    "matricula_dominio_ausente",
    "matricula_dominio_informada_somente_na_linha",
    "matricula_dominio_divergente",
    "linha_bloqueada_por_status",
    "mapeamento_matricula_ausente",
    "mapeamento_matricula_divergente_config",
    "mapeamento_matricula_ambiguo",
}


def apply_dashboard_action(
    paths: DashboardPaths,
    *,
    action_type: DashboardActionType | str,
    pending_uid: str,
    payload: Mapping[str, Any] | None = None,
    employee_registry_root: str | Path | None = None,
    rubric_catalog_root: str | Path | None = None,
    column_profile_root: str | Path | None = None,
) -> DashboardActionRecord:
    """Validate and apply one guided action against a persisted dashboard pending."""

    try:
        normalized_action_type = DashboardActionType(action_type)
    except ValueError as exc:
        raise DashboardOperationError(
            "acao_manual_invalida",
            f"Tipo de acao manual nao suportado: {action_type}.",
            source=str(action_type),
        ) from exc

    pending_uid = _required_text("pending_uid", pending_uid)
    action_payload = dict(payload or {})
    pending_item = _load_pending_from_state(paths, pending_uid)

    if normalized_action_type == DashboardActionType.EMPLOYEE_MAPPING_UPDATE:
        return _apply_employee_mapping_action(
            paths,
            pending_item,
            action_payload,
            employee_registry_root=employee_registry_root,
        )

    if normalized_action_type == DashboardActionType.EVENT_MAPPING_UPDATE:
        return _apply_event_mapping_action(
            paths,
            pending_item,
            action_payload,
            rubric_catalog_root=rubric_catalog_root,
        )

    if normalized_action_type == DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE:
        return _apply_column_mapping_profile_action(
            paths,
            pending_item,
            action_payload,
            column_profile_root=column_profile_root,
        )

    if normalized_action_type == DashboardActionType.IGNORE_PENDING:
        return _apply_ignore_action(paths, pending_item)

    if normalized_action_type == DashboardActionType.WORKBOOK_CELL_UPDATE:
        return _apply_workbook_cell_action(paths, pending_item, action_payload)

    raise DashboardOperationError(
        "acao_manual_sem_aplicador",
        f"A acao manual {normalized_action_type.value} ainda nao possui aplicador backend.",
        source=pending_item.uid,
    )


def apply_workbook_cell_correction(
    paths: DashboardPaths,
    *,
    sheet_name: str,
    cell: str,
    new_value: str | None,
    pending_uid: str | None = None,
    description: str | None = None,
) -> DashboardActionRecord:
    workbook = load_workbook(paths.editable_workbook_path)
    if sheet_name not in workbook.sheetnames:
        raise DashboardOperationError(
            "aba_para_correcao_ausente",
            f"Aba '{sheet_name}' nao encontrada na planilha editavel.",
            source=sheet_name,
        )

    worksheet = workbook[sheet_name]
    original_value = worksheet[cell].value
    worksheet[cell] = None if new_value in {"", None} else new_value
    workbook.save(paths.editable_workbook_path)

    action = DashboardActionRecord(
        action_id=f"act-{uuid4().hex[:10]}",
        action_type=DashboardActionType.WORKBOOK_CELL_UPDATE,
        pending_uid=pending_uid,
        description=description or f"Correcao aplicada em {sheet_name}!{cell}.",
        payload={
            "scope": "current_run_editable_workbook",
            "sheet_name": sheet_name,
            "cell": cell,
            "original_value": _stringify(original_value),
            "new_value": _stringify(worksheet[cell].value),
        },
    )
    _append_action(paths, action)
    return action


def upsert_employee_mapping_override(
    paths: DashboardPaths,
    *,
    employee_key: str,
    domain_registration: str,
    employee_name: str | None = None,
    pending_uid: str | None = None,
    scope: str = "current_run_editable_config",
    extra_payload: Mapping[str, Any] | None = None,
) -> DashboardActionRecord:
    payload = _load_company_config_payload(paths.editable_config_path)
    _upsert_employee_mapping_payload(
        payload,
        employee_key=employee_key,
        domain_registration=domain_registration,
        employee_name=employee_name,
    )
    _write_validated_company_config(paths.editable_config_path, payload)

    action = DashboardActionRecord(
        action_id=f"act-{uuid4().hex[:10]}",
        action_type=DashboardActionType.EMPLOYEE_MAPPING_UPDATE,
        pending_uid=pending_uid,
        description=f"Correcao de matricula aplicada para a chave '{employee_key}'.",
        payload={
            "scope": scope,
            "employee_key": employee_key,
            "domain_registration": domain_registration,
            "employee_name": employee_name,
            **dict(extra_payload or {}),
        },
    )
    _append_action(paths, action)
    return action


def upsert_event_mapping_override(
    paths: DashboardPaths,
    *,
    event_name: str,
    output_rubric: str,
    pending_uid: str | None = None,
    scope: str = "current_run_editable_config",
    extra_payload: Mapping[str, Any] | None = None,
) -> DashboardActionRecord:
    payload = _load_company_config_payload(paths.editable_config_path)
    _upsert_event_mapping_payload(
        payload,
        event_name=event_name,
        output_rubric=output_rubric,
    )
    _write_validated_company_config(paths.editable_config_path, payload)

    action = DashboardActionRecord(
        action_id=f"act-{uuid4().hex[:10]}",
        action_type=DashboardActionType.EVENT_MAPPING_UPDATE,
        pending_uid=pending_uid,
        description=f"Correcao de rubrica aplicada para o evento '{event_name}'.",
        payload={
            "scope": scope,
            "event_name": event_name,
            "output_rubric": output_rubric,
            **dict(extra_payload or {}),
        },
    )
    _append_action(paths, action)
    return action


def replay_dashboard_action_overrides(paths: DashboardPaths, state: DashboardState) -> None:
    """Reapply persisted config overrides after ConfigResolver rematerializes the run config."""

    if not state.actions:
        return

    if not paths.editable_config_path.exists():
        return

    payload = _load_company_config_payload(paths.editable_config_path)
    changed = False
    for action in state.actions:
        if action.action_type == DashboardActionType.EMPLOYEE_MAPPING_UPDATE:
            employee_key = _required_text_from_payload(action.payload, "employee_key")
            domain_registration = _required_text_from_payload(action.payload, "domain_registration")
            employee_name = _stringify(action.payload.get("employee_name"))
            _upsert_employee_mapping_payload(
                payload,
                employee_key=employee_key,
                domain_registration=domain_registration,
                employee_name=employee_name,
            )
            changed = True
        elif action.action_type == DashboardActionType.EVENT_MAPPING_UPDATE:
            event_name = _required_text_from_payload(action.payload, "event_name")
            output_rubric = _required_text_from_payload(action.payload, "output_rubric")
            _upsert_event_mapping_payload(
                payload,
                event_name=event_name,
                output_rubric=output_rubric,
            )
            changed = True

    if changed:
        _write_validated_company_config(paths.editable_config_path, payload)


def ignore_pending_for_import(
    paths: DashboardPaths,
    pending_item: DashboardPendingItem,
) -> DashboardActionRecord:
    if not pending_item.can_ignore or pending_item.ignore_mode is None:
        raise DashboardOperationError(
            "ignorar_nao_permitido",
            "Este item nao pode ser ignorado com seguranca nesta importacao.",
            source=pending_item.uid,
        )

    if pending_item.source_sheet != "LANCAMENTOS_FACEIS":
        raise DashboardOperationError(
            "ignorar_sem_apoio_operacional",
            "Somente itens originados em LANCAMENTOS_FACEIS podem ser ignorados neste MVP.",
            source=pending_item.uid,
        )

    workbook = load_workbook(paths.editable_workbook_path)
    worksheet = workbook[pending_item.source_sheet]

    if pending_item.ignore_mode == "observacao":
        cells_to_clear = (pending_item.source_cell,)
    elif pending_item.ignore_mode == "evento":
        cells_to_clear = (pending_item.source_cell,)
    elif pending_item.ignore_mode == "linha":
        if pending_item.source_row is None:
            raise DashboardOperationError(
                "linha_para_ignorar_ausente",
                "Nao foi possivel identificar a linha a ser ignorada nesta importacao.",
                source=pending_item.uid,
            )
        cells_to_clear = _event_cells_for_row(worksheet, pending_item.source_row)
    else:
        raise DashboardOperationError(
            "modo_ignorar_nao_suportado",
            f"Modo de ignorar nao suportado: {pending_item.ignore_mode}.",
            source=pending_item.uid,
        )

    original_values: dict[str, str | None] = {}
    for cell in cells_to_clear:
        if cell is None:
            continue
        original_values[cell] = _stringify(worksheet[cell].value)
        worksheet[cell] = None

    workbook.save(paths.editable_workbook_path)

    action = DashboardActionRecord(
        action_id=f"act-{uuid4().hex[:10]}",
        action_type=DashboardActionType.IGNORE_PENDING,
        pending_uid=pending_item.uid,
        description=pending_item.ignore_label or "Item ignorado nesta importacao.",
        payload={
            "scope": "current_import_only",
            "ignore_mode": pending_item.ignore_mode,
            "cleared_cells": [cell for cell in cells_to_clear if cell is not None],
            "original_values": original_values,
            "pending_code": pending_item.code,
        },
    )
    _append_action(paths, action)
    return action


def describe_ignore_strategy(
    *,
    code: str,
    source_sheet: str | None,
    source_cell: str | None,
    source_row: int | None,
    event_name: str | None,
) -> tuple[bool, str | None, str | None]:
    if source_sheet != "LANCAMENTOS_FACEIS":
        return False, None, None

    if code in IGNORABLE_NOTE_CODES and source_cell is not None:
        return True, "observacao", "Ignorar observacao nesta importacao"

    if code in IGNORABLE_EVENT_CODES and source_cell is not None and event_name is not None:
        return True, "evento", "Ignorar este evento nesta importacao"

    if code in IGNORABLE_LINE_CODES and source_row is not None:
        return True, "linha", "Ignorar esta linha nesta importacao"

    return False, None, None


def _apply_employee_mapping_action(
    paths: DashboardPaths,
    pending_item: DashboardPendingItem,
    payload: dict[str, Any],
    *,
    employee_registry_root: str | Path | None,
) -> DashboardActionRecord:
    if not pending_item.can_edit_employee_mapping:
        raise DashboardOperationError(
            "acao_incompativel_com_pendencia",
            "Esta pendencia nao permite correcao de matricula.",
            source=pending_item.uid,
        )

    if not pending_item.employee_key:
        raise DashboardOperationError(
            "pendencia_sem_funcionario",
            "A pendencia nao possui chave de funcionario suficiente para corrigir matricula.",
            source=pending_item.uid,
        )

    config_payload = _require_editable_company_config(paths)
    domain_registration = _required_text_from_payload(
        payload,
        "domain_registration",
        aliases=("matricula_dominio", "registration"),
    )
    employee_name = _stringify(payload.get("employee_name")) or pending_item.employee_name
    action_extra_payload: dict[str, Any] = {}
    scope = "current_run_editable_config"
    if _as_bool(payload.get("persist_to_employee_registry")):
        registry_path = _persist_employee_mapping_to_registry(
            company_code=_required_text("company_code", config_payload.get("company_code")),
            company_name=_stringify(config_payload.get("company_name")),
            employee_key=pending_item.employee_key,
            employee_name=_required_text("employee_name", employee_name),
            domain_registration=domain_registration,
            aliases=_aliases_from_payload(payload),
            root=employee_registry_root,
        )
        action_extra_payload = {
            "persist_to_employee_registry": True,
            "scopes": ["current_run_editable_config", "company_employee_registry"],
            "employee_registry_path": str(registry_path),
        }
    else:
        action_extra_payload = {
            "persist_to_employee_registry": False,
            "scopes": ["current_run_editable_config"],
        }

    return upsert_employee_mapping_override(
        paths,
        employee_key=pending_item.employee_key,
        employee_name=employee_name,
        domain_registration=domain_registration,
        pending_uid=pending_item.uid,
        scope=scope,
        extra_payload=action_extra_payload,
    )


def _apply_event_mapping_action(
    paths: DashboardPaths,
    pending_item: DashboardPendingItem,
    payload: dict[str, Any],
    *,
    rubric_catalog_root: str | Path | None,
) -> DashboardActionRecord:
    if not pending_item.can_edit_event_mapping:
        raise DashboardOperationError(
            "acao_incompativel_com_pendencia",
            "Esta pendencia nao permite correcao de rubrica.",
            source=pending_item.uid,
        )

    if not pending_item.event_name:
        raise DashboardOperationError(
            "pendencia_sem_evento",
            "A pendencia nao possui evento suficiente para corrigir rubrica.",
            source=pending_item.uid,
        )

    config_payload = _require_editable_company_config(paths)
    output_rubric = _required_text_from_payload(
        payload,
        "output_rubric",
        aliases=("rubrica_saida", "rubrica"),
    )
    action_extra_payload: dict[str, Any]
    scope = "current_run_editable_config"
    if _as_bool(payload.get("persist_to_rubric_catalog")):
        catalog_path = _persist_event_mapping_to_rubric_catalog(
            company_code=_required_text("company_code", config_payload.get("company_code")),
            company_name=_stringify(config_payload.get("company_name")),
            rubric_code=output_rubric,
            description=_required_text_from_payload(payload, "description"),
            canonical_event=_required_text(
                "canonical_event",
                payload.get("canonical_event") or pending_item.event_name,
            ),
            value_kind=_required_text_from_payload(payload, "value_kind"),
            nature=_stringify(payload.get("nature")) or "unknown",
            aliases=_aliases_from_payload(payload),
            root=rubric_catalog_root,
        )
        action_extra_payload = {
            "persist_to_rubric_catalog": True,
            "scopes": ["current_run_editable_config", "company_rubric_catalog"],
            "rubric_catalog_path": str(catalog_path),
            "description": _required_text_from_payload(payload, "description"),
            "canonical_event": _required_text(
                "canonical_event",
                payload.get("canonical_event") or pending_item.event_name,
            ),
            "value_kind": _required_text_from_payload(payload, "value_kind"),
            "nature": _stringify(payload.get("nature")) or "unknown",
            "aliases": _aliases_from_payload(payload),
        }
    else:
        action_extra_payload = {
            "persist_to_rubric_catalog": False,
            "scopes": ["current_run_editable_config"],
        }

    return upsert_event_mapping_override(
        paths,
        event_name=pending_item.event_name,
        output_rubric=output_rubric,
        pending_uid=pending_item.uid,
        scope=scope,
        extra_payload=action_extra_payload,
    )


def _apply_column_mapping_profile_action(
    paths: DashboardPaths,
    pending_item: DashboardPendingItem,
    payload: dict[str, Any],
    *,
    column_profile_root: str | Path | None,
) -> DashboardActionRecord:
    if pending_item.stage != "perfil_colunas" or pending_item.code not in COLUMN_MAPPING_PROFILE_CODES:
        raise DashboardOperationError(
            "acao_incompativel_com_pendencia",
            "Esta pendencia nao permite correcao de perfil de colunas.",
            source=pending_item.uid,
        )

    profile_context = _profile_context_from_state(paths)
    company_code = _required_text("company_code", profile_context.get("company_code"))
    company_name = _stringify(profile_context.get("company_name"))
    default_process = _stringify(payload.get("default_process") or profile_context.get("default_process"))
    column_name = _resolve_column_name_for_profile_action(pending_item, payload)
    column_key = _stringify(payload.get("column_key"))
    rule = _column_mapping_rule_from_payload(
        payload,
        column_name=column_name,
        column_key=column_key,
    )

    try:
        profile = load_column_mapping_profile(company_code, root=column_profile_root)
        updated_profile = upsert_column_mapping_rule(profile, rule)
    except ColumnMappingProfileError as exc:
        if exc.code != "profile_not_found":
            raise DashboardOperationError(
                "perfil_colunas_invalido",
                f"Nao foi possivel carregar o perfil de colunas da empresa: {exc}",
                source=exc.source,
            ) from exc
        updated_profile = CompanyColumnMappingProfile(
            company_code=company_code,
            company_name=company_name,
            default_process=default_process,
            mappings=[rule],
        )
    except Exception as exc:
        raise DashboardOperationError(
            "perfil_colunas_invalido",
            f"Nao foi possivel atualizar o perfil de colunas da empresa: {exc}",
            source=company_code,
        ) from exc

    if company_name and not updated_profile.company_name:
        updated_profile = updated_profile.model_copy(update={"company_name": company_name})
    if default_process and not updated_profile.default_process:
        updated_profile = updated_profile.model_copy(update={"default_process": default_process})

    try:
        profile_path = save_column_mapping_profile(updated_profile, root=column_profile_root)
    except Exception as exc:
        raise DashboardOperationError(
            "perfil_colunas_invalido",
            f"Nao foi possivel salvar o perfil de colunas da empresa: {exc}",
            source=company_code,
        ) from exc

    action = DashboardActionRecord(
        action_id=f"act-{uuid4().hex[:10]}",
        action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
        pending_uid=pending_item.uid,
        description=f"Regra de perfil de coluna aplicada para '{rule.source_column_id}'.",
        payload={
            "scope": "company_column_mapping_profile",
            "scopes": ["company_column_mapping_profile"],
            "company_code": company_code,
            "company_name": company_name,
            "column_profile_path": str(profile_path),
            "column_key": rule.column_key,
            "column_name": rule.column_name,
            "enabled": rule.enabled,
            "rubrica_target": rule.rubrica_target,
            "rubricas_target": list(rule.rubricas_target),
            "value_kind": rule.value_kind.value,
            "generation_mode": rule.generation_mode.value,
            "ignore_zero": rule.ignore_zero,
            "ignore_text": rule.ignore_text,
            "notes": rule.notes,
        },
    )
    _append_action(paths, action)
    return action


def _apply_ignore_action(
    paths: DashboardPaths,
    pending_item: DashboardPendingItem,
) -> DashboardActionRecord:
    return ignore_pending_for_import(paths, pending_item)


def _apply_workbook_cell_action(
    paths: DashboardPaths,
    pending_item: DashboardPendingItem,
    payload: dict[str, Any],
) -> DashboardActionRecord:
    if not pending_item.can_edit_workbook:
        raise DashboardOperationError(
            "acao_incompativel_com_pendencia",
            "Esta pendencia nao permite correcao direta de celula.",
            source=pending_item.uid,
        )

    if not pending_item.source_sheet or not pending_item.source_cell:
        raise DashboardOperationError(
            "pendencia_sem_celula",
            "A pendencia nao possui celula de origem suficiente para correcao.",
            source=pending_item.uid,
        )

    new_value = payload.get("new_value")
    return apply_workbook_cell_correction(
        paths,
        sheet_name=pending_item.source_sheet,
        cell=pending_item.source_cell,
        new_value=None if new_value is None else str(new_value),
        pending_uid=pending_item.uid,
    )


def _profile_context_from_state(paths: DashboardPaths) -> dict[str, Any]:
    state = load_dashboard_state(paths.state_path)
    if state.last_analysis is None:
        raise DashboardOperationError(
            "analise_ausente_para_acao",
            "Nao ha analise anterior com contexto de perfil para aplicar esta acao.",
            source=str(paths.state_path),
        )

    summary = state.last_analysis.get("summary", {})
    profile_resolution = state.last_analysis.get("profile_resolution", {})
    return {
        "company_code": profile_resolution.get("company_code") or summary.get("company_code"),
        "company_name": summary.get("company_name"),
        "default_process": summary.get("default_process"),
    }


def _resolve_column_name_for_profile_action(
    pending_item: DashboardPendingItem,
    payload: Mapping[str, Any],
) -> str | None:
    payload_column_name = _stringify(payload.get("column_name"))
    pending_column_name = pending_item.source_column_name

    if pending_column_name:
        if payload_column_name and (
            _normalize_profile_action_token(payload_column_name)
            != _normalize_profile_action_token(pending_column_name)
        ):
            raise DashboardOperationError(
                "coluna_divergente_da_pendencia",
                "A coluna informada nao corresponde a coluna pendente.",
                source=pending_item.uid,
            )
        return pending_column_name

    if payload_column_name:
        return payload_column_name

    if _stringify(payload.get("column_key")):
        return None

    raise DashboardOperationError(
        "campo_obrigatorio_ausente",
        "Campo obrigatorio ausente para acao manual: column_name.",
        source="column_name",
    )


def _column_mapping_rule_from_payload(
    payload: Mapping[str, Any],
    *,
    column_name: str | None,
    column_key: str | None,
) -> ColumnMappingRule:
    try:
        generation_mode = ColumnGenerationMode(_required_text_from_payload(payload, "generation_mode"))
    except ValueError as exc:
        raise DashboardOperationError(
            "perfil_colunas_regra_invalida",
            f"Modo de geracao invalido para perfil de coluna: {payload.get('generation_mode')}.",
            source="generation_mode",
        ) from exc
    enabled = _as_bool(payload["enabled"]) if "enabled" in payload else generation_mode != ColumnGenerationMode.IGNORE
    rubrica_target = _stringify(payload.get("rubrica_target"))
    rubricas_target = _rubricas_target_from_payload(payload)

    if generation_mode == ColumnGenerationMode.IGNORE:
        rubrica_target = None
        rubricas_target = []

    try:
        return ColumnMappingRule(
            column_key=column_key,
            column_name=column_name,
            enabled=enabled,
            rubrica_target=rubrica_target,
            rubricas_target=rubricas_target,
            value_kind=_required_text_from_payload(payload, "value_kind"),
            generation_mode=generation_mode,
            ignore_zero=_required_bool_from_payload(payload, "ignore_zero"),
            ignore_text=_required_bool_from_payload(payload, "ignore_text"),
            notes=_stringify(payload.get("notes")),
        )
    except Exception as exc:
        raise DashboardOperationError(
            "perfil_colunas_regra_invalida",
            f"Regra de perfil de coluna invalida: {exc}",
            source=column_name or column_key,
        ) from exc


def _rubricas_target_from_payload(payload: Mapping[str, Any]) -> list[str]:
    if "rubricas_target" not in payload:
        return []
    raw_targets = payload.get("rubricas_target")
    if not isinstance(raw_targets, list):
        raise DashboardOperationError(
            "campo_obrigatorio_ausente",
            "Campo rubricas_target deve ser informado como lista.",
            source="rubricas_target",
        )
    return [_required_text("rubricas_target", target) for target in raw_targets]


def _required_bool_from_payload(payload: Mapping[str, Any], field_name: str) -> bool:
    if field_name not in payload:
        raise DashboardOperationError(
            "campo_obrigatorio_ausente",
            f"Campo obrigatorio ausente para acao manual: {field_name}.",
            source=field_name,
        )
    return _as_bool(payload[field_name])


def _normalize_profile_action_token(value: str) -> str:
    return " ".join(str(value).strip().lower().split())


def _event_cells_for_row(worksheet, row_number: int) -> tuple[str | None, ...]:
    header_map = {
        str(worksheet.cell(row=1, column=column_index).value): column_index
        for column_index in range(1, worksheet.max_column + 1)
        if worksheet.cell(row=1, column=column_index).value is not None
    }

    cells: list[str] = []
    for header_name in (*EVENT_COLUMN_NAMES, "observacao_eventos"):
        column_index = header_map.get(header_name)
        if column_index is None:
            continue
        cells.append(f"{_column_letter(column_index)}{row_number}")
    return tuple(cells)


def _column_letter(column_index: int) -> str:
    quotient = column_index
    letters = ""
    while quotient:
        quotient, remainder = divmod(quotient - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _append_action(paths: DashboardPaths, action: DashboardActionRecord) -> None:
    state = load_dashboard_state(paths.state_path)
    updated_actions = [*state.actions, action]
    updated_state = state.model_copy(update={"actions": updated_actions})
    write_dashboard_state(paths.state_path, updated_state)


def _upsert_employee_mapping_payload(
    payload: dict,
    *,
    employee_key: str,
    domain_registration: str,
    employee_name: str | None,
) -> None:
    mappings = list(payload.get("employee_mappings", []))

    updated = False
    for item in mappings:
        if item.get("source_employee_key") == employee_key:
            item["domain_registration"] = domain_registration
            if employee_name:
                item["source_employee_name"] = employee_name
            item["active"] = True
            updated = True
            break

    if not updated:
        mappings.append(
            {
                "source_employee_key": employee_key,
                "source_employee_name": employee_name,
                "domain_registration": domain_registration,
                "active": True,
                "aliases": [],
                "notes": None,
            }
        )

    payload["employee_mappings"] = mappings


def _upsert_event_mapping_payload(
    payload: dict,
    *,
    event_name: str,
    output_rubric: str,
) -> None:
    mappings = list(payload.get("event_mappings", []))

    updated = False
    for item in mappings:
        if item.get("event_negocio") == event_name:
            item["rubrica_saida"] = output_rubric
            item["active"] = True
            updated = True
            break

    if not updated:
        mappings.append(
            {
                "event_negocio": event_name,
                "rubrica_saida": output_rubric,
                "active": True,
                "notes": None,
            }
        )

    payload["event_mappings"] = mappings


def _load_pending_from_state(paths: DashboardPaths, pending_uid: str) -> DashboardPendingItem:
    state = load_dashboard_state(paths.state_path)
    if state.last_analysis is None:
        raise DashboardOperationError(
            "analise_ausente_para_acao",
            "Nao ha analise anterior com pendencias para aplicar esta acao.",
            source=str(paths.state_path),
        )

    for item in state.last_analysis.get("pendings", ()):
        if item.get("uid") == pending_uid:
            return DashboardPendingItem(**item)

    raise DashboardOperationError(
        "pendencia_nao_encontrada",
        "Pendencia nao encontrada no estado atual do dashboard.",
        source=pending_uid,
    )


def _require_editable_company_config(paths: DashboardPaths) -> dict:
    if not paths.editable_config_path.exists():
        raise DashboardOperationError(
            "config_empresa_ausente_para_correcao",
            "A configuracao editavel da empresa nao esta materializada neste run.",
            source=str(paths.editable_config_path),
        )

    payload = _load_company_config_payload(paths.editable_config_path)
    if not _stringify(payload.get("company_code")):
        raise DashboardOperationError(
            "config_empresa_sem_codigo",
            "A configuracao editavel nao possui codigo de empresa.",
            source=str(paths.editable_config_path),
        )
    return payload


def _persist_employee_mapping_to_registry(
    *,
    company_code: str,
    company_name: str | None,
    employee_key: str,
    employee_name: str,
    domain_registration: str,
    aliases: list[str],
    root: str | Path | None,
) -> Path:
    try:
        registry = load_company_employee_registry(
            company_code,
            company_name=company_name,
            root=root,
        )
        employee = CompanyEmployeeRecord(
            employee_key=employee_key,
            employee_name=employee_name,
            domain_registration=domain_registration,
            aliases=aliases,
            status="active",
            source="dashboard_manual_action",
        )
        updated_registry = upsert_employee_record(registry, employee)
        return save_company_employee_registry(updated_registry, root=root)
    except Exception as exc:
        raise DashboardOperationError(
            "cadastro_funcionario_invalido",
            f"Nao foi possivel salvar o cadastro persistente do funcionario: {exc}",
            source=company_code,
        ) from exc


def _persist_event_mapping_to_rubric_catalog(
    *,
    company_code: str,
    company_name: str | None,
    rubric_code: str,
    description: str,
    canonical_event: str,
    value_kind: str,
    nature: str,
    aliases: list[str],
    root: str | Path | None,
) -> Path:
    try:
        catalog = load_company_rubric_catalog(
            company_code,
            company_name=company_name,
            root=root,
        )
        rubric = CompanyRubricRecord(
            rubric_code=rubric_code,
            description=description,
            canonical_event=canonical_event,
            value_kind=value_kind,
            nature=nature,
            aliases=aliases,
            status="active",
            source="dashboard_manual_action",
        )
        updated_catalog = upsert_rubric_record(catalog, rubric)
        return save_company_rubric_catalog(updated_catalog, root=root)
    except Exception as exc:
        raise DashboardOperationError(
            "catalogo_rubrica_invalido",
            f"Nao foi possivel salvar o catalogo persistente de rubrica: {exc}",
            source=company_code,
        ) from exc


def _aliases_from_payload(payload: Mapping[str, Any]) -> list[str]:
    raw_aliases = payload.get("aliases", [])
    if raw_aliases is None:
        return []
    if not isinstance(raw_aliases, list):
        raise DashboardOperationError(
            "aliases_invalidos",
            "Aliases de funcionario devem ser informados como lista.",
            source="aliases",
        )
    return [str(alias).strip() for alias in raw_aliases if str(alias).strip()]


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "sim", "s", "yes", "y"}


def _load_company_config_payload(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise DashboardOperationError(
            "config_empresa_ausente_para_correcao",
            "A configuracao editavel da empresa nao foi encontrada para edicao guiada.",
            source=str(path),
        ) from exc
    except json.JSONDecodeError as exc:
        raise DashboardOperationError(
            "config_empresa_invalida_no_dashboard",
            "A configuracao da empresa esta invalida para edicao guiada.",
            source=str(path),
        ) from exc


def _write_validated_company_config(path: Path, payload: dict) -> None:
    config = CompanyConfig.model_validate(payload)
    path.write_text(
        json.dumps(config.model_dump(mode="json"), ensure_ascii=True, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _stringify(value) -> str | None:
    if value is None:
        return None
    return str(value)


def _required_text_from_payload(
    payload: Mapping[str, Any],
    field_name: str,
    *,
    aliases: tuple[str, ...] = (),
) -> str:
    for key in (field_name, *aliases):
        if key in payload:
            return _required_text(field_name, payload[key])

    raise DashboardOperationError(
        "campo_obrigatorio_ausente",
        f"Campo obrigatorio ausente para acao manual: {field_name}.",
        source=field_name,
    )


def _required_text(field_name: str, value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        raise DashboardOperationError(
            "campo_obrigatorio_ausente",
            f"Campo obrigatorio ausente para acao manual: {field_name}.",
            source=field_name,
        )
    return text
