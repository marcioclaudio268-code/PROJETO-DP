"""Profile-driven conversion from uploaded workbooks into the V1 canonical workbook."""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain import NormalizedHours, decimal_to_plain_string
from ingestion import (
    InputColumnMetadata,
    InputLayoutDetection,
    InputLayoutNormalizationError,
    InputNormalizationResult,
    InputWorkbookInspection,
    NormalizationError,
    is_empty_value,
    normalize_hours_hhmm,
    normalize_money_brl,
    normalize_quantity,
    normalized_optional_text,
)
from ingestion.input_layout import (
    DEFAULT_CANONICAL_LAYOUT_VERSION,
    DEFAULT_CANONICAL_PAYROLL_TYPE,
    MONTHLY_LAYOUT_ID,
)
from ingestion.template_v1 import MAX_DATA_ROWS, create_planilha_padrao_folha_v1

from .column_mapping_profiles import (
    ColumnGenerationMode,
    ColumnMappingRule,
    ColumnValueKind,
    CompanyColumnMappingProfile,
    excel_column_to_index,
)
from .company_employee_registry import (
    find_employee_by_name_or_alias,
    load_company_employee_registry,
)
from .company_rubric_catalog import find_rubric_by_code, load_company_rubric_catalog


RUBRIC_TO_CANONICAL_EVENT_COLUMN = {
    "20": "gratificacao",
    "201": "horas_extras_70",
    "219": "horas_extras_70",
    "150": "horas_extras_50",
    "350": "horas_extras_50",
    "200": "horas_extras_100",
    "25": "hora_extra_noturna",
    "8069": "atrasos_horas",
    "8792": "faltas_dias",
    "8794": "faltas_dsr",
    "266": "mercadoria",
    "204": "mercadoria",
    "202": "mercadoria",
    "264": "mercadoria",
    "981": "desconto_adiantamento",
    "48": "vale_transporte",
}

CANONICAL_EVENT_COLUMNS = {
    "horas_extras_50": "G",
    "gratificacao": "H",
    "vale_transporte": "N",
    "mercadoria": "P",
    "faltas_dias": "R",
    "atrasos_horas": "S",
    "desconto_adiantamento": "T",
    "horas_extras_70": "V",
    "horas_extras_100": "W",
    "hora_extra_noturna": "X",
    "faltas_dsr": "Y",
}

STATUS_TOKENS = {
    "ferias": "ferias",
    "afastado": "afastado",
    "licenca": "afastado",
    "rescisao": "rescindido",
    "ignorar": "ignorar",
}
IDENTITY_COLUMN_TOKENS = {
    "matricula",
    "matricula dominio",
    "registro",
    "registro dominio",
}
FALLBACK_IDENTITY_COLUMN_TOKENS = {
    "cod",
    "codigo",
    "codigo dominio",
}


@dataclass(frozen=True, slots=True)
class _ParsedProfileValue:
    value_for_workbook: str
    is_zero: bool


@dataclass(frozen=True, slots=True)
class _ResolvedProfileRubric:
    rubric_code: str
    event_name: str
    nature: str


def normalize_workbook_with_column_profile(
    input_path: str | Path,
    *,
    inspection: InputWorkbookInspection,
    profile: CompanyColumnMappingProfile,
    employee_registry_root: str | Path | None = None,
    rubric_catalog_root: str | Path | None = None,
    output_path: str | Path | None = None,
    report_path: str | Path | None = None,
    max_data_rows: int = MAX_DATA_ROWS,
) -> InputNormalizationResult:
    """Convert an inspected non-canonical workbook into the V1 canonical workbook."""

    source_path = Path(input_path)
    workbook = load_workbook(source_path)
    normalized_workbook, manifest = build_canonical_workbook_from_column_profile(
        workbook,
        inspection=inspection,
        profile=profile,
        employee_registry_root=employee_registry_root,
        rubric_catalog_root=rubric_catalog_root,
        max_data_rows=max_data_rows,
    )

    target_path = Path(output_path) if output_path is not None else source_path
    target_path.parent.mkdir(parents=True, exist_ok=True)
    normalized_workbook.save(target_path)

    report_target = Path(report_path) if report_path is not None else None
    result = InputNormalizationResult(
        source_path=source_path,
        workbook_path=target_path,
        report_path=report_target,
        layout=inspection.detection,
        canonical_rows_written=int(manifest["counts"]["canonical_rows_written"]),
        employee_rows_written=int(manifest["counts"]["employee_rows_written"]),
        supported_cells_written=int(manifest["counts"]["source_cells_converted"]),
        unsupported_cells_preserved=int(
            manifest["counts"]["ignored_cells"] + manifest["counts"]["skipped_text_cells"]
        ),
        orphan_note_rows=0,
        manifest=manifest,
    )
    if report_target is not None:
        report_target.parent.mkdir(parents=True, exist_ok=True)
        report_target.write_text(
            json.dumps(result.as_dict(), ensure_ascii=True, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
    return result


def build_canonical_workbook_from_column_profile(
    workbook: Workbook,
    *,
    inspection: InputWorkbookInspection,
    profile: CompanyColumnMappingProfile,
    employee_registry_root: str | Path | None = None,
    rubric_catalog_root: str | Path | None = None,
    max_data_rows: int = MAX_DATA_ROWS,
) -> tuple[Workbook, dict]:
    if inspection.layout_id == "template_v1_canonico":
        raise InputLayoutNormalizationError(
            "perfil_colunas_nao_requerido",
            "Layout canonico V1 nao deve ser convertido por perfil de colunas.",
            source=inspection.selected_sheet_name,
        )
    if profile.company_code != inspection.company_code:
        raise InputLayoutNormalizationError(
            "perfil_empresa_divergente",
            "Perfil de colunas pertence a empresa diferente da planilha inspecionada.",
            source=inspection.selected_sheet_name,
            details={"inspection_company_code": inspection.company_code, "profile_company_code": profile.company_code},
        )
    if inspection.selected_sheet_name not in workbook.sheetnames:
        raise InputLayoutNormalizationError(
            "aba_detectada_ausente",
            f"A aba detectada '{inspection.selected_sheet_name}' nao existe no workbook.",
            source=inspection.selected_sheet_name,
        )

    selected_sheet = workbook[inspection.selected_sheet_name]
    _validate_position_profile_headers(selected_sheet, profile)
    normalized_workbook = create_planilha_padrao_folha_v1(max_data_rows=max_data_rows)
    _populate_parameters(normalized_workbook, inspection, profile)

    funcionarios = normalized_workbook["FUNCIONARIOS"]
    lancamentos = normalized_workbook["LANCAMENTOS_FACEIS"]
    movimentos = normalized_workbook["MOVIMENTOS_CANONICOS"]
    column_mappings = _match_profile_mappings_to_columns(profile, inspection.columns)
    employee_registry = load_company_employee_registry(
        profile.company_code,
        company_name=profile.company_name,
        root=employee_registry_root,
    )
    rubric_catalog = load_company_rubric_catalog(
        profile.company_code,
        company_name=profile.company_name,
        root=rubric_catalog_root,
    )
    employee_code_column_index = _profile_employee_code_column_index(profile)
    employee_name_column_index = _profile_employee_name_column_index(profile)
    identity_column = None if employee_code_column_index is not None else _find_domain_registration_column(inspection.columns)

    employee_rows_written = 0
    canonical_rows_written = 0
    source_cells_converted = 0
    ignored_cells = 0
    skipped_zero_cells = 0
    skipped_text_cells = 0
    generated_movements = 0

    header_row = min((column.header_row for column in inspection.columns), default=4)
    data_start_row = _profile_data_start_row(profile) or header_row + 1
    employee_code_index = employee_code_column_index or 1
    employee_name_index = employee_name_column_index or 2
    for source_row in range(data_start_row, selected_sheet.max_row + 1):
        row_values = {
            column_index: selected_sheet.cell(row=source_row, column=column_index).value
            for column_index in range(1, selected_sheet.max_column + 1)
        }
        if _row_is_empty(row_values):
            continue

        employee_key = _profile_employee_key(
            row_values=row_values,
            employee_code_index=employee_code_index,
            employee_name_index=employee_name_index,
            employee_code_column_index=employee_code_column_index,
        )
        employee_name = normalized_optional_text(row_values.get(employee_name_index))
        snapshot_registration = (
            normalized_optional_text(row_values.get(employee_code_column_index))
            if employee_code_column_index is not None
            else (
            normalized_optional_text(row_values.get(identity_column.column_index))
            if identity_column is not None
            else None
            )
        )
        if employee_key is None and employee_name is None:
            if _row_has_profile_data(row_values, column_mappings):
                raise InputLayoutNormalizationError(
                    "linha_orfa_com_dado_critico",
                    "Linha com coluna mapeada preenchida nao possui identificacao de colaborador.",
                    source=f"{selected_sheet.title}!A{source_row}",
                )
            continue

        if employee_code_column_index is not None or employee_name_column_index is not None:
            resolved_employee_name, domain_registration = _resolve_profile_employee_identity(
                employee_registry=employee_registry,
                employee_name=employee_name,
                snapshot_registration=snapshot_registration,
                selected_sheet=selected_sheet,
                source_row=source_row,
                employee_code_column_index=employee_code_column_index,
                employee_name_column_index=employee_name_column_index,
            )
            if resolved_employee_name is not None:
                employee_name = resolved_employee_name
        else:
            domain_registration = snapshot_registration

        employee_status, employee_note = _infer_employee_status(row_values)
        employee_rows_written += 1
        _write_employee_row(
            funcionarios,
            employee_rows_written + 1,
            employee_key=employee_key,
            employee_name=employee_name,
            domain_registration=domain_registration,
            status_colaborador=employee_status,
            note=employee_note,
            source_sheet=selected_sheet.title,
            source_row=source_row,
        )

        for column, mapping in column_mappings:
            raw_value = row_values.get(column.column_index)
            if is_empty_value(raw_value):
                continue
            if mapping.generation_mode == ColumnGenerationMode.IGNORE:
                ignored_cells += 1
                continue

            parsed_value = _parse_profile_value(
                raw_value,
                mapping=mapping,
                source=f"{selected_sheet.title}!{column.column_letter}{source_row}",
            )
            if parsed_value is None:
                skipped_text_cells += 1
                continue
            if parsed_value.is_zero and mapping.ignore_zero:
                skipped_zero_cells += 1
                continue

            source_cells_converted += 1
            target_rubrics = _resolve_profile_target_rubrics(
                mapping,
                source=f"{selected_sheet.title}!{column.column_letter}{source_row}",
                rubric_catalog=rubric_catalog,
            )
            for resolved_rubric in target_rubrics:
                canonical_rows_written += 1
                generated_movements += 1
                observation = _build_launch_observation(
                    inspection=inspection,
                    source_row=source_row,
                    source_column=column,
                    target_rubric=resolved_rubric.rubric_code,
                    profile=profile,
                )
                if mapping.value_column is not None:
                    _write_direct_movement_row(
                        movimentos,
                        canonical_rows_written + 1,
                        movement_id=f"profile-mov-{canonical_rows_written:05d}",
                        company_code=inspection.company_code,
                        competence=inspection.competence,
                        default_process=profile.default_process or "11",
                        employee_key=employee_key,
                        employee_name=employee_name,
                        domain_registration=domain_registration,
                        event_name=resolved_rubric.event_name,
                        informed_rubric=resolved_rubric.rubric_code,
                        output_rubric=resolved_rubric.rubric_code,
                        event_nature=resolved_rubric.nature,
                        value_kind=mapping.value_kind,
                        value=parsed_value.value_for_workbook,
                        source_sheet=selected_sheet.title,
                        source_cell=f"{column.column_letter}{source_row}",
                        source_column_name=column.column_name,
                        observation=observation,
                    )
                else:
                    event_column = _canonical_event_column_for_rubric(
                        resolved_rubric.rubric_code,
                        source=f"{selected_sheet.title}!{column.column_letter}{source_row}",
                    )
                    _write_launch_row(
                        lancamentos,
                        canonical_rows_written + 1,
                        event_column=event_column,
                        value=parsed_value.value_for_workbook,
                        employee_key=employee_key,
                        employee_name=employee_name,
                        domain_registration=domain_registration,
                        observation=observation,
                    )

    manifest = {
        "layout_id": inspection.layout_id,
        "normalizer": "profile_column_mapping",
        "selected_sheet_name": inspection.selected_sheet_name,
        "selected_sheet_reason": inspection.detection.selected_sheet_reason,
        "company_code": inspection.company_code,
        "company_name": inspection.company_name,
        "competence": inspection.competence,
        "source_company_name": inspection.detection.source_company_name,
        "source_title_text": inspection.detection.source_title_text,
        "source_sheet_names": list(inspection.source_sheet_names),
        "rules_applied": [*inspection.detection.rules_applied, "profile_column_mapping"],
        "warnings": list(inspection.warnings),
        "profile": {
            "company_code": profile.company_code,
            "company_name": profile.company_name,
            "profile_version": profile.profile_version,
            "default_process": profile.default_process,
        },
        "columns": [
            {
                "source_column": column.column_name,
                "source_column_letter": column.column_letter,
                "generation_mode": mapping.generation_mode.value,
                "target_rubrics": list(mapping.target_rubrics),
                "value_kind": mapping.value_kind.value,
            }
            for column, mapping in column_mappings
        ],
        "counts": {
            "employee_rows_written": employee_rows_written,
            "canonical_rows_written": canonical_rows_written,
            "source_cells_converted": source_cells_converted,
            "generated_movements": generated_movements,
            "ignored_cells": ignored_cells,
            "skipped_zero_cells": skipped_zero_cells,
            "skipped_text_cells": skipped_text_cells,
        },
        "row_warnings": [],
    }
    return normalized_workbook, manifest


def _match_profile_mappings_to_columns(
    profile: CompanyColumnMappingProfile,
    columns: tuple[InputColumnMetadata, ...],
) -> tuple[tuple[InputColumnMetadata, ColumnMappingRule], ...]:
    mapping_index: dict[str, ColumnMappingRule] = {}
    for mapping in profile.mappings:
        if not mapping.is_active:
            continue
        for token in (mapping.column_key, mapping.column_name, mapping.value_column):
            if token:
                mapping_index[_normalize_token(token)] = mapping

    matches: list[tuple[InputColumnMetadata, ColumnMappingRule]] = []
    for column in columns:
        tokens = (
            column.column_name,
            column.normalized_column_name,
            column.column_letter,
            str(column.column_index),
        )
        mapping = next((mapping_index[_normalize_token(token)] for token in tokens if _normalize_token(token) in mapping_index), None)
        if mapping is not None:
            matches.append((column, mapping))
    return tuple(matches)


def inspect_workbook_with_position_profile(
    input_path: str | Path,
    *,
    profile: CompanyColumnMappingProfile,
    selected_company_code: str | None = None,
    selected_company_name: str | None = None,
    selected_competence: str | None = None,
) -> InputWorkbookInspection:
    workbook = load_workbook(input_path, data_only=False)
    selected_sheet = _select_position_profile_sheet(workbook, profile)
    active_position_rules = tuple(rule for rule in profile.mappings if rule.is_active and rule.value_column)
    header_row = _profile_header_row(profile) or 1
    columns: list[InputColumnMetadata] = []
    for rule in active_position_rules:
        assert rule.value_column is not None
        column_index = excel_column_to_index(rule.value_column)
        actual_header = normalized_optional_text(selected_sheet.cell(row=rule.header_row or header_row, column=column_index).value)
        column_name = actual_header or rule.expected_header or rule.column_name or rule.value_column
        columns.append(
            InputColumnMetadata(
                column_index=column_index,
                column_letter=rule.value_column,
                column_name=column_name,
                normalized_column_name=_normalize_token(column_name),
                header_row=rule.header_row or header_row,
            )
        )

    detection = InputLayoutDetection(
        layout_id=MONTHLY_LAYOUT_ID,
        active_sheet_name=workbook.active.title if workbook.active is not None else None,
        selected_sheet_name=selected_sheet.title,
        selected_sheet_reason="position_column_profile",
        company_code=selected_company_code or profile.company_code,
        company_name=selected_company_name or profile.company_name or "",
        competence=selected_competence or _infer_competence_from_sheet_name(selected_sheet.title) or "",
        source_company_name=selected_company_name or profile.company_name or "",
        source_title_text=normalized_optional_text(selected_sheet["A1"].value),
        source_sheet_names=tuple(workbook.sheetnames),
        rules_applied=("position_column_profile",),
        warnings=(),
    )
    return InputWorkbookInspection(
        layout_id=detection.layout_id,
        company_code=detection.company_code,
        company_name=detection.company_name,
        competence=detection.competence,
        selected_sheet_name=detection.selected_sheet_name,
        source_sheet_names=detection.source_sheet_names,
        columns=tuple(columns),
        warnings=detection.warnings,
        detection=detection,
    )


def is_profile_identity_column(column: InputColumnMetadata) -> bool:
    return _normalize_token(column.column_name) in IDENTITY_COLUMN_TOKENS


def _find_domain_registration_column(
    columns: tuple[InputColumnMetadata, ...],
) -> InputColumnMetadata | None:
    explicit_column = next((column for column in columns if is_profile_identity_column(column)), None)
    if explicit_column is not None:
        return explicit_column
    return next(
        (
            column
            for column in columns
            if _normalize_token(column.column_name) in FALLBACK_IDENTITY_COLUMN_TOKENS
        ),
        None,
    )


def _profile_header_row(profile: CompanyColumnMappingProfile) -> int | None:
    rows = [rule.header_row for rule in profile.mappings if rule.is_active and rule.header_row is not None]
    return min(rows) if rows else None


def _profile_data_start_row(profile: CompanyColumnMappingProfile) -> int | None:
    rows = [rule.data_start_row for rule in profile.mappings if rule.is_active and rule.data_start_row is not None]
    return min(rows) if rows else None


def _profile_employee_code_column_index(profile: CompanyColumnMappingProfile) -> int | None:
    for rule in profile.mappings:
        if rule.is_active and rule.employee_code_column:
            return excel_column_to_index(rule.employee_code_column)
    return None


def _profile_employee_name_column_index(profile: CompanyColumnMappingProfile) -> int | None:
    for rule in profile.mappings:
        if rule.is_active and rule.employee_name_column:
            return excel_column_to_index(rule.employee_name_column)
    return None


def _profile_employee_key(
    *,
    row_values: dict[int, object],
    employee_code_index: int,
    employee_name_index: int,
    employee_code_column_index: int | None,
) -> str | None:
    value = normalized_optional_text(row_values.get(employee_code_index))
    if value is not None:
        return value
    return normalized_optional_text(row_values.get(employee_name_index))


def _resolve_profile_employee_identity(
    *,
    employee_registry,
    employee_name: str | None,
    snapshot_registration: str | None,
    selected_sheet: Worksheet,
    source_row: int,
    employee_code_column_index: int | None,
    employee_name_column_index: int | None,
) -> tuple[str | None, str]:
    if _is_valid_domain_registration(snapshot_registration):
        return employee_name, str(snapshot_registration).strip()

    if employee_name_column_index is None:
        raise InputLayoutNormalizationError(
            "perfil_colunas_sem_nome_para_resolver_funcionario",
            "Linha sem matricula valida e sem coluna de nome configurada no perfil.",
            source=f"{selected_sheet.title}!A{source_row}",
        )

    normalized_name = normalized_optional_text(employee_name)
    if normalized_name is None:
        raise InputLayoutNormalizationError(
            "funcionario_nome_nao_encontrado",
            "Funcionario '' nao encontrado no cadastro ativo da empresa. Cadastre ou revise o nome antes de gerar o TXT.",
            source=f"{selected_sheet.title}!{get_column_letter(employee_name_column_index)}{source_row}",
        )

    matches = find_employee_by_name_or_alias(employee_registry, normalized_name, active_only=True)
    if not matches:
        raise InputLayoutNormalizationError(
            "funcionario_nome_nao_encontrado",
            (
                f"Funcionario '{normalized_name}' nao encontrado no cadastro ativo da empresa. "
                "Cadastre ou revise o nome antes de gerar o TXT."
            ),
            source=f"{selected_sheet.title}!{get_column_letter(employee_name_column_index)}{source_row}",
        )
    if len(matches) > 1:
        raise InputLayoutNormalizationError(
            "funcionario_nome_ambiguo",
            (
                f"Funcionario '{normalized_name}' encontrou mais de um cadastro compativel. "
                "Revise o cadastro antes de gerar o TXT."
            ),
            source=f"{selected_sheet.title}!{get_column_letter(employee_name_column_index)}{source_row}",
        )

    employee = matches[0]
    return employee.employee_name, employee.domain_registration


def _is_valid_domain_registration(value: str | None) -> bool:
    if value is None:
        return False
    return bool(re.fullmatch(r"\d+", str(value).strip()))


def _select_position_profile_sheet(workbook: Workbook, profile: CompanyColumnMappingProfile) -> Worksheet:
    configured_names = [
        rule.sheet_name
        for rule in profile.mappings
        if rule.is_active and rule.sheet_name
    ]
    for configured_name in configured_names:
        for worksheet in workbook.worksheets:
            if _normalize_token(worksheet.title) == _normalize_token(configured_name):
                return worksheet
        raise InputLayoutNormalizationError(
            "aba_perfil_colunas_ausente",
            f"A aba configurada no perfil de colunas nao existe no workbook: {configured_name}.",
            source=configured_name,
        )
    if workbook.active is not None:
        return workbook.active
    return workbook.worksheets[0]


def _validate_position_profile_headers(
    worksheet: Worksheet,
    profile: CompanyColumnMappingProfile,
) -> None:
    for rule in profile.mappings:
        if not rule.is_active or not rule.value_column or not rule.expected_header:
            continue
        if rule.sheet_name and _normalize_token(rule.sheet_name) != _normalize_token(worksheet.title):
            continue
        column_index = excel_column_to_index(rule.value_column)
        actual = normalized_optional_text(worksheet.cell(row=rule.header_row or 1, column=column_index).value)
        expected = rule.expected_header
        if _header_matches_expected(actual, expected):
            continue
        found = actual or "[vazio]"
        raise InputLayoutNormalizationError(
            "cabecalho_perfil_divergente",
            (
                f"A coluna {rule.value_column} esperava {expected}, mas encontrou {found}. "
                "Revise o perfil de colunas."
            ),
            source=f"{worksheet.title}!{rule.value_column}{rule.header_row or 1}",
            details={
                "expected_header": expected,
                "actual_header": actual,
                "value_column": rule.value_column,
            },
        )


def _header_matches_expected(actual: str | None, expected: str) -> bool:
    actual_token = _normalize_token(actual or "")
    expected_token = _normalize_token(expected)
    if not actual_token or not expected_token:
        return False
    return expected_token in actual_token or actual_token in expected_token


def _infer_competence_from_sheet_name(sheet_name: str) -> str | None:
    match = re.search(r"\b(0?[1-9]|1[0-2])[-_/ ]?(20\d{2}|\d{2})\b", sheet_name)
    if match:
        month = int(match.group(1))
        year = match.group(2)
        if len(year) == 2:
            year = f"20{year}"
        return f"{month:02d}/{year}"

    month_names = {
        "jan": 1,
        "janeiro": 1,
        "fev": 2,
        "fevereiro": 2,
        "mar": 3,
        "marco": 3,
        "abr": 4,
        "abril": 4,
        "mai": 5,
        "maio": 5,
        "jun": 6,
        "junho": 6,
        "jul": 7,
        "julho": 7,
        "ago": 8,
        "agosto": 8,
        "set": 9,
        "setembro": 9,
        "out": 10,
        "outubro": 10,
        "nov": 11,
        "novembro": 11,
        "dez": 12,
        "dezembro": 12,
    }
    normalized = _normalize_token(sheet_name)
    for name, month in month_names.items():
        year_match = re.search(rf"\b{name}\b\D*(20\d{{2}}|\d{{2}})", normalized)
        if year_match:
            year = year_match.group(1)
            if len(year) == 2:
                year = f"20{year}"
            return f"{month:02d}/{year}"
    return None


def _parse_profile_value(
    value: object,
    *,
    mapping: ColumnMappingRule,
    source: str,
) -> _ParsedProfileValue | None:
    try:
        if mapping.value_kind == ColumnValueKind.MONETARY:
            amount = normalize_money_brl(value)
            return _ParsedProfileValue(decimal_to_plain_string(amount), amount == Decimal("0"))
        if mapping.value_kind == ColumnValueKind.QUANTITY:
            quantity = normalize_quantity(value)
            return _ParsedProfileValue(decimal_to_plain_string(quantity), quantity == Decimal("0"))
        if mapping.value_kind == ColumnValueKind.HOURS:
            hours = _normalize_profile_hours(value)
            return _ParsedProfileValue(hours.text, hours.total_minutes == 0)
    except NormalizationError as exc:
        if mapping.ignore_text:
            return None
        raise InputLayoutNormalizationError(
            "valor_perfil_invalido",
            f"Valor nao pode ser convertido pela regra da coluna '{mapping.source_column_id}': {exc}",
            source=source,
        ) from exc

    raise InputLayoutNormalizationError(
        "tipo_valor_perfil_nao_suportado",
        f"Tipo de valor nao suportado no perfil: {mapping.value_kind}.",
        source=source,
    )


def _normalize_profile_hours(value: object) -> NormalizedHours:
    if isinstance(value, int):
        return normalize_hours_hhmm(f"{value:02d}:00")
    if isinstance(value, float) and value.is_integer():
        return normalize_hours_hhmm(f"{int(value):02d}:00")
    return normalize_hours_hhmm(value)


def _resolve_profile_target_rubrics(
    mapping: ColumnMappingRule,
    *,
    source: str,
    rubric_catalog,
) -> tuple[_ResolvedProfileRubric, ...]:
    if mapping.generation_mode == ColumnGenerationMode.SINGLE_LINE and not mapping.rubrica_target:
        raise InputLayoutNormalizationError(
            "perfil_coluna_sem_rubrica_unica",
            f"Regra de perfil da coluna {mapping.value_column or mapping.source_column_id} nao possui rubrica unica configurada.",
            source=source,
        )
    if mapping.generation_mode == ColumnGenerationMode.MULTI_LINE and len(mapping.rubricas_target) == 0:
        raise InputLayoutNormalizationError(
            "perfil_coluna_sem_rubricas_multiplas",
            f"Regra de perfil da coluna {mapping.value_column or mapping.source_column_id} nao possui rubricas multiplas configuradas.",
            source=source,
        )

    resolved: list[_ResolvedProfileRubric] = []
    if mapping.value_column is None:
        for target_rubric in mapping.target_rubrics:
            resolved.append(
                _ResolvedProfileRubric(
                    rubric_code=target_rubric,
                    event_name=target_rubric,
                    nature=mapping.nature.value,
                )
            )
        return tuple(resolved)

    for target_rubric in mapping.target_rubrics:
        matches = find_rubric_by_code(rubric_catalog, target_rubric, active_only=True)
        if not matches:
            raise InputLayoutNormalizationError(
                "rubrica_perfil_inexistente_no_catalogo",
                f"Rubrica {target_rubric} nao existe no catalogo ativo da empresa.",
                source=source,
            )
        rubric = matches[0]
        resolved.append(
            _ResolvedProfileRubric(
                rubric_code=rubric.rubric_code,
                event_name=rubric.rubric_code,
                nature=rubric.nature.value,
            )
        )
    return tuple(resolved)


def _canonical_event_column_for_rubric(target_rubric: str, *, source: str) -> str:
    normalized_rubric = str(target_rubric).strip()
    event_name = RUBRIC_TO_CANONICAL_EVENT_COLUMN.get(normalized_rubric)
    if event_name is None:
        raise InputLayoutNormalizationError(
            "rubrica_sem_coluna_canonica",
            f"Rubrica '{target_rubric}' nao possui coluna canonica V1 suportada nesta rodada.",
            source=source,
        )
    return CANONICAL_EVENT_COLUMNS[event_name]


def _populate_parameters(
    workbook: Workbook,
    inspection: InputWorkbookInspection,
    profile: CompanyColumnMappingProfile,
) -> None:
    worksheet = workbook["PARAMETROS"]
    _set_parameter_value(worksheet, "empresa_codigo", inspection.company_code)
    _set_parameter_value(worksheet, "empresa_nome", inspection.company_name)
    _set_parameter_value(worksheet, "competencia", inspection.competence)
    _set_parameter_value(worksheet, "tipo_folha", DEFAULT_CANONICAL_PAYROLL_TYPE)
    _set_parameter_value(worksheet, "processo_padrao", profile.default_process or "11")
    _set_parameter_value(worksheet, "versao_layout", DEFAULT_CANONICAL_LAYOUT_VERSION)
    _set_parameter_value(worksheet, "responsavel_preenchimento", "profile_column_mapping")
    _set_parameter_value(worksheet, "observacoes_gerais", f"normalizado_por_perfil; layout={inspection.layout_id}")


def _set_parameter_value(worksheet: Worksheet, field_name: str, value: object) -> None:
    for row_number in range(2, worksheet.max_row + 1):
        current = normalized_optional_text(worksheet.cell(row=row_number, column=1).value)
        if current == field_name:
            worksheet.cell(row=row_number, column=2, value=value)
            return
    raise InputLayoutNormalizationError(
        "parametro_interno_ausente",
        f"Nao foi possivel localizar o parametro interno '{field_name}' no workbook canonico.",
        source=worksheet.title,
    )


def _write_employee_row(
    worksheet: Worksheet,
    row_number: int,
    *,
    employee_key: str | None,
    employee_name: str | None,
    domain_registration: str | None,
    status_colaborador: str,
    note: str | None,
    source_sheet: str,
    source_row: int,
) -> None:
    values = {
        "A": employee_key,
        "B": employee_name,
        "E": domain_registration,
        "H": status_colaborador,
        "I": "nao" if status_colaborador == "ignorar" else "sim",
        "J": note or f"layout={MONTHLY_LAYOUT_ID}; aba={source_sheet}; linha_origem={source_row}",
    }
    for column_letter, cell_value in values.items():
        worksheet[f"{column_letter}{row_number}"] = cell_value


def _write_launch_row(
    worksheet: Worksheet,
    row_number: int,
    *,
    event_column: str,
    value: str,
    employee_key: str | None,
    employee_name: str | None,
    domain_registration: str | None,
    observation: str,
) -> None:
    worksheet[f"B{row_number}"] = employee_key
    worksheet[f"C{row_number}"] = employee_name
    worksheet[f"D{row_number}"] = domain_registration
    worksheet[f"F{row_number}"] = observation
    worksheet[f"{event_column}{row_number}"] = value


def _write_direct_movement_row(
    worksheet: Worksheet,
    row_number: int,
    *,
    movement_id: str,
    company_code: str,
    competence: str,
    default_process: str,
    employee_key: str | None,
    employee_name: str | None,
    domain_registration: str | None,
    event_name: str,
    informed_rubric: str,
    output_rubric: str,
    event_nature: str,
    value_kind: ColumnValueKind,
    value: str,
    source_sheet: str,
    source_cell: str,
    source_column_name: str,
    observation: str,
) -> None:
    worksheet[f"A{row_number}"] = movement_id
    worksheet[f"B{row_number}"] = company_code
    worksheet[f"C{row_number}"] = competence
    worksheet[f"D{row_number}"] = DEFAULT_CANONICAL_PAYROLL_TYPE
    worksheet[f"E{row_number}"] = default_process
    worksheet[f"F{row_number}"] = employee_key
    worksheet[f"G{row_number}"] = employee_name
    worksheet[f"H{row_number}"] = domain_registration
    worksheet[f"I{row_number}"] = event_name
    worksheet[f"J{row_number}"] = informed_rubric
    worksheet[f"K{row_number}"] = output_rubric
    worksheet[f"L{row_number}"] = event_nature
    worksheet[f"M{row_number}"] = "dias" if value_kind == ColumnValueKind.QUANTITY else value_kind.value
    if value_kind == ColumnValueKind.MONETARY:
        worksheet[f"O{row_number}"] = value
        worksheet[f"P{row_number}"] = "BRL"
    elif value_kind == ColumnValueKind.HOURS:
        worksheet[f"N{row_number}"] = value
        worksheet[f"P{row_number}"] = "HH:MM"
    else:
        worksheet[f"N{row_number}"] = value
        worksheet[f"P{row_number}"] = "DIAS"
    worksheet[f"Q{row_number}"] = source_sheet
    worksheet[f"R{row_number}"] = source_cell
    worksheet[f"S{row_number}"] = source_column_name
    worksheet[f"T{row_number}"] = "nao"
    worksheet[f"W{row_number}"] = observation


def _build_launch_observation(
    *,
    inspection: InputWorkbookInspection,
    source_row: int,
    source_column: InputColumnMetadata,
    target_rubric: str,
    profile: CompanyColumnMappingProfile,
) -> str:
    return (
        f"layout={inspection.layout_id}; "
        f"aba={inspection.selected_sheet_name}; "
        f"linha_origem={source_row}; "
        f"coluna_origem={source_column.column_name}; "
        f"rubrica_target={target_rubric}; "
        f"profile_version={profile.profile_version}"
    )


def _infer_employee_status(row_values: dict[int, object]) -> tuple[str, str | None]:
    for column_index, value in row_values.items():
        if column_index in {1, 2}:
            continue
        text = normalized_optional_text(value)
        if text is None:
            continue
        normalized = _normalize_token(text)
        for token, status in STATUS_TOKENS.items():
            if token in normalized:
                return status, f"status_origem={token}"
    return "ativo", None


def _row_has_profile_data(
    row_values: dict[int, object],
    column_mappings: tuple[tuple[InputColumnMetadata, ColumnMappingRule], ...],
) -> bool:
    return any(not is_empty_value(row_values.get(column.column_index)) for column, _mapping in column_mappings)


def _row_is_empty(row_values: dict[int, object]) -> bool:
    return all(is_empty_value(value) for value in row_values.values())


def _normalize_token(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value)).strip().lower()
    text = "".join(character for character in text if not unicodedata.combining(character))
    cleaned = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(cleaned.split())


__all__ = [
    "build_canonical_workbook_from_column_profile",
    "inspect_workbook_with_position_profile",
    "is_profile_identity_column",
    "normalize_workbook_with_column_profile",
]
