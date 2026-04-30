"""Operational orchestration for the guided local dashboard."""

from __future__ import annotations

import json
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from mapping.pipeline import map_snapshot_with_company_config
from serialization.pipeline import serialize_mapped_artifact_to_txt
from validation.pipeline import validate_pipeline_v1

from ingestion.input_layout import (
    CANONICAL_LAYOUT_ID,
    InputColumnMetadata,
    InputWorkbookInspection,
    inspect_input_workbook,
    normalize_input_workbook,
)
from ingestion.pipeline import ingest_fill_and_persist_planilha_padrao_v1

from .column_mapping_profiles import (
    ColumnMappingProfileError,
    CompanyColumnMappingProfile,
    column_mapping_profile_path,
    load_column_mapping_profile,
)
from .company_employee_registry import apply_employee_registry_to_editable_config
from .company_rubric_catalog import apply_rubric_catalog_to_editable_config
from .config_resolver import ConfigResolutionResult, ConfigResolutionStatus, ConfigResolver
from .models import (
    DashboardConfigResolution,
    DashboardPaths,
    DashboardPendingItem,
    DashboardProfileResolution,
    DashboardRunResult,
    DashboardSummary,
)
from .overrides import describe_ignore_strategy, replay_dashboard_action_overrides
from .profile_normalizer import is_profile_identity_column, normalize_workbook_with_column_profile
from .storage import load_dashboard_state, write_dashboard_state


CONFIG_EDIT_EMPLOYEE_CODES = {
    "mapeamento_matricula_ausente",
    "mapeamento_matricula_divergente_config",
}
CONFIG_EDIT_EVENT_CODES = {
    "mapeamento_evento_ausente",
    "mapeamento_evento_inativo",
}


def run_dashboard_analysis(
    paths: DashboardPaths,
    *,
    config_resolver: ConfigResolver | None = None,
    column_profile_root: str | Path | None = None,
    employee_registry_root: str | Path | None = None,
    rubric_catalog_root: str | Path | None = None,
) -> DashboardRunResult:
    resolver = config_resolver or ConfigResolver()
    source_workbook_path = _resolve_source_workbook_path(paths)
    inspection = inspect_input_workbook(source_workbook_path, registry_root=resolver.registry_root)
    profile_resolution, column_profile = _resolve_column_mapping_profile(
        inspection=inspection,
        source_workbook_path=source_workbook_path,
        profile_root=column_profile_root,
    )
    if profile_resolution.status != "found" and profile_resolution.status != "not_required":
        return _build_blocked_run_without_profile(
            paths=paths,
            state=load_dashboard_state(paths.state_path),
            inspection=inspection,
            profile_resolution=profile_resolution,
        )

    if column_profile is not None:
        normalize_workbook_with_column_profile(
            source_workbook_path,
            inspection=inspection,
            profile=column_profile,
            output_path=paths.editable_workbook_path,
            report_path=paths.normalization_path,
        )
    else:
        normalize_input_workbook(
            source_workbook_path,
            output_path=paths.editable_workbook_path,
            report_path=paths.normalization_path,
            registry_root=resolver.registry_root,
        )
    ingest_fill_and_persist_planilha_padrao_v1(
        paths.editable_workbook_path,
        output_path=paths.analyzed_workbook_path,
        snapshot_path=paths.snapshot_path,
        manifest_path=paths.manifest_path,
        write_manifest_file=True,
    )
    snapshot_payload = _load_json(paths.snapshot_path)
    state = load_dashboard_state(paths.state_path)
    resolution = resolver.resolve(
        company_code=str(snapshot_payload["parameters"]["company_code"]),
        competence=str(snapshot_payload["parameters"]["competence"]),
    )

    if resolution.status != ConfigResolutionStatus.FOUND:
        return _build_blocked_run_without_config(
            paths=paths,
            state=state,
            snapshot_payload=snapshot_payload,
            resolution=resolution,
            profile_resolution=profile_resolution,
        )

    resolver.write_resolved_config(resolution, target_path=paths.editable_config_path)
    apply_employee_registry_to_editable_config(
        paths.editable_config_path,
        company_code=str(snapshot_payload["parameters"]["company_code"]),
        snapshot_payload=snapshot_payload,
        root=employee_registry_root,
    )
    apply_rubric_catalog_to_editable_config(
        paths.editable_config_path,
        company_code=str(snapshot_payload["parameters"]["company_code"]),
        snapshot_payload=snapshot_payload,
        root=rubric_catalog_root,
    )
    replay_dashboard_action_overrides(paths, state)
    map_snapshot_with_company_config(
        paths.snapshot_path,
        paths.editable_config_path,
        output_path=paths.mapped_artifact_path,
    )
    serialize_mapped_artifact_to_txt(
        paths.mapped_artifact_path,
        txt_path=paths.txt_path,
        summary_path=paths.serialization_summary_path,
    )
    validate_pipeline_v1(
        snapshot_path=paths.snapshot_path,
        mapped_artifact_path=paths.mapped_artifact_path,
        txt_path=paths.txt_path,
        serialization_summary_path=paths.serialization_summary_path,
        output_path=paths.validation_path,
    )

    mapped_payload = _load_json(paths.mapped_artifact_path)
    serialization_payload = _load_json(paths.serialization_summary_path)
    validation_payload = _load_json(paths.validation_path)
    pendings = collect_dashboard_pendings(
        paths=paths,
        snapshot_payload=snapshot_payload,
        mapped_payload=mapped_payload,
    )
    config_resolution = _resolution_to_dashboard_config(resolution)
    summary = build_dashboard_summary(
        state=state,
        snapshot_payload=snapshot_payload,
        serialization_payload=serialization_payload,
        validation_payload=validation_payload,
        pending_count=len(pendings),
        config_resolution=config_resolution,
    )

    updated_state = state.model_copy(
        update={
            "source_config_name": (
                resolution.source_path.name if resolution.source_path is not None else state.source_config_name
            ),
            "last_analysis": {
                "summary": _serialize_summary(summary),
                "config_resolution": _serialize_config_resolution(config_resolution),
                "profile_resolution": _serialize_profile_resolution(profile_resolution),
                "pendings": [_serialize_pending_item(item) for item in pendings],
            }
        }
    )
    write_dashboard_state(paths.state_path, updated_state)

    return DashboardRunResult(
        paths=paths,
        state=updated_state,
        summary=summary,
        config_resolution=config_resolution,
        profile_resolution=profile_resolution,
        pendings=pendings,
        snapshot_payload=snapshot_payload,
        mapped_payload=mapped_payload,
        serialization_payload=serialization_payload,
        validation_payload=validation_payload,
    )


def load_dashboard_run(paths: DashboardPaths) -> DashboardRunResult:
    state = load_dashboard_state(paths.state_path)
    if state.last_analysis is None:
        raise FileNotFoundError("Nenhuma analise anterior foi encontrada para este caso.")

    snapshot_payload = _load_json(paths.snapshot_path) if paths.snapshot_path.exists() else {}
    mapped_payload = _load_json(paths.mapped_artifact_path) if paths.mapped_artifact_path.exists() else {}
    serialization_payload = (
        _load_json(paths.serialization_summary_path)
        if paths.serialization_summary_path.exists()
        else {"counts": {"serialized": 0, "non_serialized": 0, "blocked_or_non_serialized": 0, "total_mapped_movements": 0}}
    )
    validation_payload = (
        _load_json(paths.validation_path)
        if paths.validation_path.exists()
        else {
            "execution": {"status": state.last_analysis["summary"]["validation_status"]},
            "fatal_errors": [],
            "inconsistencies": [],
            "recommendation": state.last_analysis["summary"]["recommendation"],
        }
    )
    summary = _deserialize_summary(state.last_analysis["summary"])
    config_resolution = _deserialize_config_resolution(state.last_analysis["config_resolution"])
    profile_resolution = _deserialize_profile_resolution(
        state.last_analysis.get("profile_resolution")
    )
    pendings = tuple(_deserialize_pending_item(item) for item in state.last_analysis["pendings"])

    return DashboardRunResult(
        paths=paths,
        state=state,
        summary=summary,
        config_resolution=config_resolution,
        profile_resolution=profile_resolution,
        pendings=pendings,
        snapshot_payload=snapshot_payload,
        mapped_payload=mapped_payload,
        serialization_payload=serialization_payload,
        validation_payload=validation_payload,
    )


def build_dashboard_summary(
    *,
    state,
    snapshot_payload: dict,
    serialization_payload: dict,
    validation_payload: dict,
    pending_count: int,
    config_resolution: DashboardConfigResolution,
) -> DashboardSummary:
    parameters = snapshot_payload["parameters"]
    counts = snapshot_payload["counts"]
    validation_status = str(validation_payload["execution"]["status"])
    txt_enabled = is_txt_download_enabled(
        validation_payload=validation_payload,
        serialization_payload=serialization_payload,
    )
    serialized_line_count = int(serialization_payload["counts"]["serialized"])
    ignored_count = sum(
        1 for action in state.actions if action.action_type.value == "ignorar_nesta_importacao"
    )

    return DashboardSummary(
        company_name=str(parameters["company_name"]),
        company_code=str(parameters["company_code"]),
        competence=str(parameters["competence"]),
        employee_count=int(counts["employees"]),
        relevant_movement_count=int(counts["movements"]),
        pending_count=pending_count,
        ignored_count=ignored_count,
        serialized_line_count=serialized_line_count,
        validation_status=validation_status,
        status_label=_humanize_validation_status(validation_status),
        recommendation=str(validation_payload["recommendation"]),
        txt_enabled=txt_enabled,
        txt_status_label="Liberado para baixar" if txt_enabled else "Ainda bloqueado",
        config_status=config_resolution.status,
        config_status_label=config_resolution.status_label,
        config_source=config_resolution.config_source,
        config_version=config_resolution.config_version,
    )


def is_txt_download_enabled(*, validation_payload: dict, serialization_payload: dict) -> bool:
    validation_status = str(validation_payload["execution"]["status"])
    serialized_count = int(serialization_payload["counts"]["serialized"])
    has_blocking_issues = bool(validation_payload.get("fatal_errors")) or bool(
        validation_payload.get("inconsistencies")
    )
    return (
        validation_status in {"success", "success_with_warnings"}
        and not has_blocking_issues
        and serialized_count > 0
    )


def collect_dashboard_pendings(
    *,
    paths: DashboardPaths,
    snapshot_payload: dict,
    mapped_payload: dict,
) -> tuple[DashboardPendingItem, ...]:
    workbook = load_workbook(paths.editable_workbook_path, data_only=False)
    config_payload = _load_json(paths.editable_config_path) if paths.editable_config_path.exists() else {}

    pendings: list[DashboardPendingItem] = []
    for pending_payload in snapshot_payload.get("pendings", ()):
        pendings.append(
            _build_pending_item(
                stage="ingestao",
                pending_payload=pending_payload,
                workbook=workbook,
                config_payload=config_payload,
            )
        )

    for pending_payload in mapped_payload.get("mapping_pendings", ()):
        pendings.append(
            _build_pending_item(
                stage="mapeamento",
                pending_payload=pending_payload,
                workbook=workbook,
                config_payload=config_payload,
            )
        )

    return tuple(pendings)


def _build_blocked_run_without_config(
    *,
    paths: DashboardPaths,
    state,
    snapshot_payload: dict,
    resolution: ConfigResolutionResult,
    profile_resolution: DashboardProfileResolution,
) -> DashboardRunResult:
    _cleanup_downstream_artifacts(paths)
    config_resolution = _resolution_to_dashboard_config(resolution)
    internal_pending = _build_config_resolution_pending(snapshot_payload, config_resolution)
    pendings = (
        *collect_dashboard_pendings(paths=paths, snapshot_payload=snapshot_payload, mapped_payload={}),
        internal_pending,
    )
    serialization_payload = {
        "counts": {
            "serialized": 0,
            "non_serialized": 0,
            "blocked_or_non_serialized": 0,
            "total_mapped_movements": 0,
        }
    }
    validation_payload = {
        "execution": {"status": "blocked"},
        "fatal_errors": [],
        "inconsistencies": [],
        "recommendation": config_resolution.message,
    }
    summary = build_dashboard_summary(
        state=state,
        snapshot_payload=snapshot_payload,
        serialization_payload=serialization_payload,
        validation_payload=validation_payload,
        pending_count=len(pendings),
        config_resolution=config_resolution,
    )
    updated_state = state.model_copy(
        update={
            "source_config_name": None,
            "last_analysis": {
                "summary": _serialize_summary(summary),
                "config_resolution": _serialize_config_resolution(config_resolution),
                "profile_resolution": _serialize_profile_resolution(profile_resolution),
                "pendings": [_serialize_pending_item(item) for item in pendings],
            },
        }
    )
    write_dashboard_state(paths.state_path, updated_state)
    return DashboardRunResult(
        paths=paths,
        state=updated_state,
        summary=summary,
        config_resolution=config_resolution,
        profile_resolution=profile_resolution,
        pendings=tuple(pendings),
        snapshot_payload=snapshot_payload,
        mapped_payload={},
        serialization_payload=serialization_payload,
        validation_payload=validation_payload,
    )


def _build_blocked_run_without_profile(
    *,
    paths: DashboardPaths,
    state,
    inspection: InputWorkbookInspection,
    profile_resolution: DashboardProfileResolution,
) -> DashboardRunResult:
    _cleanup_profile_block_artifacts(paths)
    snapshot_payload = _inspection_to_snapshot_payload(inspection)
    config_resolution = _profile_block_config_resolution(inspection, profile_resolution)
    internal_pending = _build_profile_resolution_pending(inspection, profile_resolution)
    pendings = (internal_pending,)
    serialization_payload = {
        "counts": {
            "serialized": 0,
            "non_serialized": 0,
            "blocked_or_non_serialized": 0,
            "total_mapped_movements": 0,
        }
    }
    validation_payload = {
        "execution": {"status": "blocked"},
        "fatal_errors": [],
        "inconsistencies": [],
        "recommendation": profile_resolution.message,
    }
    summary = build_dashboard_summary(
        state=state,
        snapshot_payload=snapshot_payload,
        serialization_payload=serialization_payload,
        validation_payload=validation_payload,
        pending_count=len(pendings),
        config_resolution=config_resolution,
    )
    updated_state = state.model_copy(
        update={
            "source_config_name": None,
            "last_analysis": {
                "summary": _serialize_summary(summary),
                "config_resolution": _serialize_config_resolution(config_resolution),
                "profile_resolution": _serialize_profile_resolution(profile_resolution),
                "pendings": [_serialize_pending_item(item) for item in pendings],
            },
        }
    )
    write_dashboard_state(paths.state_path, updated_state)
    return DashboardRunResult(
        paths=paths,
        state=updated_state,
        summary=summary,
        config_resolution=config_resolution,
        profile_resolution=profile_resolution,
        pendings=pendings,
        snapshot_payload=snapshot_payload,
        mapped_payload={},
        serialization_payload=serialization_payload,
        validation_payload=validation_payload,
    )


def _build_pending_item(
    *,
    stage: str,
    pending_payload: dict,
    workbook,
    config_payload: dict,
) -> DashboardPendingItem:
    source = pending_payload.get("source", {})
    code = str(pending_payload["pending_code"])
    event_name = pending_payload.get("event_name")
    source_sheet = source.get("sheet_name")
    source_cell = source.get("cell")
    source_row = int(source["row_number"]) if source.get("row_number") is not None else None
    source_column_name = source.get("column_name")
    can_edit_employee_mapping = code in CONFIG_EDIT_EMPLOYEE_CODES
    can_edit_event_mapping = code in CONFIG_EDIT_EVENT_CODES
    can_edit_workbook = bool(source_sheet and source_cell) and not (
        can_edit_employee_mapping or can_edit_event_mapping
    )
    can_ignore, ignore_mode, ignore_label = describe_ignore_strategy(
        code=code,
        source_sheet=source_sheet,
        source_cell=source_cell,
        source_row=source_row,
        event_name=event_name,
    )

    return DashboardPendingItem(
        uid=f"{stage}:{pending_payload['pending_id']}",
        stage=stage,
        pending_id=str(pending_payload["pending_id"]),
        code=code,
        severity=str(pending_payload["severity"]),
        employee_name=pending_payload.get("employee_name"),
        employee_key=pending_payload.get("employee_key"),
        event_name=event_name,
        field_label=_resolve_field_label(code=code, source_column_name=source_column_name),
        found_value=_resolve_found_value(
            code=code,
            source_sheet=source_sheet,
            source_cell=source_cell,
            workbook=workbook,
            config_payload=config_payload,
            employee_key=pending_payload.get("employee_key"),
            event_name=event_name,
        ),
        problem=str(pending_payload["description"]),
        recommended_action=str(pending_payload["recommended_action"]),
        source_sheet=source_sheet,
        source_cell=source_cell,
        source_row=source_row,
        source_column_name=source_column_name,
        can_edit_workbook=can_edit_workbook,
        can_edit_employee_mapping=can_edit_employee_mapping,
        can_edit_event_mapping=can_edit_event_mapping,
        can_ignore=can_ignore,
        ignore_mode=ignore_mode,
        ignore_label=ignore_label,
    )


def _resolve_field_label(*, code: str, source_column_name: str | None) -> str:
    if code in CONFIG_EDIT_EMPLOYEE_CODES:
        return "matricula do cadastro da empresa"
    if code in CONFIG_EDIT_EVENT_CODES:
        return "rubrica de saida"
    return source_column_name or "campo da planilha"


def _resolve_found_value(
    *,
    code: str,
    source_sheet: str | None,
    source_cell: str | None,
    workbook,
    config_payload: dict,
    employee_key: str | None,
    event_name: str | None,
) -> str | None:
    if code in CONFIG_EDIT_EMPLOYEE_CODES:
        for item in config_payload.get("employee_mappings", []):
            if item.get("source_employee_key") == employee_key:
                return _stringify(item.get("domain_registration")) or "em branco"
        return "sem cadastro"

    if code in CONFIG_EDIT_EVENT_CODES:
        for item in config_payload.get("event_mappings", []):
            if item.get("event_negocio") == event_name:
                return _stringify(item.get("rubrica_saida")) or "em branco"
        return "sem rubrica"

    if source_sheet and source_sheet in workbook.sheetnames and source_cell is not None:
        return _stringify(workbook[source_sheet][source_cell].value) or "em branco"

    return None


def _humanize_validation_status(status: str) -> str:
    if status == "success":
        return "Analise concluida e pronta para exportacao"
    if status == "success_with_warnings":
        return "Analise concluida com alertas"
    if status == "empty":
        return "Sem linhas liberadas para exportacao"
    return "Analise bloqueada"


def _resolution_to_dashboard_config(result: ConfigResolutionResult) -> DashboardConfigResolution:
    if result.status == ConfigResolutionStatus.FOUND:
        if result.config_source in {"registry_company_competence", "legacy_company_competence"}:
            status_label = "Configuracao interna encontrada para a competencia"
        elif result.config_source in {"registry_company_active", "legacy_company_active"}:
            status_label = "Configuracao ativa da empresa aplicada"
        else:
            status_label = "Configuracao interna encontrada"
    elif result.status == ConfigResolutionStatus.NOT_FOUND:
        status_label = "Configuracao interna nao encontrada"
    elif result.status == ConfigResolutionStatus.AMBIGUOUS:
        status_label = "Mais de uma configuracao interna candidata foi encontrada"
    else:
        status_label = "Configuracao interna encontrada, mas inconsistente"

    return DashboardConfigResolution(
        status=result.status.value,
        status_label=status_label,
        message=result.message,
        company_code=result.company_code,
        competence=result.competence,
        config_source=result.config_source,
        config_version=result.config_version,
        source_path=str(result.source_path) if result.source_path is not None else None,
    )


def _build_config_resolution_pending(
    snapshot_payload: dict,
    config_resolution: DashboardConfigResolution,
) -> DashboardPendingItem:
    parameters = snapshot_payload["parameters"]
    return DashboardPendingItem(
        uid=f"configuracao:{config_resolution.status}",
        stage="configuracao",
        pending_id=f"config-{config_resolution.status.lower()}",
        code=f"config_resolution_{config_resolution.status.lower()}",
        severity="bloqueante",
        employee_name=None,
        employee_key=None,
        event_name=None,
        field_label="configuracao interna da empresa",
        found_value=f"empresa={parameters['company_code']} | competencia={parameters['competence']}",
        problem=config_resolution.message,
        recommended_action=(
            "Cadastrar ou corrigir a configuracao interna da empresa no backend antes de repetir a analise."
        ),
        source_sheet=None,
        source_cell=None,
        source_row=None,
        source_column_name=None,
        can_edit_workbook=False,
        can_edit_employee_mapping=False,
        can_edit_event_mapping=False,
        can_ignore=False,
    )


def _resolve_column_mapping_profile(
    *,
    inspection: InputWorkbookInspection,
    source_workbook_path: Path,
    profile_root: str | Path | None,
) -> tuple[DashboardProfileResolution, CompanyColumnMappingProfile | None]:
    if inspection.layout_id == CANONICAL_LAYOUT_ID:
        return (
            DashboardProfileResolution(
                status="not_required",
                status_label="Perfil de colunas nao requerido",
                message="Layout canonico V1 ja esta no contrato interno e nao exige perfil de colunas.",
                company_code=inspection.company_code,
                competence=inspection.competence,
                layout_id=inspection.layout_id,
                source_path=None,
            ),
            None,
        )

    profile_path = column_mapping_profile_path(inspection.company_code, root=profile_root)
    try:
        profile = load_column_mapping_profile(inspection.company_code, root=profile_root)
    except ColumnMappingProfileError as exc:
        return (
            DashboardProfileResolution(
                status="missing",
                status_label="Perfil de mapeamento de colunas ausente",
                message=(
                    "Perfil de mapeamento de colunas nao encontrado para a empresa "
                    f"{inspection.company_code}. Cadastre o perfil antes de processar este layout."
                ),
                company_code=inspection.company_code,
                competence=inspection.competence,
                layout_id=inspection.layout_id,
                source_path=exc.source or str(profile_path),
            ),
            None,
        )
    except (ValidationError, ValueError) as exc:
        return (
            DashboardProfileResolution(
                status="invalid",
                status_label="Perfil de mapeamento de colunas invalido",
                message=f"Perfil de mapeamento de colunas invalido para a empresa {inspection.company_code}: {exc}",
                company_code=inspection.company_code,
                competence=inspection.competence,
                layout_id=inspection.layout_id,
                source_path=str(profile_path),
            ),
            None,
        )

    if profile.company_code != inspection.company_code:
        return (
            DashboardProfileResolution(
                status="invalid",
                status_label="Perfil de mapeamento de colunas invalido",
                message=(
                    "Perfil de mapeamento de colunas pertence a outra empresa. "
                    f"Esperado={inspection.company_code}; recebido={profile.company_code}."
                ),
                company_code=inspection.company_code,
                competence=inspection.competence,
                layout_id=inspection.layout_id,
                source_path=str(profile_path),
            ),
            None,
        )

    relevant_columns = _relevant_profile_columns(
        inspection=inspection,
        source_workbook_path=source_workbook_path,
    )
    missing_columns = _missing_profile_columns(profile, relevant_columns)
    if missing_columns:
        first_missing = missing_columns[0]
        return (
            DashboardProfileResolution(
                status="incomplete",
                status_label="Perfil de mapeamento de colunas incompleto",
                message=(
                    "Perfil de mapeamento de colunas incompleto para a empresa "
                    f"{inspection.company_code}. Coluna sem mapping: {first_missing}."
                ),
                company_code=inspection.company_code,
                competence=inspection.competence,
                layout_id=inspection.layout_id,
                source_path=str(profile_path),
                missing_columns=missing_columns,
            ),
            None,
        )

    return (
        DashboardProfileResolution(
            status="found",
            status_label="Perfil de mapeamento de colunas encontrado",
            message=(
                f"Perfil de mapeamento de colunas encontrado para a empresa {inspection.company_code} "
                f"com cobertura das colunas relevantes."
            ),
            company_code=inspection.company_code,
            competence=inspection.competence,
            layout_id=inspection.layout_id,
            source_path=str(profile_path),
        ),
        profile,
    )


def _relevant_profile_columns(
    *,
    inspection: InputWorkbookInspection,
    source_workbook_path: Path,
) -> tuple[InputColumnMetadata, ...]:
    workbook = load_workbook(source_workbook_path, data_only=False)
    worksheet = workbook[inspection.selected_sheet_name]
    relevant: list[InputColumnMetadata] = []
    for column in inspection.columns:
        if column.column_index <= 2:
            continue
        if is_profile_identity_column(column):
            continue
        if _column_has_any_data(worksheet, column):
            relevant.append(column)
    return tuple(relevant)


def _column_has_any_data(worksheet, column: InputColumnMetadata) -> bool:
    for row_number in range(column.header_row + 1, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=column.column_index).value
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False


def _missing_profile_columns(
    profile: CompanyColumnMappingProfile,
    columns: tuple[InputColumnMetadata, ...],
) -> tuple[str, ...]:
    missing = [
        column.column_name
        for column in columns
        if not _profile_covers_column(profile, column)
    ]
    return tuple(missing)


def _profile_covers_column(
    profile: CompanyColumnMappingProfile,
    column: InputColumnMetadata,
) -> bool:
    column_tokens = {
        _normalize_profile_token(column.column_name),
        column.normalized_column_name,
        _normalize_profile_token(column.column_letter),
        _normalize_profile_token(str(column.column_index)),
    }
    for mapping in profile.mappings:
        mapping_tokens = {
            _normalize_profile_token(token)
            for token in (mapping.column_key, mapping.column_name)
            if token
        }
        if column_tokens.intersection(mapping_tokens):
            return True
    return False


def _normalize_profile_token(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value)).strip().lower()
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(text.split())


def _inspection_to_snapshot_payload(inspection: InputWorkbookInspection) -> dict:
    return {
        "parameters": {
            "company_code": inspection.company_code,
            "company_name": inspection.company_name,
            "competence": inspection.competence,
            "payroll_type": "mensal",
            "default_process": "",
            "layout_version": inspection.layout_id,
        },
        "counts": {
            "employees": 0,
            "movements": 0,
            "pendings": 1,
        },
        "pendings": [],
    }


def _profile_block_config_resolution(
    inspection: InputWorkbookInspection,
    profile_resolution: DashboardProfileResolution,
) -> DashboardConfigResolution:
    return DashboardConfigResolution(
        status="not_run",
        status_label="Configuracao interna nao avaliada",
        message=(
            "ConfigResolver nao foi executado porque a etapa de perfil de colunas "
            f"bloqueou a analise: {profile_resolution.message}"
        ),
        company_code=inspection.company_code,
        competence=inspection.competence,
        config_source=None,
        config_version=None,
        source_path=None,
    )


def _build_profile_resolution_pending(
    inspection: InputWorkbookInspection,
    profile_resolution: DashboardProfileResolution,
) -> DashboardPendingItem:
    found_value = f"empresa={inspection.company_code} | competencia={inspection.competence}"
    if profile_resolution.missing_columns:
        found_value = f"{found_value} | coluna={profile_resolution.missing_columns[0]}"

    return DashboardPendingItem(
        uid=f"perfil_colunas:{profile_resolution.status}",
        stage="perfil_colunas",
        pending_id=f"profile-{profile_resolution.status}",
        code=f"column_mapping_profile_{profile_resolution.status}",
        severity="bloqueante",
        employee_name=None,
        employee_key=None,
        event_name=None,
        field_label="perfil de mapeamento de colunas",
        found_value=found_value,
        problem=profile_resolution.message,
        recommended_action=(
            "Cadastrar ou corrigir o perfil de mapeamento de colunas da empresa antes de repetir a analise."
        ),
        source_sheet=inspection.selected_sheet_name,
        source_cell=None,
        source_row=None,
        source_column_name=(profile_resolution.missing_columns[0] if profile_resolution.missing_columns else None),
        can_edit_workbook=False,
        can_edit_employee_mapping=False,
        can_edit_event_mapping=False,
        can_ignore=False,
    )


def _cleanup_downstream_artifacts(paths: DashboardPaths) -> None:
    for target in (
        paths.editable_config_path,
        paths.mapped_artifact_path,
        paths.txt_path,
        paths.serialization_summary_path,
        paths.validation_path,
    ):
        target.unlink(missing_ok=True)


def _cleanup_profile_block_artifacts(paths: DashboardPaths) -> None:
    for target in (
        paths.editable_workbook_path,
        paths.editable_config_path,
        paths.analyzed_workbook_path,
        paths.snapshot_path,
        paths.manifest_path,
        paths.normalization_path,
        paths.mapped_artifact_path,
        paths.txt_path,
        paths.serialization_summary_path,
        paths.validation_path,
    ):
        target.unlink(missing_ok=True)


def _resolve_source_workbook_path(paths: DashboardPaths) -> Path:
    if paths.editable_workbook_path.exists():
        return paths.editable_workbook_path
    if paths.raw_workbook_path.exists():
        return paths.raw_workbook_path
    raise FileNotFoundError("Nenhuma planilha enviada foi encontrada para esta execucao.")


def _serialize_pending_item(item: DashboardPendingItem) -> dict:
    return {
        "uid": item.uid,
        "stage": item.stage,
        "pending_id": item.pending_id,
        "code": item.code,
        "severity": item.severity,
        "employee_name": item.employee_name,
        "employee_key": item.employee_key,
        "event_name": item.event_name,
        "field_label": item.field_label,
        "found_value": item.found_value,
        "problem": item.problem,
        "recommended_action": item.recommended_action,
        "source_sheet": item.source_sheet,
        "source_cell": item.source_cell,
        "source_row": item.source_row,
        "source_column_name": item.source_column_name,
        "can_edit_workbook": item.can_edit_workbook,
        "can_edit_employee_mapping": item.can_edit_employee_mapping,
        "can_edit_event_mapping": item.can_edit_event_mapping,
        "can_ignore": item.can_ignore,
        "ignore_mode": item.ignore_mode,
        "ignore_label": item.ignore_label,
    }


def _deserialize_pending_item(payload: dict) -> DashboardPendingItem:
    return DashboardPendingItem(**payload)


def _serialize_summary(summary: DashboardSummary) -> dict:
    return {
        "company_name": summary.company_name,
        "company_code": summary.company_code,
        "competence": summary.competence,
        "employee_count": summary.employee_count,
        "relevant_movement_count": summary.relevant_movement_count,
        "pending_count": summary.pending_count,
        "ignored_count": summary.ignored_count,
        "serialized_line_count": summary.serialized_line_count,
        "validation_status": summary.validation_status,
        "status_label": summary.status_label,
        "recommendation": summary.recommendation,
        "txt_enabled": summary.txt_enabled,
        "txt_status_label": summary.txt_status_label,
        "config_status": summary.config_status,
        "config_status_label": summary.config_status_label,
        "config_source": summary.config_source,
        "config_version": summary.config_version,
    }


def _deserialize_summary(payload: dict) -> DashboardSummary:
    return DashboardSummary(**payload)


def _serialize_config_resolution(config_resolution: DashboardConfigResolution) -> dict:
    return {
        "status": config_resolution.status,
        "status_label": config_resolution.status_label,
        "message": config_resolution.message,
        "company_code": config_resolution.company_code,
        "competence": config_resolution.competence,
        "config_source": config_resolution.config_source,
        "config_version": config_resolution.config_version,
        "source_path": config_resolution.source_path,
    }


def _deserialize_config_resolution(payload: dict) -> DashboardConfigResolution:
    return DashboardConfigResolution(**payload)


def _serialize_profile_resolution(profile_resolution: DashboardProfileResolution) -> dict:
    return {
        "status": profile_resolution.status,
        "status_label": profile_resolution.status_label,
        "message": profile_resolution.message,
        "company_code": profile_resolution.company_code,
        "competence": profile_resolution.competence,
        "layout_id": profile_resolution.layout_id,
        "source_path": profile_resolution.source_path,
        "missing_columns": list(profile_resolution.missing_columns),
    }


def _deserialize_profile_resolution(payload: dict | None) -> DashboardProfileResolution:
    if payload is None:
        return DashboardProfileResolution(
            status="unknown",
            status_label="Perfil de colunas nao registrado",
            message="Analise anterior nao registrou resolucao de perfil de colunas.",
            company_code="",
            competence="",
            layout_id="",
            source_path=None,
        )
    return DashboardProfileResolution(
        status=str(payload["status"]),
        status_label=str(payload["status_label"]),
        message=str(payload["message"]),
        company_code=str(payload["company_code"]),
        competence=str(payload["competence"]),
        layout_id=str(payload["layout_id"]),
        source_path=payload.get("source_path"),
        missing_columns=tuple(payload.get("missing_columns", ())),
    )


def _load_json(path: str | Path) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def _stringify(value) -> str | None:
    if value is None:
        return None
    return str(value)
