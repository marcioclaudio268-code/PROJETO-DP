from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from dashboard import (
    DashboardActionType,
    ColumnGenerationMode,
    ColumnMappingRule,
    ColumnValueKind,
    CompanyEmployeeRecord,
    CompanyEmployeeRegistry,
    CompanyColumnMappingProfile,
    CompanyRubricCatalog,
    CompanyRubricRecord,
    ConfigResolutionStatus,
    ConfigResolver,
    analyze_report_import,
    apply_dashboard_action,
    apply_workbook_cell_correction,
    apply_report_employee_suggestions,
    apply_report_rubric_suggestions,
    company_employee_registry_path,
    company_rubric_catalog_path,
    create_dashboard_run_from_paths,
    ignore_pending_for_import,
    load_company_employee_registry,
    load_company_rubric_catalog,
    load_dashboard_run,
    run_dashboard_analysis,
    save_column_mapping_profile,
    save_company_employee_registry,
    save_company_rubric_catalog,
)
from dashboard.txt_audit import build_txt_audit
from ingestion import save_planilha_padrao_folha_v1


REPO_ROOT = Path(__file__).resolve().parents[2]
MONTHLY_FIXTURE = REPO_ROOT / "data" / "runs" / "dashboard_v1" / "run-15f4250f85" / "inputs" / "input.xlsx"


def _write_internal_config(
    root: Path,
    *,
    company_code: str,
    file_name: str,
    competence: str,
    config_version: str = "cfg-v1",
    payload_override: dict | None = None,
) -> Path:
    payload = {
        "company_code": company_code,
        "company_name": "Dela More",
        "default_process": "11",
        "competence": competence,
        "config_version": config_version,
        "event_mappings": [
            {
                "event_negocio": "gratificacao",
                "rubrica_saida": "201",
            },
            {
                "event_negocio": "horas_extras_50",
                "rubrica_saida": "350",
            },
        ],
        "employee_mappings": [
            {
                "source_employee_key": "col-001",
                "source_employee_name": "Ana Lima",
                "domain_registration": "123",
            },
            {
                "source_employee_key": "col-999",
                "source_employee_name": "Nova Pessoa",
                "domain_registration": "999",
            },
        ],
        "pending_policy": {
            "review_required_event_negocios": [],
            "review_required_fields": [],
            "block_on_ambiguous_observations": True,
            "block_on_unmapped_employee": True,
            "block_on_unmapped_event": True,
        },
        "validation_flags": {},
    }
    if payload_override:
        payload.update(payload_override)
    path = root / company_code / file_name
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=True, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return path


def _write_monthly_column_profile(
    root: Path,
    *,
    company_code: str = "528",
    omit_column: str | None = None,
) -> Path:
    raw_mappings = [
        ("H. EXTRA 50% COD.150", True, "150", None, ColumnValueKind.HOURS, ColumnGenerationMode.SINGLE_LINE),
        ("H. EXTRA 100% COD.200  FERIADO", True, "200", None, ColumnValueKind.HOURS, ColumnGenerationMode.SINGLE_LINE),
        ("ADIANTAMENTO COD. 981", False, None, None, ColumnValueKind.MONETARY, ColumnGenerationMode.IGNORE),
        ("ATRASOS       COD. 8069", True, "8069", None, ColumnValueKind.HOURS, ColumnGenerationMode.SINGLE_LINE),
        ("CONSUMO              COD. 266", True, "266", None, ColumnValueKind.MONETARY, ColumnGenerationMode.SINGLE_LINE),
        ("VALE TRANSPORTE COD. 48", False, None, None, ColumnValueKind.MONETARY, ColumnGenerationMode.IGNORE),
    ]
    mappings = []
    for column_name, enabled, rubrica_target, rubricas_target, value_kind, generation_mode in raw_mappings:
        if column_name == omit_column:
            continue
        mappings.append(
            ColumnMappingRule(
                column_name=column_name,
                enabled=enabled,
                rubrica_target=rubrica_target,
                rubricas_target=rubricas_target or [],
                value_kind=value_kind,
                generation_mode=generation_mode,
                ignore_zero=True,
                ignore_text=True,
            )
        )

    profile = CompanyColumnMappingProfile(
        company_code=company_code,
        company_name="FRIED FISH VILAREJO",
        default_process="11",
        mappings=mappings,
    )
    return save_column_mapping_profile(profile, root=root)


def _write_position_column_profile(root: Path) -> Path:
    profile = CompanyColumnMappingProfile(
        company_code="755",
        company_name="GUSTAVO LOPES LACERDA",
        default_process="11",
        mappings=[
            ColumnMappingRule(
                sheet_name="abril",
                header_row=2,
                data_start_row=3,
                employee_code_column="A",
                employee_name_column="B",
                value_column="T",
                expected_header="HORA 50%",
                enabled=True,
                rubrica_target="201",
                value_kind=ColumnValueKind.HOURS,
                nature="provento",
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
                status="active",
            )
        ],
    )
    return save_column_mapping_profile(profile, root=root)


def _write_position_profile_workbook(path: Path, *, header: str = "HORA 50%") -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "abril"
    worksheet["A2"] = "CODIGO"
    worksheet["B2"] = "NOME"
    worksheet["T2"] = header
    worksheet["A3"] = "304"
    worksheet["B3"] = "ADILSON RAFAEL DE SOUSA"
    worksheet["T3"] = "01:30"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _prepare_single_row_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "input.xlsx"
    save_planilha_padrao_folha_v1(workbook_path)

    workbook = load_workbook(workbook_path)
    parametros = workbook["PARAMETROS"]
    parametros["B2"] = "72"
    parametros["B3"] = "Dela More"
    parametros["B4"] = "03/2024"
    parametros["B5"] = "mensal"
    parametros["B6"] = "11"
    parametros["B7"] = "v1"

    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B2"] = "col-999"
    lancamentos["C2"] = "Nova Pessoa"
    lancamentos["H2"] = 100
    workbook.save(workbook_path)
    return workbook_path


def _prepare_non_automatable_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "input.xlsx"
    save_planilha_padrao_folha_v1(workbook_path)

    workbook = load_workbook(workbook_path)
    parametros = workbook["PARAMETROS"]
    parametros["B2"] = "72"
    parametros["B3"] = "Dela More"
    parametros["B4"] = "03/2024"
    parametros["B5"] = "mensal"
    parametros["B6"] = "11"
    parametros["B7"] = "v1"

    funcionarios = workbook["FUNCIONARIOS"]
    funcionarios["A2"] = "col-001"
    funcionarios["B2"] = "Ana Lima"
    funcionarios["E2"] = "123"
    funcionarios["H2"] = "ativo"
    funcionarios["I2"] = "sim"

    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B2"] = "col-001"
    lancamentos["C2"] = "Ana Lima"
    lancamentos["D2"] = "123"
    lancamentos["H2"] = 100
    lancamentos["N2"] = "revisar"
    workbook.save(workbook_path)

    return workbook_path


def test_dashboard_happy_path_enables_txt(tmp_path: Path) -> None:
    workbook_path = REPO_ROOT / "data" / "golden" / "v1" / "happy_path" / "input.xlsx"
    configs_root = tmp_path / "configs"
    _write_internal_config(configs_root, company_code="72", file_name="03-2024.json", competence="03/2024")
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )
    persisted = load_dashboard_run(paths)

    assert result.summary.txt_enabled is True
    assert result.summary.serialized_line_count == 2
    assert result.summary.pending_count == 0
    assert result.summary.config_status == ConfigResolutionStatus.FOUND.value
    assert result.profile_resolution.status == "not_required"
    assert persisted.summary.txt_enabled is True
    assert persisted.summary.serialized_line_count == 2

    audit = build_txt_audit(result)
    assert audit.summary.total_lines == 2
    assert {row.check_status for row in audit.employee_rows} == {"OK"}
    assert audit.divergences == ()


def test_dashboard_default_master_data_allows_xlsx_only_flow(tmp_path: Path) -> None:
    workbook_path = REPO_ROOT / "data" / "golden" / "v1" / "happy_path" / "input.xlsx"
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(paths)

    assert result.summary.txt_enabled is True
    assert result.summary.config_status == ConfigResolutionStatus.FOUND.value
    assert result.summary.config_source == "registry_company_active"
    assert result.summary.config_version == "cfg-v1"
    assert result.profile_resolution.status == "not_required"


def test_dashboard_blocks_when_selected_company_differs_from_detected_company(tmp_path: Path) -> None:
    workbook_path = REPO_ROOT / "data" / "golden" / "v1" / "happy_path" / "input.xlsx"
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        selected_company_code="528",
        selected_company_name="FRIED FISH VILAREJO",
    )

    assert result.summary.txt_enabled is False
    assert result.summary.validation_status == "blocked"
    assert result.config_resolution.status == "empresa_selecionada_divergente"
    assert any(item.code == "empresa_selecionada_divergente" for item in result.pendings)
    assert result.summary.company_code == "72"


def test_dashboard_can_fix_missing_registration_and_reprocess(tmp_path: Path) -> None:
    workbook_path = _prepare_single_row_workbook(tmp_path)
    configs_root = tmp_path / "configs"
    _write_internal_config(configs_root, company_code="72", file_name="active.json", competence="01/2024")
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    initial = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )
    assert initial.summary.txt_enabled is False
    pending = next(item for item in initial.pendings if item.code == "matricula_dominio_ausente")

    apply_workbook_cell_correction(
        paths,
        sheet_name=pending.source_sheet or "",
        cell=pending.source_cell or "",
        new_value="999",
        pending_uid=pending.uid,
    )
    updated = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )

    assert updated.summary.txt_enabled is True
    assert updated.summary.serialized_line_count == 1
    assert not any(item.code == "matricula_dominio_ausente" for item in updated.pendings)
    assert updated.summary.config_source == "legacy_company_active"


def test_dashboard_can_ignore_non_automatizable_event_and_reprocess(tmp_path: Path) -> None:
    workbook_path = _prepare_non_automatable_workbook(tmp_path)
    configs_root = tmp_path / "configs"
    _write_internal_config(configs_root, company_code="72", file_name="03-2024.json", competence="03/2024")
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    initial = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )
    assert initial.summary.pending_count == 1
    pending = next(item for item in initial.pendings if item.code == "evento_nao_automatizavel")

    ignore_pending_for_import(paths, pending)
    updated = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )

    assert updated.summary.pending_count == 0
    assert updated.summary.ignored_count == 1
    assert updated.summary.txt_enabled is True
    workbook = load_workbook(paths.editable_workbook_path)
    assert workbook["LANCAMENTOS_FACEIS"]["N2"].value is None


def test_dashboard_can_apply_event_mapping_action_and_reprocess(tmp_path: Path) -> None:
    workbook_path = _prepare_single_row_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    workbook["LANCAMENTOS_FACEIS"]["D2"] = "999"
    workbook.save(workbook_path)

    configs_root = tmp_path / "configs"
    _write_internal_config(
        configs_root,
        company_code="72",
        file_name="03-2024.json",
        competence="03/2024",
        payload_override={"event_mappings": []},
    )
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    initial = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )
    pending = next(item for item in initial.pendings if item.code == "mapeamento_evento_ausente")

    apply_dashboard_action(
        paths,
        action_type=DashboardActionType.EVENT_MAPPING_UPDATE,
        pending_uid=pending.uid,
        payload={"output_rubric": "201"},
    )
    updated = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )

    assert updated.summary.txt_enabled is True
    assert updated.summary.serialized_line_count == 1
    assert not any(item.code == "mapeamento_evento_ausente" for item in updated.pendings)
    assert json.loads(paths.editable_config_path.read_text(encoding="utf-8"))["event_mappings"] == [
        {
            "active": True,
            "event_negocio": "gratificacao",
            "notes": None,
            "rubrica_saida": "201",
        }
    ]


def test_dashboard_applies_employee_registry_to_editable_config(tmp_path: Path) -> None:
    workbook_path = _prepare_single_row_workbook(tmp_path)
    configs_root = tmp_path / "configs"
    _write_internal_config(
        configs_root,
        company_code="72",
        file_name="03-2024.json",
        competence="03/2024",
        payload_override={"employee_mappings": []},
    )
    employee_registry_root = tmp_path / "employee_registries"
    save_company_employee_registry(
        CompanyEmployeeRegistry(
            company_code="72",
            company_name="Dela More",
            employees=[
                CompanyEmployeeRecord(
                    employee_key="col-999",
                    employee_name="Nova Pessoa",
                    domain_registration="999",
                    aliases=["N Pessoa"],
                    source="test",
                )
            ],
        ),
        root=employee_registry_root,
    )
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
        employee_registry_root=employee_registry_root,
    )

    config_payload = json.loads(paths.editable_config_path.read_text(encoding="utf-8"))
    assert config_payload["employee_mappings"] == [
        {
            "active": True,
            "aliases": [],
            "domain_registration": "999",
            "notes": "Preenchido a partir do cadastro persistente de funcionarios.",
            "source_employee_key": "col-999",
            "source_employee_name": "Nova Pessoa",
        }
    ]
    assert result.mapped_payload["config"]["active_employee_mappings"] == 1


def test_dashboard_does_not_apply_ambiguous_employee_registry_match(tmp_path: Path) -> None:
    workbook_path = _prepare_single_row_workbook(tmp_path)
    configs_root = tmp_path / "configs"
    _write_internal_config(
        configs_root,
        company_code="72",
        file_name="03-2024.json",
        competence="03/2024",
        payload_override={"employee_mappings": []},
    )
    employee_registry_root = tmp_path / "employee_registries"
    save_company_employee_registry(
        CompanyEmployeeRegistry(
            company_code="72",
            company_name="Dela More",
            employees=[
                CompanyEmployeeRecord(
                    employee_key="col-101",
                    employee_name="Pessoa Um",
                    domain_registration="101",
                    aliases=["Nova Pessoa"],
                    source="test",
                ),
                CompanyEmployeeRecord(
                    employee_key="col-102",
                    employee_name="Pessoa Dois",
                    domain_registration="102",
                    aliases=["Nova Pessoa"],
                    source="test",
                ),
            ],
        ),
        root=employee_registry_root,
    )
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
        employee_registry_root=employee_registry_root,
    )

    config_payload = json.loads(paths.editable_config_path.read_text(encoding="utf-8"))
    assert config_payload["employee_mappings"] == []


def test_dashboard_applies_rubric_catalog_to_editable_config(tmp_path: Path) -> None:
    workbook_path = _prepare_single_row_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    workbook["LANCAMENTOS_FACEIS"]["D2"] = "999"
    workbook.save(workbook_path)

    configs_root = tmp_path / "configs"
    _write_internal_config(
        configs_root,
        company_code="72",
        file_name="03-2024.json",
        competence="03/2024",
        payload_override={"event_mappings": []},
    )
    rubric_catalog_root = tmp_path / "rubric_catalogs"
    save_company_rubric_catalog(
        CompanyRubricCatalog(
            company_code="72",
            company_name="Dela More",
            rubrics=[
                CompanyRubricRecord(
                    rubric_code="201",
                    description="GRATIFICACAO",
                    canonical_event="gratificacao",
                    value_kind="monetario",
                    nature="provento",
                    aliases=["GRAT"],
                    source="test",
                )
            ],
        ),
        root=rubric_catalog_root,
    )
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
        rubric_catalog_root=rubric_catalog_root,
    )

    assert result.summary.txt_enabled is True
    assert result.summary.serialized_line_count == 1
    config_payload = json.loads(paths.editable_config_path.read_text(encoding="utf-8"))
    assert config_payload["event_mappings"] == [
        {
            "active": True,
            "event_negocio": "gratificacao",
            "notes": "Preenchido a partir do catalogo persistente de rubricas.",
            "rubrica_saida": "201",
        }
    ]


def test_dashboard_does_not_apply_ambiguous_rubric_catalog_match(tmp_path: Path) -> None:
    workbook_path = _prepare_single_row_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    workbook["LANCAMENTOS_FACEIS"]["D2"] = "999"
    workbook.save(workbook_path)

    configs_root = tmp_path / "configs"
    _write_internal_config(
        configs_root,
        company_code="72",
        file_name="03-2024.json",
        competence="03/2024",
        payload_override={"event_mappings": []},
    )
    rubric_catalog_root = tmp_path / "rubric_catalogs"
    save_company_rubric_catalog(
        CompanyRubricCatalog(
            company_code="72",
            company_name="Dela More",
            rubrics=[
                CompanyRubricRecord(
                    rubric_code="201",
                    description="GRATIFICACAO A",
                    canonical_event="gratificacao",
                    value_kind="monetario",
                    nature="provento",
                    source="test",
                ),
                CompanyRubricRecord(
                    rubric_code="202",
                    description="GRATIFICACAO B",
                    canonical_event="gratificacao",
                    value_kind="monetario",
                    nature="provento",
                    source="test",
                ),
            ],
        ),
        root=rubric_catalog_root,
    )
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
        rubric_catalog_root=rubric_catalog_root,
    )

    config_payload = json.loads(paths.editable_config_path.read_text(encoding="utf-8"))
    assert config_payload["event_mappings"] == []
    assert any(item.code == "mapeamento_evento_ausente" for item in result.pendings)


def test_assisted_report_importer_only_persists_after_explicit_apply(tmp_path: Path) -> None:
    employee_root = tmp_path / "employee_registries"
    rubric_root = tmp_path / "rubric_catalogs"
    report_bytes = "\n".join(
        [
            "Empresa: 900 - Empresa Relatorio",
            "Competencia: 04/2026",
            "Funcionario: Matricula: 123 Nome: Ana Lima",
            (
                "Rubrica: 350 - HORAS EXTRAS 50 "
                "Evento canonico: horas_extras_50 Tipo: horas Natureza: provento Total: 12,50"
            ),
        ]
    ).encode("utf-8")

    analysis = analyze_report_import(
        file_name="resumo.txt",
        file_bytes=report_bytes,
        selected_company_code="900",
        selected_company_name="Empresa Relatorio",
        employee_registry_root=employee_root,
        rubric_catalog_root=rubric_root,
    )

    assert analysis.blocked_reason is None
    assert not company_employee_registry_path("900", root=employee_root).exists()
    assert not company_rubric_catalog_path("900", root=rubric_root).exists()
    assert not list(tmp_path.rglob("*.txt"))

    employee_result = apply_report_employee_suggestions(
        analysis,
        selected_domain_registrations=("123",),
        root=employee_root,
    )

    assert employee_result.applied == 1
    assert load_company_employee_registry("900", root=employee_root).employees[0].employee_name == "Ana Lima"
    assert not company_rubric_catalog_path("900", root=rubric_root).exists()
    assert not list(tmp_path.rglob("*.txt"))

    rubric_result = apply_report_rubric_suggestions(
        analysis,
        selected_rubric_codes=("350",),
        root=rubric_root,
    )

    assert rubric_result.applied == 1
    assert load_company_rubric_catalog("900", root=rubric_root).rubrics[0].canonical_event == "horas_extras_50"
    assert not list(tmp_path.rglob("*.txt"))


def test_dashboard_returns_internal_pending_when_config_is_missing(tmp_path: Path) -> None:
    workbook_path = REPO_ROOT / "data" / "golden" / "v1" / "happy_path" / "input.xlsx"
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=tmp_path / "configs"),
    )

    assert result.summary.txt_enabled is False
    assert result.summary.config_status == ConfigResolutionStatus.NOT_FOUND.value
    assert any(item.stage == "configuracao" for item in result.pendings)


def test_dashboard_returns_internal_pending_when_config_is_ambiguous(tmp_path: Path) -> None:
    workbook_path = REPO_ROOT / "data" / "golden" / "v1" / "happy_path" / "input.xlsx"
    configs_root = tmp_path / "configs"
    _write_internal_config(configs_root, company_code="72", file_name="03-2024.json", competence="03/2024")
    _write_internal_config(configs_root, company_code="72", file_name="03_2024.json", competence="03/2024")
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )

    assert result.summary.txt_enabled is False
    assert result.summary.config_status == ConfigResolutionStatus.AMBIGUOUS.value
    assert any(item.stage == "configuracao" for item in result.pendings)


def test_dashboard_returns_internal_pending_when_config_is_mismatch(tmp_path: Path) -> None:
    workbook_path = REPO_ROOT / "data" / "golden" / "v1" / "happy_path" / "input.xlsx"
    configs_root = tmp_path / "configs"
    _write_internal_config(
        configs_root,
        company_code="72",
        file_name="03-2024.json",
        competence="03/2024",
        payload_override={"company_code": "99"},
    )
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
    )

    assert result.summary.txt_enabled is False
    assert result.summary.config_status == ConfigResolutionStatus.MISMATCH.value
    assert any(item.stage == "configuracao" for item in result.pendings)


def test_dashboard_normalizes_monthly_layout_before_canonical_ingestion(tmp_path: Path) -> None:
    profile_root = tmp_path / "profiles"
    _write_monthly_column_profile(profile_root)
    paths = create_dashboard_run_from_paths(MONTHLY_FIXTURE, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(paths, column_profile_root=profile_root)

    assert result.summary.company_code == "528"
    assert result.summary.competence == "03/2026"
    assert result.profile_resolution.status == "found"
    assert result.summary.config_status == ConfigResolutionStatus.FOUND.value
    assert paths.normalization_path.exists()
    assert paths.editable_workbook_path.exists()

    normalization_payload = json.loads(paths.normalization_path.read_text(encoding="utf-8"))
    assert normalization_payload["manifest"]["normalizer"] == "profile_column_mapping"
    assert normalization_payload["manifest"]["counts"]["source_cells_converted"] > 0


def test_dashboard_selected_company_context_preserves_company_528_flow(tmp_path: Path) -> None:
    profile_root = tmp_path / "profiles"
    _write_monthly_column_profile(profile_root)
    paths = create_dashboard_run_from_paths(MONTHLY_FIXTURE, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        column_profile_root=profile_root,
        selected_company_code="528",
        selected_company_name="FRIED FISH VILAREJO",
        selected_competence="03/2026",
    )

    assert result.summary.company_code == "528"
    assert result.summary.competence == "03/2026"
    assert result.profile_resolution.status == "found"
    assert result.summary.config_status == ConfigResolutionStatus.FOUND.value
    assert result.config_resolution.status == ConfigResolutionStatus.FOUND.value
    assert not any(item.code == "empresa_selecionada_divergente" for item in result.pendings)


def test_dashboard_uses_position_profile_when_monthly_contract_is_invalid(tmp_path: Path) -> None:
    workbook_path = _write_position_profile_workbook(tmp_path / "fechamento-gustavo.xlsx")
    profile_root = tmp_path / "profiles"
    _write_position_column_profile(profile_root)
    configs_root = tmp_path / "configs"
    _write_internal_config(
        configs_root,
        company_code="755",
        file_name="04-2026.json",
        competence="04/2026",
        payload_override={
            "company_name": "GUSTAVO LOPES LACERDA",
            "event_mappings": [
                {
                    "event_negocio": "horas_extras_70",
                    "rubrica_saida": "201",
                }
            ],
            "employee_mappings": [
                {
                    "source_employee_key": "304",
                    "source_employee_name": "ADILSON RAFAEL DE SOUSA",
                    "domain_registration": "304",
                }
            ],
        },
    )
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(
        paths,
        config_resolver=ConfigResolver(registry_root=tmp_path / "master", legacy_root=configs_root),
        column_profile_root=profile_root,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="042026",
    )

    normalization_payload = json.loads(paths.normalization_path.read_text(encoding="utf-8"))
    mapped_payload = json.loads(paths.mapped_artifact_path.read_text(encoding="utf-8"))

    assert result.profile_resolution.status == "found"
    assert result.summary.company_code == "755"
    assert result.summary.competence == "04/2026"
    assert result.summary.serialized_line_count == 1
    assert normalization_payload["manifest"]["normalizer"] == "profile_column_mapping"
    assert normalization_payload["manifest"]["columns"][0]["source_column_letter"] == "T"
    assert mapped_payload["mapped_movements"][0]["output_rubric"] == "201"


def test_dashboard_blocks_monthly_layout_when_column_profile_is_missing(tmp_path: Path) -> None:
    paths = create_dashboard_run_from_paths(MONTHLY_FIXTURE, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(paths, column_profile_root=tmp_path / "profiles")

    assert result.summary.txt_enabled is False
    assert result.profile_resolution.status == "missing"
    assert result.summary.company_code == "528"
    assert result.summary.competence == "03/2026"
    assert any(item.code == "column_mapping_profile_missing" for item in result.pendings)
    assert not paths.editable_workbook_path.exists()
    assert not paths.normalization_path.exists()


def test_dashboard_blocks_monthly_layout_when_column_profile_is_incomplete(tmp_path: Path) -> None:
    profile_root = tmp_path / "profiles"
    _write_monthly_column_profile(profile_root, omit_column="CONSUMO              COD. 266")
    paths = create_dashboard_run_from_paths(MONTHLY_FIXTURE, runs_root=tmp_path / "runs")

    result = run_dashboard_analysis(paths, column_profile_root=profile_root)

    assert result.summary.txt_enabled is False
    assert result.profile_resolution.status == "incomplete"
    assert "CONSUMO              COD. 266" in result.profile_resolution.missing_columns
    pending = next(item for item in result.pendings if item.code == "column_mapping_profile_incomplete")
    assert pending.source_column_name == "CONSUMO              COD. 266"
    assert not paths.editable_workbook_path.exists()


def test_dashboard_can_apply_column_profile_action_and_reprocess(tmp_path: Path) -> None:
    profile_root = tmp_path / "profiles"
    _write_monthly_column_profile(profile_root, omit_column="CONSUMO              COD. 266")
    paths = create_dashboard_run_from_paths(MONTHLY_FIXTURE, runs_root=tmp_path / "runs")

    initial = run_dashboard_analysis(paths, column_profile_root=profile_root)
    pending = next(item for item in initial.pendings if item.code == "column_mapping_profile_incomplete")

    action = apply_dashboard_action(
        paths,
        action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
        pending_uid=pending.uid,
        payload={
            "column_name": "CONSUMO              COD. 266",
            "rubrica_target": "266",
            "value_kind": "monetario",
            "generation_mode": "single_line",
            "ignore_zero": True,
            "ignore_text": True,
        },
        column_profile_root=profile_root,
    )
    updated = run_dashboard_analysis(paths, column_profile_root=profile_root)

    assert action.action_type == DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE
    assert updated.profile_resolution.status == "found"
    assert not any(item.code == "column_mapping_profile_incomplete" for item in updated.pendings)
    assert paths.normalization_path.exists()


def test_dashboard_turns_orphan_profile_row_into_editable_pending_and_reprocesses(tmp_path: Path) -> None:
    workbook_path = tmp_path / "orphan-row.xlsx"
    workbook = load_workbook(MONTHLY_FIXTURE)
    worksheet = workbook["mar 26"]
    worksheet["A20"] = None
    worksheet["B20"] = None
    worksheet["C20"] = "01:00"
    workbook.save(workbook_path)

    profile_root = tmp_path / "profiles"
    _write_monthly_column_profile(profile_root)
    paths = create_dashboard_run_from_paths(workbook_path, runs_root=tmp_path / "runs")

    initial = run_dashboard_analysis(paths, column_profile_root=profile_root)

    assert initial.summary.txt_enabled is False
    assert initial.summary.validation_status == "blocked"
    assert paths.editable_workbook_path.exists()
    pending = next(item for item in initial.pendings if item.code == "linha_com_lancamento_sem_colaborador")
    assert pending.stage == "ingestao"
    assert pending.severity == "bloqueante"
    assert pending.source_sheet == "mar 26"
    assert pending.source_cell == "A20"
    assert pending.source_row == 20
    assert pending.can_edit_workbook is True

    persisted = load_dashboard_run(paths)
    assert any(item.code == "linha_com_lancamento_sem_colaborador" for item in persisted.pendings)

    apply_workbook_cell_correction(
        paths,
        sheet_name=pending.source_sheet or "",
        cell=pending.source_cell or "",
        new_value="col-orphan-20",
        pending_uid=pending.uid,
    )
    updated = run_dashboard_analysis(paths, column_profile_root=profile_root)

    assert not any(item.code == "linha_com_lancamento_sem_colaborador" for item in updated.pendings)
    assert updated.profile_resolution.status == "found"
    assert paths.normalization_path.exists()
