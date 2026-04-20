from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from domain import PendingCode, ValueType
from ingestion import (
    FatalIngestionCode,
    TemplateV1IngestionError,
    ingest_and_fill_planilha_padrao_v1,
    ingest_fill_and_persist_planilha_padrao_v1,
    load_planilha_padrao_folha_v1,
    save_planilha_padrao_folha_v1,
)


def _prepare_base_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "planilha.xlsx"
    save_planilha_padrao_folha_v1(workbook_path)

    workbook = load_workbook(workbook_path)

    parametros = workbook["PARAMETROS"]
    parametros["B2"] = "72"
    parametros["B3"] = "Dela More"
    parametros["B4"] = "03/2024"
    parametros["B5"] = "mensal"
    parametros["B6"] = "11"
    parametros["B7"] = "v1"

    funcionarios = workbook["FUNCIONARIOS"]
    funcionarios["A2"] = "col-001"
    funcionarios["B2"] = "Ana Lima"
    funcionarios["E2"] = "123"
    funcionarios["H2"] = "ativo"
    funcionarios["I2"] = "sim"

    funcionarios["A3"] = "col-002"
    funcionarios["B3"] = "Bruno Souza"
    funcionarios["E3"] = ""
    funcionarios["H3"] = "ativo"
    funcionarios["I3"] = "sim"

    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B2"] = "col-001"
    lancamentos["C2"] = "Ana Lima"
    lancamentos["D2"] = "123"
    lancamentos["H2"] = 100
    lancamentos["G2"] = "02:30"

    workbook.save(workbook_path)
    return workbook_path


def test_ingestion_generates_monetary_and_hour_movements(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    assert len(result.movements) == 2
    monetary = next(movement for movement in result.movements if movement.event_name == "gratificacao")
    hourly = next(movement for movement in result.movements if movement.event_name == "horas_extras_50")

    assert monetary.value_type == ValueType.MONETARY
    assert monetary.amount_for_sheet() == "100"
    assert hourly.value_type == ValueType.HOURS
    assert hourly.quantity_for_sheet() == "02:30"


def test_ingestion_does_not_generate_movement_from_observacao_eventos_and_creates_pending(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B3"] = "col-001"
    lancamentos["C3"] = "Ana Lima"
    lancamentos["D3"] = "123"
    lancamentos["U3"] = "avaliar manualmente"
    workbook.save(workbook_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    assert len(result.movements) == 2
    assert any(pending.pending_code == PendingCode.AMBIGUOUS_EVENT_NOTE for pending in result.pendings)


def test_ingestion_creates_pending_for_vale_transporte(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B3"] = "col-001"
    lancamentos["C3"] = "Ana Lima"
    lancamentos["D3"] = "123"
    lancamentos["N3"] = "sim"
    workbook.save(workbook_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    assert any(pending.pending_code == PendingCode.NON_AUTOMATABLE_EVENT for pending in result.pendings)
    assert all(movement.event_name != "vale_transporte" for movement in result.movements)


def test_registration_equal_between_registry_and_human_line_has_no_conflict_pending(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    movement = next(movement for movement in result.movements if movement.employee_key == "col-001")

    assert movement.domain_registration == "123"
    assert not any(
        pending.pending_code in {PendingCode.DOMAIN_REGISTRATION_CONFLICT, PendingCode.DOMAIN_REGISTRATION_LINE_ONLY}
        for pending in result.pendings
    )


def test_registration_absent_on_line_and_present_in_registry_uses_registry_without_pending(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    workbook["LANCAMENTOS_FACEIS"]["D2"] = ""
    workbook.save(workbook_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    movement = next(movement for movement in result.movements if movement.employee_key == "col-001")

    assert movement.domain_registration == "123"
    assert not any(
        pending.pending_code == PendingCode.DOMAIN_REGISTRATION_LINE_ONLY for pending in result.pendings
    )


def test_registration_present_on_line_and_absent_in_registry_generates_pending(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B3"] = "col-002"
    lancamentos["C3"] = "Bruno Souza"
    lancamentos["D3"] = "222"
    lancamentos["H3"] = "50,00"
    workbook.save(workbook_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    movement = next(movement for movement in result.movements if movement.employee_key == "col-002")

    assert movement.domain_registration == "222"
    assert movement.blocked is True
    assert any(pending.pending_code == PendingCode.DOMAIN_REGISTRATION_LINE_ONLY for pending in result.pendings)


def test_registration_divergence_generates_blocking_pending(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    workbook["LANCAMENTOS_FACEIS"]["D2"] = "999"
    workbook.save(workbook_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    movement = next(movement for movement in result.movements if movement.employee_key == "col-001")
    assert movement.domain_registration is None
    assert movement.blocked is True
    assert any(pending.pending_code == PendingCode.DOMAIN_REGISTRATION_CONFLICT for pending in result.pendings)


def test_registry_inconsistency_for_employee_key_generates_pending(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    funcionarios = workbook["FUNCIONARIOS"]
    funcionarios["A4"] = "col-001"
    funcionarios["B4"] = "Ana Lima Duplicada"
    funcionarios["E4"] = "777"
    funcionarios["H4"] = "ativo"
    funcionarios["I4"] = "sim"
    workbook.save(workbook_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    assert any(pending.pending_code == PendingCode.EMPLOYEE_REGISTRY_INCONSISTENT for pending in result.pendings)
    movement = next(movement for movement in result.movements if movement.employee_key == "col-001")
    assert movement.blocked is True


def test_ingestion_creates_pending_for_missing_registration(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B3"] = "col-002"
    lancamentos["C3"] = "Bruno Souza"
    lancamentos["H3"] = "50,00"
    workbook.save(workbook_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    assert any(pending.pending_code == PendingCode.DOMAIN_REGISTRATION_MISSING for pending in result.pendings)
    blocked_movement = next(movement for movement in result.movements if movement.employee_key == "col-002")
    assert blocked_movement.blocked is True


def test_ingestion_rejects_invalid_payroll_type_as_structural_error(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    workbook["PARAMETROS"]["B5"] = "adiantamento"
    workbook.save(workbook_path)

    with pytest.raises(TemplateV1IngestionError) as exc_info:
        load_planilha_padrao_folha_v1(workbook_path)

    assert exc_info.value.code == FatalIngestionCode.PAYROLL_TYPE_NOT_SUPPORTED


def test_ingestion_rejects_invalid_competence_as_structural_error(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    workbook["PARAMETROS"]["B4"] = "2024-03"
    workbook.save(workbook_path)

    with pytest.raises(TemplateV1IngestionError) as exc_info:
        load_planilha_padrao_folha_v1(workbook_path)

    assert exc_info.value.code == FatalIngestionCode.COMPETENCE_INVALID


def test_pending_operational_case_does_not_raise_fatal_error(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B3"] = "col-002"
    lancamentos["C3"] = "Bruno Souza"
    lancamentos["H3"] = "50,00"
    workbook.save(workbook_path)

    result = load_planilha_padrao_folha_v1(workbook_path)

    assert len(result.pendings) >= 1


def test_ingestion_writes_technical_tabs(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B3"] = "col-002"
    lancamentos["C3"] = "Bruno Souza"
    lancamentos["D3"] = "222"
    lancamentos["H3"] = "50,00"
    lancamentos["N3"] = "revisar"
    workbook.save(workbook_path)

    result = ingest_and_fill_planilha_padrao_v1(workbook_path)
    workbook = load_workbook(workbook_path)

    movimentos = workbook["MOVIMENTOS_CANONICOS"]
    pendencias = workbook["PENDENCIAS"]

    assert len(result.movements) >= 2
    assert movimentos["A2"].value is not None
    assert movimentos["I2"].value in {"horas_extras_50", "gratificacao"}
    assert pendencias["A2"].value is not None
    assert "!" in pendencias["I2"].value


def test_pipeline_persists_snapshot_and_manifest(tmp_path: Path):
    workbook_path = _prepare_base_workbook(tmp_path)

    artifacts = ingest_fill_and_persist_planilha_padrao_v1(workbook_path)

    assert artifacts.snapshot_path.exists()
    assert artifacts.manifest_path is not None
    assert artifacts.manifest_path.exists()

    snapshot_payload = json.loads(artifacts.snapshot_path.read_text(encoding="utf-8"))
    manifest_payload = json.loads(artifacts.manifest_path.read_text(encoding="utf-8"))

    assert snapshot_payload["snapshot_version"] == "ingestion_snapshot_v1"
    assert "parameters" in snapshot_payload
    assert "movements" in snapshot_payload
    assert "pendings" in snapshot_payload
    assert manifest_payload["movement_count"] == len(artifacts.result.movements)
    assert manifest_payload["pending_count"] == len(artifacts.result.pendings)
    assert "canonical_snapshot" in manifest_payload["artifact_hashes"]
