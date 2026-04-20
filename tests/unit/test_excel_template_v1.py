from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ingestion.template_v1 import (
    FUNCIONARIOS_HEADERS,
    LANCAMENTOS_FACEIS_HEADERS,
    MOVIMENTOS_CANONICOS_HEADERS,
    PARAMETROS_HEADERS,
    PENDENCIAS_HEADERS,
    SHEET_ORDER,
    TEMPLATE_V1_FILENAME,
    create_planilha_padrao_folha_v1,
    save_planilha_padrao_folha_v1,
)


def _header_values(worksheet, total_columns: int) -> tuple[str, ...]:
    return tuple(worksheet.cell(row=1, column=index).value for index in range(1, total_columns + 1))


def test_template_workbook_has_expected_sheets_and_headers():
    workbook = create_planilha_padrao_folha_v1()

    assert tuple(workbook.sheetnames) == SHEET_ORDER
    assert _header_values(workbook["PARAMETROS"], len(PARAMETROS_HEADERS)) == PARAMETROS_HEADERS
    assert _header_values(workbook["FUNCIONARIOS"], len(FUNCIONARIOS_HEADERS)) == FUNCIONARIOS_HEADERS
    assert _header_values(
        workbook["LANCAMENTOS_FACEIS"], len(LANCAMENTOS_FACEIS_HEADERS)
    ) == LANCAMENTOS_FACEIS_HEADERS
    assert _header_values(
        workbook["MOVIMENTOS_CANONICOS"], len(MOVIMENTOS_CANONICOS_HEADERS)
    ) == MOVIMENTOS_CANONICOS_HEADERS
    assert _header_values(workbook["PENDENCIAS"], len(PENDENCIAS_HEADERS)) == PENDENCIAS_HEADERS


def test_template_workbook_configures_main_human_sheet():
    workbook = create_planilha_padrao_folha_v1()
    worksheet = workbook["LANCAMENTOS_FACEIS"]

    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:U1000"
    assert worksheet["N1"].value == "vale_transporte"

    validations = list(worksheet.data_validations.dataValidation)
    formulas = {validation.formula1 for validation in validations}
    assert '"sim,nao,parcial,revisar"' in formulas
    assert "0" in formulas
    assert any("LEN(G2)=5" in formula for formula in formulas)


def test_template_file_is_generated_and_loadable(tmp_path: Path):
    output_path = tmp_path / TEMPLATE_V1_FILENAME
    save_planilha_padrao_folha_v1(output_path)

    assert output_path.exists()

    workbook = load_workbook(output_path)
    assert tuple(workbook.sheetnames) == SHEET_ORDER


def test_versioned_template_artifact_exists_in_repo():
    artifact_path = Path("data") / "templates" / TEMPLATE_V1_FILENAME

    assert artifact_path.exists()

    workbook = load_workbook(artifact_path)
    assert tuple(workbook.sheetnames) == SHEET_ORDER
