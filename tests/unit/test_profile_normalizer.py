from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from dashboard.column_mapping_profiles import (
    ColumnGenerationMode,
    ColumnMappingRule,
    ColumnValueKind,
    CompanyColumnMappingProfile,
)
from dashboard.company_employee_registry import (
    CompanyEmployeeRecord,
    CompanyEmployeeRegistry,
    save_company_employee_registry,
)
from dashboard.company_rubric_catalog import (
    CompanyRubricCatalog,
    CompanyRubricRecord,
    save_company_rubric_catalog,
)
from dashboard.profile_normalizer import (
    build_canonical_workbook_from_column_profile,
    inspect_workbook_with_position_profile,
    normalize_workbook_with_column_profile,
)
from ingestion import (
    InputColumnMetadata,
    InputLayoutDetection,
    InputLayoutNormalizationError,
    InputWorkbookInspection,
    MONTHLY_LAYOUT_ID,
    ingest_template_v1_workbook,
)


def _build_profile_source_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "mar 26"
    sheet["A1"] = "Empresa Teste"
    sheet["A2"] = "Lancamentos marco 2026"
    headers = (
        "COD.",
        "NOME",
        "GRAT.",
        "ATRASO",
        "FALTA",
        "ADIANT. QUINZ",
        "GRAT. ZERO",
        "GRAT. TEXTO",
        "MATRICULA",
    )
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column_index, value=header)

    row = ("col-001", "Ana Lima", "150,00", "01:30", 2, "Sim", 0, "ferias", "123")
    for column_index, value in enumerate(row, start=1):
        sheet.cell(row=6, column=column_index, value=value)
    return workbook


def _build_inspection() -> InputWorkbookInspection:
    detection = InputLayoutDetection(
        layout_id=MONTHLY_LAYOUT_ID,
        active_sheet_name="mar 26",
        selected_sheet_name="mar 26",
        selected_sheet_reason="active_sheet",
        company_code="72",
        company_name="Empresa Teste",
        competence="03/2026",
        source_company_name="Empresa Teste",
        source_title_text="Lancamentos marco 2026",
        source_sheet_names=("mar 26",),
        rules_applied=("test_contract",),
    )
    headers = (
        "COD.",
        "NOME",
        "GRAT.",
        "ATRASO",
        "FALTA",
        "ADIANT. QUINZ",
        "GRAT. ZERO",
        "GRAT. TEXTO",
        "MATRICULA",
    )
    columns = tuple(
        InputColumnMetadata(
            column_index=index,
            column_letter=letter,
            column_name=header,
            normalized_column_name=header.lower(),
            header_row=4,
        )
        for index, letter, header in (
            (1, "A", headers[0]),
            (2, "B", headers[1]),
            (3, "C", headers[2]),
            (4, "D", headers[3]),
            (5, "E", headers[4]),
            (6, "F", headers[5]),
            (7, "G", headers[6]),
            (8, "H", headers[7]),
            (9, "I", headers[8]),
        )
    )
    return InputWorkbookInspection(
        layout_id=MONTHLY_LAYOUT_ID,
        company_code="72",
        company_name="Empresa Teste",
        competence="03/2026",
        selected_sheet_name="mar 26",
        source_sheet_names=("mar 26",),
        columns=columns,
        warnings=(),
        detection=detection,
    )


def _build_profile() -> CompanyColumnMappingProfile:
    return CompanyColumnMappingProfile(
        company_code="72",
        company_name="Empresa Teste",
        default_process="11",
        mappings=[
            ColumnMappingRule(
                column_name="GRAT.",
                enabled=True,
                rubrica_target="20",
                value_kind=ColumnValueKind.MONETARY,
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
            ColumnMappingRule(
                column_name="ATRASO",
                enabled=True,
                rubrica_target="8069",
                value_kind=ColumnValueKind.HOURS,
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
            ColumnMappingRule(
                column_name="FALTA",
                enabled=True,
                rubricas_target=["8792", "8794"],
                value_kind=ColumnValueKind.QUANTITY,
                generation_mode=ColumnGenerationMode.MULTI_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
            ColumnMappingRule(
                column_name="ADIANT. QUINZ",
                enabled=False,
                value_kind=ColumnValueKind.MONETARY,
                generation_mode=ColumnGenerationMode.IGNORE,
                ignore_zero=True,
                ignore_text=True,
            ),
            ColumnMappingRule(
                column_name="GRAT. ZERO",
                enabled=True,
                rubrica_target="20",
                value_kind=ColumnValueKind.MONETARY,
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
            ColumnMappingRule(
                column_name="GRAT. TEXTO",
                enabled=True,
                rubrica_target="20",
                value_kind=ColumnValueKind.MONETARY,
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
        ],
    )


def _non_empty_launch_rows(workbook) -> list[int]:
    worksheet = workbook["LANCAMENTOS_FACEIS"]
    rows = []
    for row_number in range(2, 30):
        if worksheet[f"B{row_number}"].value is not None:
            rows.append(row_number)
    return rows


def test_profile_normalizer_translates_supported_column_rules_to_canonical_workbook() -> None:
    normalized, manifest = build_canonical_workbook_from_column_profile(
        _build_profile_source_workbook(),
        inspection=_build_inspection(),
        profile=_build_profile(),
    )

    lancamentos = normalized["LANCAMENTOS_FACEIS"]
    assert _non_empty_launch_rows(normalized) == [2, 3, 4, 5]
    assert lancamentos["B2"].value == "col-001"
    assert lancamentos["C2"].value == "Ana Lima"
    assert lancamentos["D2"].value == "123"
    assert lancamentos["H2"].value == "150"
    assert lancamentos["S3"].value == "01:30"
    assert lancamentos["R4"].value == "2"
    assert lancamentos["Y5"].value == "2"
    assert "rubrica_target=8792" in lancamentos["F4"].value
    assert "rubrica_target=8794" in lancamentos["F5"].value

    assert manifest["normalizer"] == "profile_column_mapping"
    assert manifest["counts"]["source_cells_converted"] == 3
    assert manifest["counts"]["generated_movements"] == 4
    assert manifest["counts"]["ignored_cells"] == 1
    assert manifest["counts"]["skipped_zero_cells"] == 1
    assert manifest["counts"]["skipped_text_cells"] == 1


def test_profile_normalizer_output_remains_ingestable_by_v1_loader() -> None:
    normalized, _manifest = build_canonical_workbook_from_column_profile(
        _build_profile_source_workbook(),
        inspection=_build_inspection(),
        profile=_build_profile(),
    )

    result = ingest_template_v1_workbook(normalized)

    assert [movement.event_name for movement in result.movements] == [
        "gratificacao",
        "atrasos_horas",
        "faltas_dias",
        "faltas_dsr",
    ]
    assert result.movements[0].amount_for_sheet() == "150"
    assert result.movements[0].domain_registration == "123"
    assert result.movements[1].quantity_for_sheet() == "01:30"
    assert result.movements[2].quantity_for_sheet() == "2"
    assert result.movements[3].quantity_for_sheet() == "2"


def test_profile_normalizer_persists_canonical_workbook_and_report(tmp_path: Path) -> None:
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "canonical.xlsx"
    report_path = tmp_path / "normalization.json"
    _build_profile_source_workbook().save(source_path)

    result = normalize_workbook_with_column_profile(
        source_path,
        inspection=_build_inspection(),
        profile=_build_profile(),
        output_path=output_path,
        report_path=report_path,
    )

    assert result.workbook_path == output_path
    assert result.report_path == report_path
    assert result.canonical_rows_written == 4
    assert output_path.exists()
    assert report_path.exists()

    persisted = load_workbook(output_path)
    assert persisted["PARAMETROS"]["B2"].value == "72"
    assert persisted["PARAMETROS"]["B6"].value == "11"
    assert persisted["FUNCIONARIOS"]["E2"].value == "123"


def test_profile_normalizer_routes_critical_rubrics_to_distinct_canonical_columns() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "mar 26"
    sheet["A1"] = "Empresa Teste"
    headers = (
        "COD.",
        "NOME",
        "EXTRA 70%",
        "EXTRA 100%",
        "EXTRA NOTURNA",
        "DESPESAS",
        "MATRICULA",
    )
    values = ("col-001", "Ana Lima", "01:00", "02:00", "03:00", "10,00", "123")
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=4, column=column_index, value=header)
        sheet.cell(row=6, column=column_index, value=values[column_index - 1])

    detection = InputLayoutDetection(
        layout_id=MONTHLY_LAYOUT_ID,
        active_sheet_name="mar 26",
        selected_sheet_name="mar 26",
        selected_sheet_reason="unit_test",
        company_code="72",
        company_name="Empresa Teste",
        competence="03/2026",
        source_company_name="Empresa Teste",
        source_title_text=None,
        source_sheet_names=("mar 26",),
        rules_applied=("unit_test",),
    )
    inspection = InputWorkbookInspection(
        layout_id=MONTHLY_LAYOUT_ID,
        company_code="72",
        company_name="Empresa Teste",
        competence="03/2026",
        selected_sheet_name="mar 26",
        source_sheet_names=("mar 26",),
        columns=tuple(
            InputColumnMetadata(
                column_index=index,
                column_letter=letter,
                column_name=header,
                normalized_column_name=header.lower(),
                header_row=4,
            )
            for index, letter, header in (
                (1, "A", headers[0]),
                (2, "B", headers[1]),
                (3, "C", headers[2]),
                (4, "D", headers[3]),
                (5, "E", headers[4]),
                (6, "F", headers[5]),
                (7, "G", headers[6]),
            )
        ),
        warnings=(),
        detection=detection,
    )
    profile = CompanyColumnMappingProfile(
        company_code="72",
        company_name="Empresa Teste",
        default_process="11",
        mappings=[
            ColumnMappingRule(
                column_name="EXTRA 70%",
                enabled=True,
                rubrica_target="201",
                value_kind=ColumnValueKind.HOURS,
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
            ColumnMappingRule(
                column_name="EXTRA 100%",
                enabled=True,
                rubrica_target="200",
                value_kind=ColumnValueKind.HOURS,
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
            ColumnMappingRule(
                column_name="EXTRA NOTURNA",
                enabled=True,
                rubrica_target="25",
                value_kind=ColumnValueKind.HOURS,
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
            ColumnMappingRule(
                column_name="DESPESAS",
                enabled=True,
                rubrica_target="204",
                value_kind=ColumnValueKind.MONETARY,
                generation_mode=ColumnGenerationMode.SINGLE_LINE,
                ignore_zero=True,
                ignore_text=True,
            ),
        ],
    )

    normalized, _manifest = build_canonical_workbook_from_column_profile(
        workbook,
        inspection=inspection,
        profile=profile,
    )

    lancamentos = normalized["LANCAMENTOS_FACEIS"]
    assert lancamentos["V2"].value == "01:00"
    assert lancamentos["W3"].value == "02:00"
    assert lancamentos["X4"].value == "03:00"
    assert lancamentos["P5"].value == "10"
    assert lancamentos["H2"].value is None


def _build_position_profile(
    header: str = "HORA 50%",
    *,
    row_control_column: str | None = None,
    ignore_row_tokens: tuple[str, ...] = (),
    stop_row_tokens: tuple[str, ...] = (),
) -> CompanyColumnMappingProfile:
    return CompanyColumnMappingProfile(
        company_code="755",
        company_name="GUSTAVO LOPES LACERDA",
        default_process="11",
        mappings=[
            ColumnMappingRule(
                sheet_name="abril",
                header_row=2,
                data_start_row=3,
                employee_code_column="A",
                employee_name_column="B",
                row_control_column=row_control_column,
                ignore_row_tokens=list(ignore_row_tokens),
                stop_row_tokens=list(stop_row_tokens),
                value_column="T",
                expected_header=header,
                enabled=True,
                rubrica_target="201",
                value_kind="horas",
                nature="provento",
                generation_mode="single_line",
                ignore_zero=True,
                ignore_text=True,
                status="active",
            )
        ],
    )


def _build_position_workbook(
    header: str = "HORA 50%",
    *,
    employee_code_header: str = "CODIGO",
    employee_name_header: str = "NOME",
    employee_code_value: str = "304",
    employee_name_value: str = "ADILSON RAFAEL DE SOUSA",
    rows: list[tuple[object, object, object]] | None = None,
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "abril"
    sheet["A2"] = employee_code_header
    sheet["B2"] = employee_name_header
    sheet["T2"] = header
    source_rows = rows or [(employee_code_value, employee_name_value, "01:30")]
    for row_offset, (row_employee_code, row_employee_name, row_value) in enumerate(source_rows, start=3):
        sheet.cell(row=row_offset, column=1, value=row_employee_code)
        sheet.cell(row=row_offset, column=2, value=row_employee_name)
        sheet.cell(row=row_offset, column=20, value=row_value)
    return workbook


def _write_active_registry(root: Path, employees: list[CompanyEmployeeRecord]) -> None:
    save_company_employee_registry(
        CompanyEmployeeRegistry(
            company_code="755",
            company_name="GUSTAVO LOPES LACERDA",
            employees=employees,
        ),
        root=root,
    )


def _write_active_rubric_catalog(root: Path, rubrics: list[CompanyRubricRecord]) -> None:
    save_company_rubric_catalog(
        CompanyRubricCatalog(
            company_code="755",
            company_name="GUSTAVO LOPES LACERDA",
            rubrics=rubrics,
        ),
        root=root,
    )


def test_position_profile_validates_expected_header_and_generates_rubric_column(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook().save(workbook_path)
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="horas_extras_70",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile()
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )

    assert inspection.columns[0].column_letter == "T"
    assert manifest["counts"]["generated_movements"] == 1
    assert normalized["FUNCIONARIOS"]["E2"].value == "304"
    assert normalized["MOVIMENTOS_CANONICOS"]["I2"].value == "201"
    assert normalized["MOVIMENTOS_CANONICOS"]["J2"].value == "201"
    assert normalized["MOVIMENTOS_CANONICOS"]["N2"].value == "01:30"


def test_position_profile_blocks_when_expected_header_differs(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook(header="HORAS 100%").save(workbook_path)
    profile = _build_position_profile(header="HORA 50%")
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        build_canonical_workbook_from_column_profile(
            load_workbook(workbook_path),
            inspection=inspection,
            profile=profile,
            employee_registry_root=tmp_path / "employees",
        )

    assert exc_info.value.code == "cabecalho_perfil_divergente"
    assert "A coluna T esperava HORA 50%" in str(exc_info.value)


def test_position_profile_resolves_employee_by_name_when_registration_is_blank(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook(employee_code_value="", employee_name_value="ADILSON RAFAEL DE SOUSA").save(workbook_path)
    _write_active_registry(
        tmp_path / "employees",
        [
            CompanyEmployeeRecord(
                employee_name="ADILSON RAFAEL DE SOUSA",
                domain_registration="304",
                source="test",
            )
        ],
    )
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="horas_extras_70",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile()
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, _manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )

    assert normalized["FUNCIONARIOS"]["B2"].value == "ADILSON RAFAEL DE SOUSA"
    assert normalized["FUNCIONARIOS"]["E2"].value == "304"
    assert normalized["MOVIMENTOS_CANONICOS"]["H2"].value == "304"


def test_position_profile_resolves_employee_by_name_when_registration_column_has_text(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook(
        employee_code_value="ADILSON RAFAEL DE SOUSA",
        employee_name_value=" Adilson  Rafael de Sousa ",
    ).save(workbook_path)
    _write_active_registry(
        tmp_path / "employees",
        [
            CompanyEmployeeRecord(
                employee_name="ADILSON RAFAEL DE SOUSA",
                domain_registration="304",
                source="test",
            )
        ],
    )
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="horas_extras_70",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile()
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, _manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )

    assert normalized["FUNCIONARIOS"]["E2"].value == "304"
    assert normalized["FUNCIONARIOS"]["B2"].value == "ADILSON RAFAEL DE SOUSA"


def test_position_profile_resolves_employee_by_alias_when_registration_is_invalid(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook(
        employee_code_value="DANIELA BOTURA",
        employee_name_value=" daniela  botura ",
    ).save(workbook_path)
    _write_active_registry(
        tmp_path / "employees",
        [
            CompanyEmployeeRecord(
                employee_name="DANIELA PRISCILLA BOTURA MONTEIRO",
                domain_registration="384",
                aliases=["DANIELA BOTURA"],
                source="test",
            )
        ],
    )
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="horas_extras_70",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile()
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, _manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )

    assert normalized["FUNCIONARIOS"]["E2"].value == "384"
    assert normalized["FUNCIONARIOS"]["B2"].value == "DANIELA PRISCILLA BOTURA MONTEIRO"
    assert normalized["MOVIMENTOS_CANONICOS"]["H2"].value == "384"


def test_position_profile_blocks_when_employee_name_is_not_found(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook(employee_code_value="", employee_name_value="FUNCIONARIO INEXISTENTE").save(workbook_path)
    _write_active_registry(
        tmp_path / "employees",
        [
            CompanyEmployeeRecord(
                employee_name="ADILSON RAFAEL DE SOUSA",
                domain_registration="304",
                source="test",
            )
        ],
    )
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="horas_extras_70",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile()
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        build_canonical_workbook_from_column_profile(
            load_workbook(workbook_path),
            inspection=inspection,
            profile=profile,
            employee_registry_root=tmp_path / "employees",
            rubric_catalog_root=tmp_path / "rubrics",
        )

    assert exc_info.value.code == "funcionario_nome_nao_encontrado"
    assert "Funcionario 'FUNCIONARIO INEXISTENTE' nao encontrado no cadastro ativo da empresa." in str(exc_info.value)


def test_position_profile_blocks_when_employee_name_is_ambiguous(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook(employee_code_value="", employee_name_value="ADILSON RAFAEL DE SOUSA").save(workbook_path)
    _write_active_registry(
        tmp_path / "employees",
        [
            CompanyEmployeeRecord(
                employee_name="ADILSON RAFAEL DE SOUSA",
                domain_registration="304",
                source="test",
            ),
            CompanyEmployeeRecord(
                employee_name="ADILSON RAFAEL DE SOUSA",
                domain_registration="999",
                aliases=["ADILSON RAFAEL DE SOUSA"],
                source="test",
            ),
        ],
    )
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="horas_extras_70",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile()
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        build_canonical_workbook_from_column_profile(
            load_workbook(workbook_path),
            inspection=inspection,
            profile=profile,
            employee_registry_root=tmp_path / "employees",
            rubric_catalog_root=tmp_path / "rubrics",
        )

    assert exc_info.value.code == "funcionario_nome_ambiguo"
    assert "encontrou mais de um cadastro compativel" in str(exc_info.value)


def test_position_profile_blocks_when_employee_alias_is_ambiguous(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook(employee_code_value="", employee_name_value="DANIELA BOTURA").save(workbook_path)
    _write_active_registry(
        tmp_path / "employees",
        [
            CompanyEmployeeRecord(
                employee_name="DANIELA PRISCILLA BOTURA MONTEIRO",
                domain_registration="384",
                aliases=["DANIELA BOTURA"],
                source="test",
            ),
            CompanyEmployeeRecord(
                employee_name="DANIELA BOTURA SILVA",
                domain_registration="999",
                aliases=["daniela botura"],
                source="test",
            ),
        ],
    )
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="horas_extras_70",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile()
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        build_canonical_workbook_from_column_profile(
            load_workbook(workbook_path),
            inspection=inspection,
            profile=profile,
            employee_registry_root=tmp_path / "employees",
            rubric_catalog_root=tmp_path / "rubrics",
        )

    assert exc_info.value.code == "funcionario_nome_ambiguo"


def test_position_profile_blocks_without_valid_registration_and_without_name_column(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook(employee_code_value="ADILSON RAFAEL DE SOUSA").save(workbook_path)
    profile = CompanyColumnMappingProfile(
        company_code="755",
        company_name="GUSTAVO LOPES LACERDA",
        default_process="11",
        mappings=[
            ColumnMappingRule(
                sheet_name="abril",
                header_row=2,
                data_start_row=3,
                employee_code_column="A",
                value_column="T",
                expected_header="HORA 50%",
                enabled=True,
                rubrica_target="201",
                value_kind="horas",
                nature="provento",
                generation_mode="single_line",
                ignore_zero=True,
                ignore_text=True,
                status="active",
            )
        ],
    )
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        build_canonical_workbook_from_column_profile(
            load_workbook(workbook_path),
            inspection=inspection,
            profile=profile,
            employee_registry_root=tmp_path / "employees",
            rubric_catalog_root=tmp_path / "rubrics",
        )

    assert exc_info.value.code == "perfil_colunas_sem_nome_para_resolver_funcionario"
    assert "Linha sem matricula valida e sem coluna de nome configurada no perfil." in str(exc_info.value)


def test_position_profile_generates_direct_rubric_without_canonical_column_map(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    workbook = _build_position_workbook(header="QUEBRA DE CAIXA")
    workbook.active["T3"] = "100,00"
    workbook.save(workbook_path)
    profile = CompanyColumnMappingProfile(
        company_code="755",
        company_name="GUSTAVO LOPES LACERDA",
        default_process="11",
        mappings=[
            ColumnMappingRule(
                sheet_name="abril",
                header_row=2,
                data_start_row=3,
                employee_code_column="A",
                employee_name_column="B",
                value_column="T",
                expected_header="QUEBRA DE CAIXA",
                enabled=True,
                rubrica_target="8907",
                value_kind="monetario",
                nature="provento",
                generation_mode="single_line",
                ignore_zero=True,
                ignore_text=True,
                status="active",
            )
        ],
    )
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="8907",
                description="QUEBRA DE CAIXA",
                canonical_event="8907",
                value_kind="monetario",
                nature="provento",
                source="test",
            )
        ],
    )
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, _manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )
    result = ingest_template_v1_workbook(normalized)

    assert len(result.movements) == 1
    assert result.movements[0].event_name == "8907"
    assert result.movements[0].informed_rubric == "8907"
    assert result.movements[0].output_rubric == "8907"
    assert result.movements[0].amount_for_sheet() == "100"


def test_position_profile_blocks_when_direct_rubric_is_missing_from_active_catalog(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    workbook = _build_position_workbook(header="QUEBRA DE CAIXA")
    workbook.active["T3"] = "100,00"
    workbook.save(workbook_path)
    profile = CompanyColumnMappingProfile(
        company_code="755",
        company_name="GUSTAVO LOPES LACERDA",
        default_process="11",
        mappings=[
            ColumnMappingRule(
                sheet_name="abril",
                header_row=2,
                data_start_row=3,
                employee_code_column="A",
                employee_name_column="B",
                value_column="T",
                expected_header="QUEBRA DE CAIXA",
                enabled=True,
                rubrica_target="8907",
                value_kind="monetario",
                nature="provento",
                generation_mode="single_line",
                ignore_zero=True,
                ignore_text=True,
                status="active",
            )
        ],
    )
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        build_canonical_workbook_from_column_profile(
            load_workbook(workbook_path),
            inspection=inspection,
            profile=profile,
            employee_registry_root=tmp_path / "employees",
            rubric_catalog_root=tmp_path / "rubrics",
        )

    assert exc_info.value.code == "rubrica_perfil_inexistente_no_catalogo"
    assert "Rubrica 8907 nao existe no catalogo ativo da empresa." in str(exc_info.value)


def test_position_profile_multi_line_generates_direct_rubrics_without_canonical_column_map(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    workbook = _build_position_workbook(header="FALTAS INJUSTIFICADAS")
    workbook.active["T3"] = "2"
    workbook.save(workbook_path)
    profile = CompanyColumnMappingProfile(
        company_code="755",
        company_name="GUSTAVO LOPES LACERDA",
        default_process="11",
        mappings=[
            ColumnMappingRule(
                sheet_name="abril",
                header_row=2,
                data_start_row=3,
                employee_code_column="A",
                employee_name_column="B",
                value_column="T",
                expected_header="FALTAS INJUSTIFICADAS",
                enabled=True,
                rubricas_target=["8792", "8794"],
                value_kind="quantidade",
                nature="desconto",
                generation_mode="multi_line",
                ignore_zero=True,
                ignore_text=True,
                status="active",
            )
        ],
    )
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="8792",
                description="DIAS FALTAS",
                canonical_event="8792",
                value_kind="quantidade",
                nature="desconto",
                source="test",
            ),
            CompanyRubricRecord(
                rubric_code="8794",
                description="DIAS FALTAS DSR",
                canonical_event="8794",
                value_kind="quantidade",
                nature="desconto",
                source="test",
            ),
        ],
    )
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, _manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )
    result = ingest_template_v1_workbook(normalized)

    assert [(movement.event_name, movement.output_rubric) for movement in result.movements] == [
        ("8792", "8792"),
        ("8794", "8794"),
    ]


def test_position_profile_blocks_when_single_line_rule_has_no_rubrica_target(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento.xlsx"
    _build_position_workbook().save(workbook_path)
    invalid_rule = ColumnMappingRule.model_construct(
        column_name=None,
        column_key=None,
        sheet_name="abril",
        header_row=2,
        data_start_row=3,
        employee_code_column="A",
        employee_name_column="B",
        value_column="T",
        expected_header="HORA 50%",
        enabled=True,
        rubrica_target=None,
        rubricas_target=[],
        value_kind=ColumnValueKind.HOURS,
        nature="provento",
        generation_mode=ColumnGenerationMode.SINGLE_LINE,
        ignore_zero=True,
        ignore_text=True,
        status="active",
        notes=None,
    )
    profile = CompanyColumnMappingProfile.model_construct(
        company_code="755",
        company_name="GUSTAVO LOPES LACERDA",
        default_process="11",
        mappings=[invalid_rule],
        profile_version="column-mapping-v1",
        notes=None,
    )
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=CompanyColumnMappingProfile(
            company_code="755",
            company_name="GUSTAVO LOPES LACERDA",
            default_process="11",
            mappings=[
                ColumnMappingRule(
                    sheet_name="abril",
                    header_row=2,
                    data_start_row=3,
                    employee_code_column="A",
                    employee_name_column="B",
                    value_column="T",
                    expected_header="HORA 50%",
                    enabled=True,
                    rubrica_target="201",
                    value_kind="horas",
                    nature="provento",
                    generation_mode="single_line",
                    ignore_zero=True,
                    ignore_text=True,
                    status="active",
                )
            ],
        ),
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        build_canonical_workbook_from_column_profile(
            load_workbook(workbook_path),
            inspection=inspection,
            profile=profile,
            employee_registry_root=tmp_path / "employees",
            rubric_catalog_root=tmp_path / "rubrics",
        )

    assert exc_info.value.code == "perfil_coluna_sem_rubrica_unica"
    assert "Regra de perfil da coluna T nao possui rubrica unica configurada." in str(exc_info.value)


def test_position_profile_stops_before_resolving_control_row(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento-stop.xlsx"
    _build_position_workbook(
        rows=[
            ("304", "ADILSON RAFAEL DE SOUSA", "01:30"),
            (" total ", None, None),
            ("399", "ADRIANE MARTINS DE ALMEIDA", "02:00"),
        ]
    ).save(workbook_path)
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="201",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile(row_control_column="A", stop_row_tokens=("TOTAL",))
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )

    assert manifest["counts"]["employee_rows_written"] == 1
    assert normalized["FUNCIONARIOS"]["B2"].value == "ADILSON RAFAEL DE SOUSA"
    assert normalized["FUNCIONARIOS"]["B3"].value is None
    assert normalized["MOVIMENTOS_CANONICOS"]["H2"].value == "304"
    assert normalized["MOVIMENTOS_CANONICOS"]["H3"].value is None


def test_position_profile_ignores_control_row_and_continues_processing(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento-ignore.xlsx"
    _build_position_workbook(
        rows=[
            (" Tótal ", None, None),
            ("399", "ADRIANE MARTINS DE ALMEIDA", "02:00"),
        ]
    ).save(workbook_path)
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="201",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile(row_control_column="A", ignore_row_tokens=("TOTAL",))
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )

    assert manifest["counts"]["employee_rows_written"] == 1
    assert normalized["FUNCIONARIOS"]["B2"].value == "ADRIANE MARTINS DE ALMEIDA"
    assert normalized["FUNCIONARIOS"]["E2"].value == "399"
    assert normalized["MOVIMENTOS_CANONICOS"]["N2"].value == "02:00"


def test_position_profile_without_control_rule_keeps_previous_employee_resolution_behavior(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento-sem-controle.xlsx"
    _build_position_workbook(rows=[("TOTAL", None, None)]).save(workbook_path)
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="201",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile()
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        build_canonical_workbook_from_column_profile(
            load_workbook(workbook_path),
            inspection=inspection,
            profile=profile,
            employee_registry_root=tmp_path / "employees",
            rubric_catalog_root=tmp_path / "rubrics",
        )

    assert exc_info.value.code == "funcionario_nome_nao_encontrado"


def test_position_profile_total_control_row_does_not_generate_employee_pending(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fechamento-total.xlsx"
    _build_position_workbook(rows=[("TOTAL", None, None)]).save(workbook_path)
    _write_active_rubric_catalog(
        tmp_path / "rubrics",
        [
            CompanyRubricRecord(
                rubric_code="201",
                description="HORA 50%",
                canonical_event="201",
                value_kind="horas",
                nature="provento",
                source="test",
            )
        ],
    )
    profile = _build_position_profile(row_control_column="A", stop_row_tokens=("TOTAL",))
    inspection = inspect_workbook_with_position_profile(
        workbook_path,
        profile=profile,
        selected_company_code="755",
        selected_company_name="GUSTAVO LOPES LACERDA",
        selected_competence="04/2026",
    )

    normalized, manifest = build_canonical_workbook_from_column_profile(
        load_workbook(workbook_path),
        inspection=inspection,
        profile=profile,
        employee_registry_root=tmp_path / "employees",
        rubric_catalog_root=tmp_path / "rubrics",
    )

    assert manifest["counts"]["employee_rows_written"] == 0
    assert normalized["FUNCIONARIOS"]["B2"].value is None
