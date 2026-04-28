from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from ingestion import (
    CANONICAL_LAYOUT_ID,
    MONTHLY_LAYOUT_ID,
    InputLayoutNormalizationError,
    detect_input_layout,
    inspect_input_workbook,
    load_planilha_padrao_folha_v1,
    normalize_input_workbook,
    save_planilha_padrao_folha_v1,
)


REPO_ROOT = Path(__file__).resolve().parents[2]
MASTER_ROOT = REPO_ROOT / "data" / "company_master"
MONTHLY_FIXTURE = REPO_ROOT / "data" / "runs" / "dashboard_v1" / "run-15f4250f85" / "inputs" / "input.xlsx"


def test_detect_input_layout_identifies_canonical_template(tmp_path: Path) -> None:
    workbook_path = tmp_path / "template.xlsx"
    save_planilha_padrao_folha_v1(workbook_path)

    workbook = load_workbook(workbook_path)
    parametros = workbook["PARAMETROS"]
    parametros["B2"] = "72"
    parametros["B3"] = "Dela More"
    parametros["B4"] = "03/2024"
    parametros["B5"] = "mensal"
    parametros["B6"] = "11"
    parametros["B7"] = "v1"
    detection = detect_input_layout(workbook, registry_root=MASTER_ROOT)

    assert detection.layout_id == CANONICAL_LAYOUT_ID
    assert detection.selected_sheet_name == "LANCAMENTOS_FACEIS"


def test_detect_input_layout_identifies_monthly_layout_from_real_workbook() -> None:
    workbook = load_workbook(MONTHLY_FIXTURE)
    detection = detect_input_layout(workbook, registry_root=MASTER_ROOT)

    assert detection.layout_id == MONTHLY_LAYOUT_ID
    assert detection.company_code == "528"
    assert detection.competence == "03/2026"
    assert detection.selected_sheet_name == "mar 26"


def test_inspect_input_workbook_returns_monthly_metadata_and_real_columns() -> None:
    inspection = inspect_input_workbook(MONTHLY_FIXTURE, registry_root=MASTER_ROOT)

    assert inspection.layout_id == MONTHLY_LAYOUT_ID
    assert inspection.company_code == "528"
    assert inspection.company_name
    assert inspection.competence == "03/2026"
    assert inspection.selected_sheet_name == "mar 26"
    assert "jan 26" in inspection.source_sheet_names

    columns_by_letter = {column.column_letter: column for column in inspection.columns}
    assert columns_by_letter["A"].column_name == "COD."
    assert columns_by_letter["B"].column_name == "NOME"
    assert columns_by_letter["C"].column_name == "H. EXTRA 50% COD.150"
    assert columns_by_letter["H"].column_name == "FALTAS             COD. 8792"
    assert columns_by_letter["J"].column_name == "ATRASOS       COD. 8069"
    assert all(column.header_row == 4 for column in inspection.columns)


def test_inspect_input_workbook_returns_canonical_metadata_and_columns(tmp_path: Path) -> None:
    workbook_path = tmp_path / "template.xlsx"
    save_planilha_padrao_folha_v1(workbook_path)

    workbook = load_workbook(workbook_path)
    parametros = workbook["PARAMETROS"]
    parametros["B2"] = "72"
    parametros["B3"] = "Dela More"
    parametros["B4"] = "03/2024"
    parametros["B5"] = "mensal"
    parametros["B6"] = "11"
    parametros["B7"] = "v1"
    workbook.save(workbook_path)

    inspection = inspect_input_workbook(workbook_path, registry_root=MASTER_ROOT)

    assert inspection.layout_id == CANONICAL_LAYOUT_ID
    assert inspection.company_code == "72"
    assert inspection.competence == "03/2024"
    assert inspection.selected_sheet_name == "LANCAMENTOS_FACEIS"
    assert [column.column_name for column in inspection.columns[:4]] == [
        "linha_status",
        "chave_colaborador",
        "nome_colaborador",
        "matricula_dominio",
    ]
    assert all(column.header_row == 1 for column in inspection.columns)


def test_normalize_monthly_workbook_builds_canonical_workbook_and_report(tmp_path: Path) -> None:
    output_path = tmp_path / "normalized.xlsx"
    report_path = tmp_path / "normalization.json"

    result = normalize_input_workbook(
        MONTHLY_FIXTURE,
        output_path=output_path,
        report_path=report_path,
        registry_root=MASTER_ROOT,
    )

    assert result.layout.layout_id == MONTHLY_LAYOUT_ID
    assert result.layout.company_code == "528"
    assert result.layout.competence == "03/2026"
    assert result.workbook_path == output_path
    assert output_path.exists()
    assert report_path.exists()

    report_payload = json.loads(report_path.read_text(encoding="utf-8"))
    assert report_payload["layout"]["layout_id"] == MONTHLY_LAYOUT_ID
    assert report_payload["source_path"].endswith("input.xlsx")

    normalized_workbook = load_workbook(output_path)
    assert normalized_workbook.sheetnames[:3] == ["PARAMETROS", "FUNCIONARIOS", "LANCAMENTOS_FACEIS"]

    ingestion_result = load_planilha_padrao_folha_v1(output_path)
    assert ingestion_result.parameters.company_code == "528"
    assert ingestion_result.parameters.competence == "03/2026"


def test_unknown_layout_fails_clear(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "entrada"
    sheet["A1"] = "qualquer coisa"
    workbook_path = tmp_path / "unknown.xlsx"
    workbook.save(workbook_path)

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        detect_input_layout(load_workbook(workbook_path), registry_root=MASTER_ROOT)

    assert exc_info.value.code == "layout_desconhecido"


def test_monthly_layout_missing_company_name_fails_clear(tmp_path: Path) -> None:
    workbook_path = tmp_path / "monthly_missing_company.xlsx"
    workbook = load_workbook(MONTHLY_FIXTURE)
    workbook.active["A1"] = None
    workbook.save(workbook_path)

    with pytest.raises(InputLayoutNormalizationError) as exc_info:
        detect_input_layout(load_workbook(workbook_path), registry_root=MASTER_ROOT)

    assert exc_info.value.code == "empresa_nao_identificada"
