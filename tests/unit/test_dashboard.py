from __future__ import annotations

import importlib.util
import json
import shutil
from dataclasses import asdict
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from dashboard import (
    CompanyAdminEntry,
    ConfigResolutionStatus,
    ConfigResolver,
    DashboardActionType,
    DashboardOperationError,
    DashboardPendingItem,
    apply_dashboard_action,
    apply_workbook_cell_correction,
    create_dashboard_run_from_paths,
    ignore_pending_for_import,
    is_txt_download_enabled,
    load_company_employee_registry,
    load_company_rubric_catalog,
    load_column_mapping_profile,
    load_dashboard_state,
    list_company_admin_entries,
    save_column_mapping_profile_rule,
    save_company_admin_entry,
    upsert_employee_mapping_override,
    upsert_event_mapping_override,
    save_employee_registry_record,
    save_rubric_catalog_record,
    write_dashboard_state,
)
from ingestion import save_planilha_padrao_folha_v1


REPO_ROOT = Path(__file__).resolve().parents[2]


class _FakeStreamlit:
    def __init__(
        self,
        *,
        uploaded_workbook=None,
        button_result: bool = False,
        selected_label: str | None = None,
        form_submit_result: bool = False,
        text_inputs: dict[str, str] | None = None,
        checkboxes: dict[str, bool] | None = None,
        selectboxes: dict[str, str] | None = None,
    ) -> None:
        self.session_state: dict[str, object] = {}
        self.errors: list[str] = []
        self.infos: list[str] = []
        self.markdowns: list[str] = []
        self.subheaders: list[str] = []
        self.tables: list[object] = []
        self.tab_labels: list[str] = []
        self.selectbox_calls: list[dict[str, object]] = []
        self.button_calls: list[dict[str, object]] = []
        self.file_uploader_calls: list[dict[str, object]] = []
        self.event_log: list[str] = []
        self.uploaded_workbook = uploaded_workbook
        self.button_result = button_result
        self.selected_label = selected_label
        self.form_submit_result = form_submit_result
        self.text_inputs = text_inputs or {}
        self.checkboxes = checkboxes or {}
        self.selectboxes = selectboxes or {}
        self.rerun_called = False

    def set_page_config(self, **kwargs) -> None:
        return None

    def title(self, *args, **kwargs) -> None:
        return None

    def caption(self, *args, **kwargs) -> None:
        return None

    def subheader(self, message: str, *args, **kwargs) -> None:
        self.subheaders.append(message)

    def write(self, *args, **kwargs) -> None:
        return None

    def error(self, message: str) -> None:
        self.errors.append(message)

    def success(self, *args, **kwargs) -> None:
        return None

    def warning(self, *args, **kwargs) -> None:
        return None

    def info(self, message: str, *args, **kwargs) -> None:
        self.infos.append(message)

    def markdown(self, message: str, *args, **kwargs) -> None:
        self.markdowns.append(message)

    def table(self, data, *args, **kwargs) -> None:
        self.tables.append(data)

    def tabs(self, labels):
        self.tab_labels = list(labels)

        class _TabContext:
            def __enter__(self_inner):
                return self

            def __exit__(self_inner, exc_type, exc, tb):
                return False

        return [_TabContext() for _label in labels]

    def metric(self, *args, **kwargs) -> None:
        return None

    def columns(self, count: int):
        return [self for _ in range(count)]

    def selectbox(self, label, options=None, *args, **kwargs):
        self.selectbox_calls.append({"label": label, "options": list(options or [])})
        self.event_log.append(f"selectbox:{label}")
        if label == "Corrigir item selecionado" and self.selected_label is not None:
            return self.selected_label
        return self.selectboxes.get(label, options[0] if options else "Selecione um item")

    def form(self, *args, **kwargs):
        class _FormContext:
            def __enter__(self_inner):
                return self

            def __exit__(self_inner, exc_type, exc, tb):
                return False

        return _FormContext()

    def form_submit_button(self, *args, **kwargs):
        return self.form_submit_result

    def text_input(self, label, value="", *args, **kwargs):
        return self.text_inputs.get(label, value)

    def checkbox(self, label, value=False, *args, **kwargs):
        return self.checkboxes.get(label, value)

    def file_uploader(self, *args, **kwargs):
        self.file_uploader_calls.append({"args": args, "kwargs": kwargs})
        self.event_log.append("file_uploader")
        return self.uploaded_workbook

    def button(self, *args, **kwargs):
        self.button_calls.append({"args": args, "kwargs": kwargs})
        if kwargs.get("disabled"):
            return False
        return self.button_result

    def rerun(self):
        self.rerun_called = True


def _load_dashboard_v1_module():
    module_path = REPO_ROOT / "app" / "dashboard_v1.py"
    spec = importlib.util.spec_from_file_location("dashboard_v1_under_test", module_path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _prepare_workbook_and_config(tmp_path: Path) -> tuple[Path, Path]:
    workbook_path = tmp_path / "input.xlsx"
    config_path = tmp_path / "company_config.json"
    save_planilha_padrao_folha_v1(workbook_path)

    workbook = load_workbook(workbook_path)
    parametros = workbook["PARAMETROS"]
    parametros["B2"] = "72"
    parametros["B3"] = "Dela More"
    parametros["B4"] = "03/2024"
    parametros["B5"] = "mensal"
    parametros["B6"] = "11"
    parametros["B7"] = "v1"

    funcionarios = workbook["FUNCIONARIOS"]
    funcionarios["A2"] = "col-001"
    funcionarios["B2"] = "Ana Lima"
    funcionarios["E2"] = "123"
    funcionarios["H2"] = "ativo"
    funcionarios["I2"] = "sim"

    lancamentos = workbook["LANCAMENTOS_FACEIS"]
    lancamentos["B2"] = "col-001"
    lancamentos["C2"] = "Ana Lima"
    lancamentos["D2"] = "123"
    lancamentos["H2"] = 100
    workbook.save(workbook_path)

    config_payload = {
        "company_code": "72",
        "company_name": "Dela More",
        "default_process": "11",
        "competence": "03/2024",
        "config_version": "cfg-v1",
        "event_mappings": [
            {
                "event_negocio": "gratificacao",
                "rubrica_saida": "201",
            }
        ],
        "employee_mappings": [
            {
                "source_employee_key": "col-001",
                "source_employee_name": "Ana Lima",
                "domain_registration": "123",
            }
        ],
        "pending_policy": {
            "review_required_event_negocios": [],
            "review_required_fields": [],
            "block_on_ambiguous_observations": True,
            "block_on_unmapped_employee": True,
            "block_on_unmapped_event": True,
        },
        "validation_flags": {},
    }
    config_path.write_text(
        json.dumps(config_payload, ensure_ascii=True, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return workbook_path, config_path


def _write_internal_config(
    root: Path,
    *,
    company_code: str,
    file_name: str,
    competence: str,
    config_version: str = "cfg-v1",
    payload_override: dict | None = None,
) -> Path:
    payload = {
        "company_code": company_code,
        "company_name": "Dela More",
        "default_process": "11",
        "competence": competence,
        "config_version": config_version,
        "event_mappings": [
            {
                "event_negocio": "gratificacao",
                "rubrica_saida": "201",
            }
        ],
        "employee_mappings": [
            {
                "source_employee_key": "col-001",
                "source_employee_name": "Ana Lima",
                "domain_registration": "123",
            }
        ],
        "pending_policy": {
            "review_required_event_negocios": [],
            "review_required_fields": [],
            "block_on_ambiguous_observations": True,
            "block_on_unmapped_employee": True,
            "block_on_unmapped_event": True,
        },
        "validation_flags": {},
    }
    if payload_override:
        payload.update(payload_override)
    path = root / company_code / file_name
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=True, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return path


def _materialize_editable_workbook(paths) -> None:
    shutil.copy2(paths.raw_workbook_path, paths.editable_workbook_path)


def _persist_pending_for_action(paths, pending: DashboardPendingItem) -> None:
    state = load_dashboard_state(paths.state_path)
    updated_state = state.model_copy(update={"last_analysis": {"pendings": [asdict(pending)]}})
    write_dashboard_state(paths.state_path, updated_state)


def _employee_mapping_pending() -> DashboardPendingItem:
    return DashboardPendingItem(
        uid="mapeamento:pend-employee",
        stage="mapeamento",
        pending_id="pend-employee",
        code="mapeamento_matricula_ausente",
        severity="bloqueante",
        employee_name="Bruno Souza",
        employee_key="col-002",
        event_name=None,
        field_label="matricula do cadastro da empresa",
        found_value="sem cadastro",
        problem="Matricula nao encontrada no cadastro da empresa.",
        recommended_action="Informar a matricula Dominio do funcionario.",
        source_sheet=None,
        source_cell=None,
        source_row=None,
        source_column_name=None,
        can_edit_workbook=False,
        can_edit_employee_mapping=True,
        can_edit_event_mapping=False,
        can_ignore=False,
    )


def _event_mapping_pending() -> DashboardPendingItem:
    return DashboardPendingItem(
        uid="mapeamento:pend-event",
        stage="mapeamento",
        pending_id="pend-event",
        code="mapeamento_evento_ausente",
        severity="bloqueante",
        employee_name="Ana Lima",
        employee_key="col-001",
        event_name="horas_extras_50",
        field_label="rubrica de saida",
        found_value="sem rubrica",
        problem="Evento sem rubrica de saida.",
        recommended_action="Informar a rubrica de saida.",
        source_sheet="LANCAMENTOS_FACEIS",
        source_cell="G2",
        source_row=2,
        source_column_name="horas_extras_50",
        can_edit_workbook=False,
        can_edit_employee_mapping=False,
        can_edit_event_mapping=True,
        can_ignore=True,
        ignore_mode="evento",
        ignore_label="Ignorar este evento nesta importacao",
    )


def _column_profile_pending() -> DashboardPendingItem:
    return DashboardPendingItem(
        uid="perfil_colunas:incomplete",
        stage="perfil_colunas",
        pending_id="profile-incomplete",
        code="column_mapping_profile_incomplete",
        severity="bloqueante",
        employee_name=None,
        employee_key=None,
        event_name=None,
        field_label="perfil de mapeamento de colunas",
        found_value="empresa=528 | competencia=03/2026 | coluna=EXTRA 100%",
        problem="Perfil de mapeamento de colunas incompleto.",
        recommended_action="Cadastrar regra de perfil para a coluna.",
        source_sheet="mar 26",
        source_cell=None,
        source_row=None,
        source_column_name="EXTRA 100%",
        can_edit_workbook=False,
        can_edit_employee_mapping=False,
        can_edit_event_mapping=False,
        can_ignore=False,
    )


def _persist_column_profile_pending_for_action(paths, pending: DashboardPendingItem) -> None:
    state = load_dashboard_state(paths.state_path)
    updated_state = state.model_copy(
        update={
            "last_analysis": {
                "summary": {
                    "company_code": "528",
                    "company_name": "Empresa Mensal",
                },
                "profile_resolution": {
                    "status": "incomplete",
                    "status_label": "Perfil de mapeamento de colunas incompleto",
                    "message": "Perfil incompleto.",
                    "company_code": "528",
                    "competence": "03/2026",
                    "layout_id": "resumo_mensal_por_abas",
                    "source_path": None,
                    "missing_columns": ["EXTRA 100%"],
                },
                "pendings": [asdict(pending)],
            }
        }
    )
    write_dashboard_state(paths.state_path, updated_state)


def _fake_dashboard_result(*, paths, pendings: list[DashboardPendingItem]):
    return SimpleNamespace(
        paths=paths,
        pendings=pendings,
        state=SimpleNamespace(actions=[]),
        summary=SimpleNamespace(),
    )


def test_company_admin_entry_can_save_minimal_company_config(tmp_path: Path) -> None:
    saved = save_company_admin_entry(
        company_code="900",
        company_name="Empresa Teste",
        default_process="11",
        competence="04/2026",
        is_active=True,
        root=tmp_path / "master",
    )

    entries = list_company_admin_entries(root=tmp_path / "master")
    config_payload = json.loads((tmp_path / "master" / "company_configs.json").read_text(encoding="utf-8"))[0][
        "config_payload_internal"
    ]

    assert saved.company_code == "900"
    assert entries[0].company_name == "Empresa Teste"
    assert entries[0].default_process == "11"
    assert config_payload["event_mappings"] == []
    assert config_payload["employee_mappings"] == []


def test_dashboard_company_employee_helper_saves_registry(tmp_path: Path) -> None:
    save_employee_registry_record(
        company_code="900",
        company_name="Empresa Teste",
        domain_registration="123",
        employee_name="Ana Lima",
        aliases="ANA, A LIMA",
        root=tmp_path,
    )

    registry = load_company_employee_registry("900", root=tmp_path)
    assert registry.employees[0].domain_registration == "123"
    assert registry.employees[0].employee_name == "Ana Lima"
    assert registry.employees[0].aliases == ["ANA", "A LIMA"]


def test_dashboard_rubric_helper_saves_catalog(tmp_path: Path) -> None:
    save_rubric_catalog_record(
        company_code="900",
        company_name="Empresa Teste",
        rubric_code="201",
        description="GRATIFICACAO",
        canonical_event="gratificacao",
        value_kind="monetario",
        nature="provento",
        aliases="GRAT",
        root=tmp_path,
    )

    catalog = load_company_rubric_catalog("900", root=tmp_path)
    assert catalog.rubrics[0].rubric_code == "201"
    assert catalog.rubrics[0].canonical_event == "gratificacao"
    assert catalog.rubrics[0].aliases == ["GRAT"]


def test_dashboard_column_profile_helper_saves_ignore_rule_without_rubric(tmp_path: Path) -> None:
    save_column_mapping_profile_rule(
        company_code="900",
        company_name="Empresa Teste",
        default_process="11",
        column_name="ADIANTAMENTO",
        value_kind="monetario",
        generation_mode="ignore",
        ignore_zero=True,
        ignore_text=True,
        root=tmp_path,
    )

    profile = load_column_mapping_profile("900", root=tmp_path)
    rule = profile.mappings[0]
    assert rule.column_name == "ADIANTAMENTO"
    assert rule.enabled is False
    assert rule.generation_mode.value == "ignore"
    assert rule.rubrica_target is None
    assert rule.rubricas_target == []


def test_dashboard_column_profile_helper_rejects_ignore_rule_with_rubric(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="ignore mappings cannot define rubrica targets"):
        save_column_mapping_profile_rule(
            company_code="900",
            column_name="ADIANTAMENTO",
            value_kind="monetario",
            generation_mode="ignore",
            rubrica_target="981",
            ignore_zero=True,
            ignore_text=True,
            root=tmp_path,
        )


def test_txt_download_enabled_requires_success_and_serialized_lines() -> None:
    validation_payload = {
        "execution": {"status": "success_with_warnings"},
        "fatal_errors": [],
        "inconsistencies": [],
    }
    serialization_payload = {"counts": {"serialized": 2}}

    assert is_txt_download_enabled(
        validation_payload=validation_payload,
        serialization_payload=serialization_payload,
    )

    assert not is_txt_download_enabled(
        validation_payload=validation_payload,
        serialization_payload={"counts": {"serialized": 0}},
    )
    assert not is_txt_download_enabled(
        validation_payload={
            "execution": {"status": "blocked"},
            "fatal_errors": [],
            "inconsistencies": [],
        },
        serialization_payload=serialization_payload,
    )


def test_apply_workbook_cell_correction_updates_workbook_and_records_action(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    _materialize_editable_workbook(paths)

    action = apply_workbook_cell_correction(
        paths,
        sheet_name="LANCAMENTOS_FACEIS",
        cell="H2",
        new_value="150,00",
        pending_uid="ingestao:pend-001",
    )

    workbook = load_workbook(paths.editable_workbook_path)
    assert workbook["LANCAMENTOS_FACEIS"]["H2"].value == "150,00"
    state = load_dashboard_state(paths.state_path)
    assert len(state.actions) == 1
    assert state.actions[0].action_id == action.action_id
    assert state.actions[0].payload["cell"] == "H2"


def test_ignore_pending_clears_source_cell_and_records_action(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    _materialize_editable_workbook(paths)

    workbook = load_workbook(paths.editable_workbook_path)
    workbook["LANCAMENTOS_FACEIS"]["N2"] = "revisar"
    workbook.save(paths.editable_workbook_path)

    pending = DashboardPendingItem(
        uid="ingestao:pend-vale",
        stage="ingestao",
        pending_id="pend-vale",
        code="evento_nao_automatizavel",
        severity="media",
        employee_name="Ana Lima",
        employee_key="col-001",
        event_name="vale_transporte",
        field_label="vale_transporte",
        found_value="revisar",
        problem="Vale transporte exige revisao.",
        recommended_action="Avaliar manualmente.",
        source_sheet="LANCAMENTOS_FACEIS",
        source_cell="N2",
        source_row=2,
        source_column_name="vale_transporte",
        can_edit_workbook=True,
        can_edit_employee_mapping=False,
        can_edit_event_mapping=False,
        can_ignore=True,
        ignore_mode="evento",
        ignore_label="Ignorar este evento nesta importacao",
    )

    ignore_pending_for_import(paths, pending)

    workbook = load_workbook(paths.editable_workbook_path)
    assert workbook["LANCAMENTOS_FACEIS"]["N2"].value is None
    state = load_dashboard_state(paths.state_path)
    assert state.actions[-1].action_type.value == "ignorar_nesta_importacao"
    assert state.actions[-1].payload["cleared_cells"] == ["N2"]


def test_upsert_event_mapping_override_updates_config(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")

    upsert_event_mapping_override(
        paths,
        event_name="horas_extras_50",
        output_rubric="350",
        pending_uid="mapeamento:pend-001",
    )

    payload = json.loads(paths.editable_config_path.read_text(encoding="utf-8"))
    mapping = next(
        item for item in payload["event_mappings"] if item["event_negocio"] == "horas_extras_50"
    )
    assert mapping["rubrica_saida"] == "350"
    assert mapping["active"] is True


def test_upsert_employee_mapping_override_updates_config(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")

    upsert_employee_mapping_override(
        paths,
        employee_key="col-002",
        employee_name="Bruno Souza",
        domain_registration="456",
        pending_uid="mapeamento:pend-002",
    )

    payload = json.loads(paths.editable_config_path.read_text(encoding="utf-8"))
    mapping = next(
        item for item in payload["employee_mappings"] if item["source_employee_key"] == "col-002"
    )
    assert mapping["domain_registration"] == "456"
    assert mapping["source_employee_name"] == "Bruno Souza"
    assert mapping["active"] is True


def test_apply_dashboard_action_registers_employee_mapping_correction(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _employee_mapping_pending()
    _persist_pending_for_action(paths, pending)

    action = apply_dashboard_action(
        paths,
        action_type=DashboardActionType.EMPLOYEE_MAPPING_UPDATE,
        pending_uid=pending.uid,
        payload={"domain_registration": "456"},
    )

    payload = json.loads(paths.editable_config_path.read_text(encoding="utf-8"))
    mapping = next(
        item for item in payload["employee_mappings"] if item["source_employee_key"] == "col-002"
    )
    assert mapping["domain_registration"] == "456"
    assert mapping["source_employee_name"] == "Bruno Souza"
    assert action.action_type == DashboardActionType.EMPLOYEE_MAPPING_UPDATE
    assert action.payload["scope"] == "current_run_editable_config"

    state = load_dashboard_state(paths.state_path)
    assert state.actions[-1].action_id == action.action_id


def test_apply_dashboard_action_persists_employee_mapping_when_explicit(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _employee_mapping_pending()
    _persist_pending_for_action(paths, pending)

    action = apply_dashboard_action(
        paths,
        action_type=DashboardActionType.EMPLOYEE_MAPPING_UPDATE,
        pending_uid=pending.uid,
        payload={
            "domain_registration": "456",
            "persist_to_employee_registry": True,
            "aliases": ["BRUNO"],
        },
        employee_registry_root=tmp_path / "employee_registries",
    )

    registry = load_company_employee_registry("72", root=tmp_path / "employee_registries")
    assert registry.employees[0].employee_key == "col-002"
    assert registry.employees[0].employee_name == "Bruno Souza"
    assert registry.employees[0].domain_registration == "456"
    assert registry.employees[0].aliases == ["BRUNO"]
    assert action.payload["persist_to_employee_registry"] is True
    assert action.payload["scopes"] == ["current_run_editable_config", "company_employee_registry"]


def test_apply_dashboard_action_registers_event_mapping_correction(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _event_mapping_pending()
    _persist_pending_for_action(paths, pending)

    action = apply_dashboard_action(
        paths,
        action_type=DashboardActionType.EVENT_MAPPING_UPDATE,
        pending_uid=pending.uid,
        payload={"output_rubric": "350"},
    )

    payload = json.loads(paths.editable_config_path.read_text(encoding="utf-8"))
    mapping = next(
        item for item in payload["event_mappings"] if item["event_negocio"] == "horas_extras_50"
    )
    assert mapping["rubrica_saida"] == "350"
    assert action.action_type == DashboardActionType.EVENT_MAPPING_UPDATE
    assert action.payload["scope"] == "current_run_editable_config"

    state = load_dashboard_state(paths.state_path)
    assert state.actions[-1].action_id == action.action_id


def test_apply_dashboard_action_registers_ignore_for_current_import(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    _materialize_editable_workbook(paths)

    workbook = load_workbook(paths.editable_workbook_path)
    workbook["LANCAMENTOS_FACEIS"]["G2"] = 2
    workbook.save(paths.editable_workbook_path)

    pending = _event_mapping_pending()
    _persist_pending_for_action(paths, pending)

    action = apply_dashboard_action(
        paths,
        action_type=DashboardActionType.IGNORE_PENDING,
        pending_uid=pending.uid,
    )

    workbook = load_workbook(paths.editable_workbook_path)
    assert workbook["LANCAMENTOS_FACEIS"]["G2"].value is None
    assert action.action_type == DashboardActionType.IGNORE_PENDING
    assert action.payload["scope"] == "current_import_only"

    state = load_dashboard_state(paths.state_path)
    assert state.actions[-1].action_id == action.action_id


def test_apply_dashboard_action_persists_event_mapping_when_explicit(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _event_mapping_pending()
    _persist_pending_for_action(paths, pending)

    action = apply_dashboard_action(
        paths,
        action_type=DashboardActionType.EVENT_MAPPING_UPDATE,
        pending_uid=pending.uid,
        payload={
            "output_rubric": "350",
            "description": "HORAS EXTRAS 50%",
            "value_kind": "horas",
            "nature": "provento",
            "aliases": ["H. EXTRA 50"],
            "persist_to_rubric_catalog": True,
        },
        rubric_catalog_root=tmp_path / "rubric_catalogs",
    )

    catalog = load_company_rubric_catalog("72", root=tmp_path / "rubric_catalogs")
    assert catalog.rubrics[0].rubric_code == "350"
    assert catalog.rubrics[0].canonical_event == "horas_extras_50"
    assert catalog.rubrics[0].value_kind.value == "horas"
    assert catalog.rubrics[0].aliases == ["H. EXTRA 50"]
    assert action.payload["persist_to_rubric_catalog"] is True
    assert action.payload["scopes"] == ["current_run_editable_config", "company_rubric_catalog"]


def test_apply_dashboard_action_rejects_incomplete_rubric_catalog_payload(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _event_mapping_pending()
    _persist_pending_for_action(paths, pending)

    with pytest.raises(DashboardOperationError) as exc_info:
        apply_dashboard_action(
            paths,
            action_type=DashboardActionType.EVENT_MAPPING_UPDATE,
            pending_uid=pending.uid,
            payload={
                "output_rubric": "350",
                "persist_to_rubric_catalog": True,
            },
            rubric_catalog_root=tmp_path / "rubric_catalogs",
        )

    assert exc_info.value.code == "campo_obrigatorio_ausente"


def test_apply_dashboard_action_creates_column_mapping_profile_rule(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _column_profile_pending()
    _persist_column_profile_pending_for_action(paths, pending)

    action = apply_dashboard_action(
        paths,
        action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
        pending_uid=pending.uid,
        payload={
            "column_name": "EXTRA 100%",
            "rubrica_target": "200",
            "value_kind": "horas",
            "generation_mode": "single_line",
            "ignore_zero": True,
            "ignore_text": True,
        },
        column_profile_root=tmp_path / "profiles",
    )

    profile = load_column_mapping_profile("528", root=tmp_path / "profiles")
    assert profile.mappings[0].column_name == "EXTRA 100%"
    assert profile.mappings[0].rubrica_target == "200"
    assert action.action_type == DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE
    assert action.payload["scope"] == "company_column_mapping_profile"
    assert action.payload["column_profile_path"].endswith("528.json")


def test_apply_dashboard_action_updates_column_mapping_profile_with_multi_rubrics(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _column_profile_pending()
    _persist_column_profile_pending_for_action(paths, pending)

    apply_dashboard_action(
        paths,
        action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
        pending_uid=pending.uid,
        payload={
            "column_name": "EXTRA 100%",
            "rubrica_target": "200",
            "value_kind": "horas",
            "generation_mode": "single_line",
            "ignore_zero": True,
            "ignore_text": True,
        },
        column_profile_root=tmp_path / "profiles",
    )
    apply_dashboard_action(
        paths,
        action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
        pending_uid=pending.uid,
        payload={
            "column_name": "EXTRA 100%",
            "rubricas_target": ["8792", "8794"],
            "value_kind": "quantidade",
            "generation_mode": "multi_line",
            "ignore_zero": True,
            "ignore_text": True,
        },
        column_profile_root=tmp_path / "profiles",
    )

    profile = load_column_mapping_profile("528", root=tmp_path / "profiles")
    assert len(profile.mappings) == 1
    assert profile.mappings[0].rubricas_target == ["8792", "8794"]
    assert profile.mappings[0].generation_mode.value == "multi_line"


def test_apply_dashboard_action_marks_column_mapping_profile_rule_as_ignored(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _column_profile_pending()
    _persist_column_profile_pending_for_action(paths, pending)

    apply_dashboard_action(
        paths,
        action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
        pending_uid=pending.uid,
        payload={
            "column_name": "EXTRA 100%",
            "value_kind": "monetario",
            "generation_mode": "ignore",
            "ignore_zero": True,
            "ignore_text": True,
        },
        column_profile_root=tmp_path / "profiles",
    )

    profile = load_column_mapping_profile("528", root=tmp_path / "profiles")
    assert profile.mappings[0].enabled is False
    assert profile.mappings[0].generation_mode.value == "ignore"
    assert profile.mappings[0].target_rubrics == ()


def test_apply_dashboard_action_rejects_ignored_column_mapping_with_single_rubric(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _column_profile_pending()
    _persist_column_profile_pending_for_action(paths, pending)
    profile_root = tmp_path / "profiles"

    with pytest.raises(DashboardOperationError) as exc_info:
        apply_dashboard_action(
            paths,
            action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
            pending_uid=pending.uid,
            payload={
                "column_name": "EXTRA 100%",
                "rubrica_target": "999",
                "value_kind": "monetario",
                "generation_mode": "ignore",
                "ignore_zero": True,
                "ignore_text": True,
            },
            column_profile_root=profile_root,
        )

    assert exc_info.value.code == "perfil_colunas_regra_invalida"
    assert not (profile_root / "528.json").exists()


def test_apply_dashboard_action_rejects_ignored_column_mapping_with_multiple_rubrics(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _column_profile_pending()
    _persist_column_profile_pending_for_action(paths, pending)
    profile_root = tmp_path / "profiles"

    with pytest.raises(DashboardOperationError) as exc_info:
        apply_dashboard_action(
            paths,
            action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
            pending_uid=pending.uid,
            payload={
                "column_name": "EXTRA 100%",
                "rubricas_target": ["999", "998"],
                "value_kind": "monetario",
                "generation_mode": "ignore",
                "ignore_zero": True,
                "ignore_text": True,
            },
            column_profile_root=profile_root,
        )

    assert exc_info.value.code == "perfil_colunas_regra_invalida"
    assert not (profile_root / "528.json").exists()


def test_apply_dashboard_action_rejects_invalid_column_mapping_profile_payload(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _column_profile_pending()
    _persist_column_profile_pending_for_action(paths, pending)

    with pytest.raises(DashboardOperationError) as exc_info:
        apply_dashboard_action(
            paths,
            action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
            pending_uid=pending.uid,
            payload={
                "column_name": "EXTRA 100%",
                "rubrica_target": "200",
                "value_kind": "horas",
                "generation_mode": "automatico",
                "ignore_zero": True,
                "ignore_text": True,
            },
            column_profile_root=tmp_path / "profiles",
        )

    assert exc_info.value.code == "perfil_colunas_regra_invalida"


def test_apply_dashboard_action_rejects_column_mapping_profile_action_for_other_pending(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _event_mapping_pending()
    _persist_pending_for_action(paths, pending)

    with pytest.raises(DashboardOperationError) as exc_info:
        apply_dashboard_action(
            paths,
            action_type=DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE,
            pending_uid=pending.uid,
            payload={
                "column_name": "EXTRA 100%",
                "rubrica_target": "200",
                "value_kind": "horas",
                "generation_mode": "single_line",
                "ignore_zero": True,
                "ignore_text": True,
            },
            column_profile_root=tmp_path / "profiles",
        )

    assert exc_info.value.code == "acao_incompativel_com_pendencia"


def test_apply_dashboard_action_rejects_invalid_payload(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _event_mapping_pending()
    _persist_pending_for_action(paths, pending)

    with pytest.raises(DashboardOperationError) as exc_info:
        apply_dashboard_action(
            paths,
            action_type=DashboardActionType.EVENT_MAPPING_UPDATE,
            pending_uid=pending.uid,
            payload={"output_rubric": "  "},
        )

    assert exc_info.value.code == "campo_obrigatorio_ausente"


def test_apply_dashboard_action_rejects_incompatible_pending(tmp_path: Path) -> None:
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _employee_mapping_pending()
    _persist_pending_for_action(paths, pending)

    with pytest.raises(DashboardOperationError) as exc_info:
        apply_dashboard_action(
            paths,
            action_type=DashboardActionType.EVENT_MAPPING_UPDATE,
            pending_uid=pending.uid,
            payload={"output_rubric": "350"},
        )

    assert exc_info.value.code == "acao_incompativel_com_pendencia"


def test_dashboard_main_shows_last_error_before_run_root_guard(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit()
    fake_st.session_state[module.ERROR_KEY] = "A analise nao conseguiu ser concluida: boom"
    module.st = fake_st
    module._render_upload_area = lambda: None
    module._render_assisted_report_importer_tab = lambda: None
    module._render_company_registration_tab = lambda: None
    module._render_employees_tab = lambda: None
    module._render_rubrics_tab = lambda: None
    module._render_column_profile_tab = lambda: None
    module._render_txt_audit_tab = lambda result: None

    build_called = {"count": 0}

    def _fail_build_dashboard_paths(*args, **kwargs):
        build_called["count"] += 1
        raise AssertionError("build_dashboard_paths nao deveria ser chamado sem run_root")

    module.build_dashboard_paths = _fail_build_dashboard_paths

    module.main()

    assert fake_st.errors == ["A analise nao conseguiu ser concluida: boom"]
    assert build_called["count"] == 0


def test_dashboard_main_renders_registration_tabs() -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit()
    module.st = fake_st
    module._render_import_tab = lambda result: None
    module._render_assisted_report_importer_tab = lambda: None
    module._render_company_registration_tab = lambda: None
    module._render_employees_tab = lambda: None
    module._render_rubrics_tab = lambda: None
    module._render_column_profile_tab = lambda: None
    module._render_txt_audit_tab = lambda result: None

    module.main()

    assert tuple(fake_st.tab_labels) == module.TAB_LABELS
    assert "Importador assistido de relatorios" in module.TAB_LABELS


def test_render_assisted_report_importer_requires_company_before_upload() -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit()
    module.st = fake_st
    module._render_company_selector = lambda *args, **kwargs: None

    module._render_assisted_report_importer_tab()

    assert fake_st.file_uploader_calls == []


def test_render_assisted_report_importer_shows_report_upload_after_company_selection() -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit()
    module.st = fake_st
    selected_company = CompanyAdminEntry(
        company_code="72",
        company_name="Dela More",
        status="active",
        is_active=True,
        company_id="company:72",
        default_process="11",
        competence="03/2024",
    )
    module._render_company_selector = lambda *args, **kwargs: selected_company

    module._render_assisted_report_importer_tab()

    assert fake_st.file_uploader_calls[0]["args"][0] == "Relatorio de folha/resumo (.pdf, .txt, .csv, .xlsx)"


def test_render_upload_area_shows_company_selector_before_upload(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit()
    module.st = fake_st
    entry = CompanyAdminEntry(
        company_code="72",
        company_name="Dela More",
        status="active",
        is_active=True,
        company_id="company:72",
        default_process="11",
        competence="03/2024",
    )
    module.list_company_admin_entries = lambda: (entry,)
    module.get_company_admin_entry = lambda company_code: entry

    module._render_upload_area()

    assert fake_st.selectbox_calls[0]["label"] == "Empresa para esta importacao"
    assert fake_st.event_log.index("selectbox:Empresa para esta importacao") < fake_st.event_log.index("file_uploader")


def test_render_upload_area_blocks_import_without_selected_company(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit(
        uploaded_workbook=SimpleNamespace(name="entrada.xlsx", getvalue=lambda: b"conteudo"),
        button_result=True,
        text_inputs={"Competencia informada (opcional)": "03/2024"},
    )
    module.st = fake_st
    module._render_company_selector = lambda *args, **kwargs: None

    calls = {"create": 0}
    module.create_dashboard_run_from_uploads = lambda *args, **kwargs: calls.__setitem__("create", calls["create"] + 1)

    module._render_upload_area()

    assert calls["create"] == 0
    assert fake_st.button_calls[-1]["kwargs"]["disabled"] is True


def test_render_upload_area_clears_previous_error_after_successful_analysis(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit(
        uploaded_workbook=SimpleNamespace(name="entrada.xlsx", getvalue=lambda: b"conteudo"),
        button_result=True,
        text_inputs={"Competencia informada (opcional)": "03/2024"},
    )
    fake_st.session_state[module.ERROR_KEY] = "erro antigo"
    module.st = fake_st
    selected_company = CompanyAdminEntry(
        company_code="72",
        company_name="Dela More",
        status="active",
        is_active=True,
        company_id="company:72",
        default_process="11",
        competence="03/2024",
    )
    module._render_company_selector = lambda *args, **kwargs: selected_company

    calls: dict[str, object] = {}

    def _create_dashboard_run_from_uploads(*, workbook_name: str, workbook_bytes: bytes):
        calls["create"] = {
            "workbook_name": workbook_name,
            "workbook_bytes": workbook_bytes,
        }
        return SimpleNamespace(run_root=tmp_path / "runs" / "run-001")

    def _run_dashboard_analysis(paths, **kwargs):
        calls["run"] = paths
        calls["run_kwargs"] = kwargs

    module.create_dashboard_run_from_uploads = _create_dashboard_run_from_uploads
    module.run_dashboard_analysis = _run_dashboard_analysis

    module._render_upload_area()

    assert calls["create"] == {"workbook_name": "entrada.xlsx", "workbook_bytes": b"conteudo"}
    assert calls["run"].run_root == tmp_path / "runs" / "run-001"
    assert calls["run_kwargs"]["selected_company_code"] == "72"
    assert calls["run_kwargs"]["selected_company_name"] == "Dela More"
    assert calls["run_kwargs"]["selected_competence"] == "03/2024"
    assert fake_st.session_state[module.RUN_ROOT_KEY] == str(tmp_path / "runs" / "run-001")
    assert fake_st.session_state[module.ERROR_KEY] is None
    assert fake_st.rerun_called is True


def test_dashboard_main_continues_normal_flow_when_run_root_exists(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit()
    module.st = fake_st
    module._render_upload_area = lambda: None
    module._render_assisted_report_importer_tab = lambda: None
    module._render_company_registration_tab = lambda: None
    module._render_employees_tab = lambda: None
    module._render_rubrics_tab = lambda: None
    module._render_column_profile_tab = lambda: None

    state_path = tmp_path / "state.json"
    state_path.write_text("{}", encoding="utf-8")
    fake_paths = SimpleNamespace(state_path=state_path)
    module.build_dashboard_paths = lambda run_root: fake_paths

    called = {"load": 0, "summary": 0, "txt_audit": 0, "pendings": 0, "history": 0, "downloads": 0}

    def _load_dashboard_run(paths):
        called["load"] += 1
        return SimpleNamespace(
            summary=SimpleNamespace(),
            pendings=[],
            state=SimpleNamespace(actions=[]),
            paths=paths,
        )

    module.load_dashboard_run = _load_dashboard_run
    module._render_summary = lambda result: called.__setitem__("summary", called["summary"] + 1)
    module._render_txt_audit = lambda result: called.__setitem__("txt_audit", called["txt_audit"] + 1)
    module._render_pendings = lambda result: called.__setitem__("pendings", called["pendings"] + 1)
    module._render_actions_history = lambda result: called.__setitem__("history", called["history"] + 1)
    module._render_downloads = lambda result: called.__setitem__("downloads", called["downloads"] + 1)

    fake_st.session_state[module.RUN_ROOT_KEY] = str(tmp_path / "run-root")

    module.main()

    assert called == {"load": 1, "summary": 1, "txt_audit": 1, "pendings": 1, "history": 1, "downloads": 1}
    assert fake_st.errors == []


def test_render_pendings_displays_operational_columns(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit()
    module.st = fake_st
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _event_mapping_pending()

    module._render_pendings(_fake_dashboard_result(paths=paths, pendings=[pending]))

    assert fake_st.tables
    assert fake_st.tables[0][0]["Severidade"] == "bloqueante"
    assert fake_st.tables[0][0]["Etapa"] == "mapeamento"
    assert fake_st.tables[0][0]["Codigo"] == "mapeamento_evento_ausente"
    assert fake_st.tables[0][0]["Evento"] == "horas_extras_50"


def test_render_txt_audit_is_hidden_when_no_serialized_lines(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    fake_st = _FakeStreamlit()
    module.st = fake_st
    txt_path = tmp_path / "input.txt"
    txt_path.write_text("", encoding="utf-8")

    module._render_txt_audit(
        SimpleNamespace(
            paths=SimpleNamespace(txt_path=txt_path),
            summary=SimpleNamespace(serialized_line_count=0),
        )
    )

    assert fake_st.subheaders == []
    assert fake_st.infos == []
    assert fake_st.tables == []


def test_render_pendings_employee_form_calls_manual_action_and_reprocesses(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _employee_mapping_pending()
    fake_st = _FakeStreamlit(
        selected_label=pending.selection_label(),
        form_submit_result=True,
        text_inputs={"Matricula Dominio corrigida": "456"},
        checkboxes={"Salvar no cadastro da empresa para proximas importacoes": True},
    )
    module.st = fake_st
    calls: dict[str, object] = {}

    def _apply_dashboard_action(*args, **kwargs):
        calls["action"] = kwargs

    def _run_dashboard_analysis(paths_arg):
        calls["run"] = paths_arg

    module.apply_dashboard_action = _apply_dashboard_action
    module.run_dashboard_analysis = _run_dashboard_analysis

    module._render_pendings(_fake_dashboard_result(paths=paths, pendings=[pending]))

    assert calls["action"]["action_type"] == DashboardActionType.EMPLOYEE_MAPPING_UPDATE
    assert calls["action"]["pending_uid"] == pending.uid
    assert calls["action"]["payload"] == {
        "domain_registration": "456",
        "persist_to_employee_registry": True,
    }
    assert calls["run"] == paths
    assert fake_st.rerun_called is True


def test_render_pendings_rubric_form_calls_manual_action_with_explicit_persistence(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _event_mapping_pending()
    fake_st = _FakeStreamlit(
        selected_label=pending.selection_label(),
        form_submit_result=True,
        text_inputs={
            "Rubrica de saida corrigida": "350",
            "Descricao da rubrica": "HORAS EXTRAS 50%",
            "Evento canonico": "horas_extras_50",
        },
        checkboxes={"Salvar no catalogo de rubricas da empresa para proximas importacoes": True},
        selectboxes={"Tipo do valor": "horas", "Natureza": "provento"},
    )
    module.st = fake_st
    calls: dict[str, object] = {}
    module.apply_dashboard_action = lambda *args, **kwargs: calls.setdefault("action", kwargs)
    module.run_dashboard_analysis = lambda paths_arg: calls.setdefault("run", paths_arg)

    module._render_pendings(_fake_dashboard_result(paths=paths, pendings=[pending]))

    assert calls["action"]["action_type"] == DashboardActionType.EVENT_MAPPING_UPDATE
    assert calls["action"]["payload"] == {
        "output_rubric": "350",
        "persist_to_rubric_catalog": True,
        "description": "HORAS EXTRAS 50%",
        "value_kind": "horas",
        "canonical_event": "horas_extras_50",
        "nature": "provento",
    }
    assert calls["run"] == paths


def test_render_pendings_column_profile_ignore_does_not_send_rubric(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _column_profile_pending()
    fake_st = _FakeStreamlit(
        selected_label=pending.selection_label(),
        form_submit_result=True,
        text_inputs={"Coluna": "EXTRA 100%"},
        selectboxes={
            "Modo de geracao": "ignore",
            "Tipo do valor da coluna": "monetario",
        },
    )
    module.st = fake_st
    calls: dict[str, object] = {}
    module.apply_dashboard_action = lambda *args, **kwargs: calls.setdefault("action", kwargs)
    module.run_dashboard_analysis = lambda paths_arg: calls.setdefault("run", paths_arg)

    module._render_pendings(_fake_dashboard_result(paths=paths, pendings=[pending]))

    assert calls["action"]["action_type"] == DashboardActionType.COLUMN_MAPPING_PROFILE_UPDATE
    assert calls["action"]["payload"] == {
        "column_name": "EXTRA 100%",
        "value_kind": "monetario",
        "generation_mode": "ignore",
        "ignore_zero": True,
        "ignore_text": True,
    }
    assert "rubrica_target" not in calls["action"]["payload"]
    assert "rubricas_target" not in calls["action"]["payload"]
    assert calls["run"] == paths


def test_render_pendings_ignore_button_calls_manual_action_and_reprocesses(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _event_mapping_pending()
    fake_st = _FakeStreamlit(
        selected_label=pending.selection_label(),
        form_submit_result=False,
        button_result=True,
    )
    module.st = fake_st
    calls: dict[str, object] = {}
    module.apply_dashboard_action = lambda *args, **kwargs: calls.setdefault("action", kwargs)
    module.run_dashboard_analysis = lambda paths_arg: calls.setdefault("run", paths_arg)

    module._render_pendings(_fake_dashboard_result(paths=paths, pendings=[pending]))

    assert calls["action"]["action_type"] == DashboardActionType.IGNORE_PENDING
    assert calls["action"]["pending_uid"] == pending.uid
    assert calls["run"] == paths


def test_render_pendings_shows_backend_error_without_masking(tmp_path: Path) -> None:
    module = _load_dashboard_v1_module()
    workbook_path, config_path = _prepare_workbook_and_config(tmp_path)
    paths = create_dashboard_run_from_paths(workbook_path, config_path, runs_root=tmp_path / "runs")
    pending = _employee_mapping_pending()
    fake_st = _FakeStreamlit(
        selected_label=pending.selection_label(),
        form_submit_result=True,
        text_inputs={"Matricula Dominio corrigida": "456"},
    )
    module.st = fake_st

    def _raise_backend_error(*args, **kwargs):
        raise RuntimeError("falha backend")

    module.apply_dashboard_action = _raise_backend_error
    module.run_dashboard_analysis = lambda paths_arg: None

    module._render_pendings(_fake_dashboard_result(paths=paths, pendings=[pending]))

    assert fake_st.session_state[module.ERROR_KEY] == "Falha ao salvar a matricula: falha backend"
    assert fake_st.rerun_called is True


def test_config_resolver_prefers_specific_company_competence(tmp_path: Path) -> None:
    root = tmp_path / "configs"
    _write_internal_config(root, company_code="72", file_name="03-2024.json", competence="03/2024", config_version="cfg-specific")
    _write_internal_config(root, company_code="72", file_name="active.json", competence="01/2024", config_version="cfg-active")

    result = ConfigResolver(registry_root=tmp_path / "master", legacy_root=root).resolve(
        company_code="72", competence="03/2024"
    )

    assert result.status == ConfigResolutionStatus.FOUND
    assert result.config_source == "legacy_company_competence"
    assert result.config_version == "cfg-specific"


def test_config_resolver_returns_not_found_when_no_internal_config_exists(tmp_path: Path) -> None:
    result = ConfigResolver(registry_root=tmp_path / "master", legacy_root=tmp_path / "configs").resolve(
        company_code="72", competence="03/2024"
    )

    assert result.status == ConfigResolutionStatus.NOT_FOUND


def test_config_resolver_returns_ambiguous_when_multiple_specific_candidates_exist(tmp_path: Path) -> None:
    root = tmp_path / "configs"
    _write_internal_config(root, company_code="72", file_name="03-2024.json", competence="03/2024")
    _write_internal_config(root, company_code="72", file_name="03_2024.json", competence="03/2024")

    result = ConfigResolver(registry_root=tmp_path / "master", legacy_root=root).resolve(
        company_code="72", competence="03/2024"
    )

    assert result.status == ConfigResolutionStatus.AMBIGUOUS


def test_config_resolver_returns_mismatch_for_wrong_company(tmp_path: Path) -> None:
    root = tmp_path / "configs"
    _write_internal_config(
        root,
        company_code="72",
        file_name="03-2024.json",
        competence="03/2024",
        payload_override={"company_code": "99"},
    )

    result = ConfigResolver(registry_root=tmp_path / "master", legacy_root=root).resolve(
        company_code="72", competence="03/2024"
    )

    assert result.status == ConfigResolutionStatus.MISMATCH


def test_config_resolver_falls_back_to_active_json(tmp_path: Path) -> None:
    root = tmp_path / "configs"
    _write_internal_config(root, company_code="72", file_name="active.json", competence="01/2024", config_version="cfg-active")

    result = ConfigResolver(registry_root=tmp_path / "master", legacy_root=root).resolve(
        company_code="72", competence="03/2024"
    )

    assert result.status == ConfigResolutionStatus.FOUND
    assert result.config_source == "legacy_company_active"
    assert result.config_version == "cfg-active"
    assert result.config_payload is not None
    assert result.config_payload["competence"] == "03/2024"
