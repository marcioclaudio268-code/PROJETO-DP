from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ingestion import ingest_template_v1_workbook, save_planilha_padrao_folha_v1
from serialization import deserialize_mapped_artifact, encode_mapped_movement_to_txt_line


def _mapped_payload_from_ingested_movements(movements: list[dict]) -> dict:
    return {
        "artifact_version": "mapping_result_v1",
        "execution": {
            "engine_version": "0.1.0",
            "status": "success",
        },
        "snapshot": {
            "snapshot_version": "ingestion_snapshot_v1",
            "company_code": "755",
            "company_name": "GUSTAVO LOPES LACERDA",
            "competence": "04/2026",
            "layout_version": "v1",
            "movement_count": len(movements),
            "pending_count": 0,
            "execution_status": "success",
        },
        "config": {
            "company_code": "755",
            "company_name": "GUSTAVO LOPES LACERDA",
            "competence": "04/2026",
            "config_version": "test-contract",
            "default_process": "11",
            "active_event_mappings": len(movements),
            "active_employee_mappings": 1,
        },
        "mapped_movements": movements,
        "mapping_pendings": [],
        "counts": {
            "mapped_movements": len(movements),
            "ready_movements": len(movements),
            "blocked_movements": 0,
            "mapping_pendings": 0,
            "blocking_mapping_pendings": 0,
        },
    }


def _movement_payload(
    *,
    movement_id: str,
    registration: str,
    competence: str,
    rubric: str,
    value_type: str,
    amount: str | None = None,
    quantity: str | None = None,
    hours_text: str | None = None,
    event_name: str | None = None,
) -> dict:
    hours = None
    if hours_text is not None:
        hour_text, minute_text = hours_text.split(":")
        hours = {
            "text": hours_text,
            "total_minutes": int(hour_text) * 60 + int(minute_text),
        }

    return {
        "canonical_movement_id": movement_id,
        "company_code": "755",
        "competence": competence,
        "payroll_type": "mensal",
        "default_process": "11",
        "employee_key": registration,
        "employee_name": "ADILSON RAFAEL DE SOUSA",
        "event_name": event_name or rubric,
        "value_type": value_type,
        "quantity": quantity,
        "hours": hours,
        "amount": amount,
        "source": {
            "sheet_name": "MOVIMENTOS_CANONICOS",
            "row_number": 2,
            "cell": "A2",
            "column_name": "rubrica_saida",
        },
        "canonical_domain_registration": registration,
        "resolved_domain_registration": registration,
        "output_rubric": rubric,
        "status": "pronto_para_serializer",
        "canonical_blocked": False,
        "inherited_pending_codes": [],
        "inherited_pending_messages": [],
        "mapping_pending_codes": [],
        "mapping_pending_messages": [],
        "observation": None,
        "informed_rubric": rubric,
        "event_nature": None,
        "serialization_unit": "BRL" if value_type == "monetario" else ("HH:MM" if value_type == "horas" else "DIAS"),
    }


def test_prefilled_movimentos_canonicos_flow_preserves_txt_contract(tmp_path: Path) -> None:
    workbook_path = tmp_path / "prefilled.xlsx"
    save_planilha_padrao_folha_v1(workbook_path)
    workbook = load_workbook(workbook_path)

    parametros = workbook["PARAMETROS"]
    parametros["B2"] = "755"
    parametros["B3"] = "GUSTAVO LOPES LACERDA"
    parametros["B4"] = "04/2026"
    parametros["B5"] = "mensal"
    parametros["B6"] = "11"
    parametros["B7"] = "v1"

    movimentos = workbook["MOVIMENTOS_CANONICOS"]
    movimentos["A2"] = "mov-00001"
    movimentos["B2"] = "755"
    movimentos["C2"] = "04/2026"
    movimentos["D2"] = "mensal"
    movimentos["E2"] = "11"
    movimentos["F2"] = "304"
    movimentos["G2"] = "ADILSON RAFAEL DE SOUSA"
    movimentos["H2"] = "304"
    movimentos["I2"] = "8907"
    movimentos["J2"] = "8907"
    movimentos["K2"] = "8907"
    movimentos["L2"] = "provento"
    movimentos["M2"] = "monetario"
    movimentos["O2"] = "100,00"
    movimentos["P2"] = "BRL"
    movimentos["Q2"] = "ABRIL 26"
    movimentos["R2"] = "B3"
    movimentos["S2"] = "B"

    workbook.save(workbook_path)

    ingested = ingest_template_v1_workbook(load_workbook(workbook_path))
    movement = ingested.movements[0]
    artifact = deserialize_mapped_artifact(
        _mapped_payload_from_ingested_movements(
            [
                _movement_payload(
                    movement_id=movement.movement_id,
                    registration=movement.domain_registration or "",
                    competence=movement.competence,
                    rubric=movement.output_rubric or "",
                    value_type=movement.value_type.value,
                    amount=movement.amount_for_sheet(),
                    event_name=movement.event_name,
                )
            ]
        )
    )

    assert encode_mapped_movement_to_txt_line(artifact.movements[0]) == "1000000003042026048907110000100000000000755"


def test_reduced_official_contract_lines_match_expected_txt_output() -> None:
    movements = [
        _movement_payload(
            movement_id="mov-00001",
            registration="304",
            competence="202604",
            rubric="8907",
            value_type="monetario",
            amount="100",
        ),
        _movement_payload(
            movement_id="mov-00002",
            registration="304",
            competence="202604",
            rubric="48",
            value_type="dias",
            quantity="6",
        ),
        _movement_payload(
            movement_id="mov-00003",
            registration="304",
            competence="202604",
            rubric="981",
            value_type="monetario",
            amount="325",
        ),
        _movement_payload(
            movement_id="mov-00004",
            registration="304",
            competence="202604",
            rubric="201",
            value_type="horas",
            hours_text="15:07",
        ),
        _movement_payload(
            movement_id="mov-00005",
            registration="304",
            competence="202604",
            rubric="200",
            value_type="horas",
            hours_text="27:02",
        ),
        _movement_payload(
            movement_id="mov-00006",
            registration="304",
            competence="202604",
            rubric="211",
            value_type="monetario",
            amount="102.50",
        ),
    ]
    artifact = deserialize_mapped_artifact(_mapped_payload_from_ingested_movements(movements))

    assert [encode_mapped_movement_to_txt_line(movement) for movement in artifact.movements] == [
        "1000000003042026048907110000100000000000755",
        "1000000003042026040048110000006000000000755",
        "1000000003042026040981110000325000000000755",
        "1000000003042026040201110000015070000000755",
        "1000000003042026040200110000027020000000755",
        "1000000003042026040211110000102500000000755",
    ]
