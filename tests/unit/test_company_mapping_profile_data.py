from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from dashboard import ColumnValueKind, CompanyColumnMappingProfile, load_column_mapping_profile
from dashboard.service import _resolve_column_mapping_profile
from ingestion import InputColumnMetadata, InputLayoutDetection, InputWorkbookInspection, MONTHLY_LAYOUT_ID


REPO_ROOT = Path(__file__).resolve().parents[2]
PROFILE_ROOT = REPO_ROOT / "data" / "company_mapping_profiles"


def _mapping_by_column(profile: CompanyColumnMappingProfile, column_name: str):
    return next(mapping for mapping in profile.mappings if mapping.column_name == column_name)


def _load_real_profile(company_code: str) -> CompanyColumnMappingProfile:
    return load_column_mapping_profile(company_code, root=PROFILE_ROOT)


def test_real_berbella_profile_loads_with_consolidated_mappings() -> None:
    profile = _load_real_profile("887")

    assert profile.company_name == "BERBELLA LTDA"
    assert _mapping_by_column(profile, "GRAT.").target_rubrics == ("20",)
    assert _mapping_by_column(profile, "EXTRA 70%").target_rubrics == ("201",)
    assert _mapping_by_column(profile, "EXTRA 100%").target_rubrics == ("200",)
    assert _mapping_by_column(profile, "EXTRA NOTURNA").target_rubrics == ("25",)
    assert _mapping_by_column(profile, "ATRASO").target_rubrics == ("8069",)
    assert _mapping_by_column(profile, "FALTA").target_rubrics == ("8792", "8794")
    assert _mapping_by_column(profile, "DESPESAS").target_rubrics == ("204",)


def test_real_mqb_profile_loads_with_consolidated_mappings() -> None:
    profile = _load_real_profile("1016")

    assert profile.company_name == "MAIS QUE BOLO DOCES E SALGADOS LTDA"
    assert _mapping_by_column(profile, "GRAT.").target_rubrics == ("20",)
    assert _mapping_by_column(profile, "EXTRA 70%").target_rubrics == ("201",)
    assert _mapping_by_column(profile, "EXTRA 100%").target_rubrics == ("200",)
    assert _mapping_by_column(profile, "HORA EXTRA NOTURNA").target_rubrics == ("25",)
    assert _mapping_by_column(profile, "ATRASO").target_rubrics == ("8069",)
    assert _mapping_by_column(profile, "FALTA").target_rubrics == ("8792", "8794")
    assert _mapping_by_column(profile, "DESPESAS").target_rubrics == ("202",)


def test_real_saad_profile_loads_with_consolidated_mappings() -> None:
    profile = _load_real_profile("448")

    assert profile.company_name == "SAAD E TOSSI LTDA ME"
    assert _mapping_by_column(profile, "GRAT.").target_rubrics == ("20",)
    assert _mapping_by_column(profile, "EXTRA 70%").target_rubrics == ("219",)
    assert _mapping_by_column(profile, "EXTRA 100%").target_rubrics == ("200",)
    assert _mapping_by_column(profile, "HORA EXTRA NOTURNA").target_rubrics == ("25",)
    assert _mapping_by_column(profile, "HORA EXTRA NOTURNA 100%").target_rubrics == ("25",)
    assert _mapping_by_column(profile, "ATRASO").target_rubrics == ("8069",)
    assert _mapping_by_column(profile, "FALTA").target_rubrics == ("8792", "8794")
    assert _mapping_by_column(profile, "DESPESAS").target_rubrics == ("264",)


def test_real_profiles_preserve_company_specific_rubric_differences() -> None:
    berbella = _load_real_profile("887")
    mqb = _load_real_profile("1016")
    saad = _load_real_profile("448")

    assert _mapping_by_column(berbella, "DESPESAS").target_rubrics == ("204",)
    assert _mapping_by_column(mqb, "DESPESAS").target_rubrics == ("202",)
    assert _mapping_by_column(saad, "DESPESAS").target_rubrics == ("264",)

    assert _mapping_by_column(berbella, "EXTRA 70%").target_rubrics == ("201",)
    assert _mapping_by_column(mqb, "EXTRA 70%").target_rubrics == ("201",)
    assert _mapping_by_column(saad, "EXTRA 70%").target_rubrics == ("219",)


@pytest.mark.parametrize("company_code", ["887", "1016", "448"])
def test_dashboard_profile_resolution_finds_materialized_company_profiles(
    tmp_path: Path,
    company_code: str,
) -> None:
    profile = _load_real_profile(company_code)
    source_path, inspection = _build_profile_resolution_fixture(tmp_path, profile)

    resolution, loaded_profile = _resolve_column_mapping_profile(
        inspection=inspection,
        source_workbook_path=source_path,
        profile_root=PROFILE_ROOT,
    )

    assert resolution.status == "found"
    assert resolution.company_code == company_code
    assert loaded_profile == profile


def _build_profile_resolution_fixture(
    tmp_path: Path,
    profile: CompanyColumnMappingProfile,
) -> tuple[Path, InputWorkbookInspection]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "mar 26"
    sheet["A1"] = profile.company_name
    sheet["A2"] = "Lancamentos de teste por perfil"
    sheet["A4"] = "COD."
    sheet["B4"] = "NOME"
    sheet["A6"] = "col-001"
    sheet["B6"] = "Ana Lima"

    columns = [
        InputColumnMetadata(1, "A", "COD.", "cod", 4),
        InputColumnMetadata(2, "B", "NOME", "nome", 4),
    ]
    for offset, mapping in enumerate(profile.mappings, start=3):
        column_letter = _column_letter(offset)
        column_name = mapping.column_name or mapping.column_key or f"COLUNA_{offset}"
        sheet.cell(row=4, column=offset, value=column_name)
        sheet.cell(row=6, column=offset, value=_sample_value_for_mapping(mapping.value_kind))
        columns.append(
            InputColumnMetadata(
                column_index=offset,
                column_letter=column_letter,
                column_name=column_name,
                normalized_column_name=column_name.lower(),
                header_row=4,
            )
        )

    path = tmp_path / f"{profile.company_code}.xlsx"
    workbook.save(path)

    detection = InputLayoutDetection(
        layout_id=MONTHLY_LAYOUT_ID,
        active_sheet_name="mar 26",
        selected_sheet_name="mar 26",
        selected_sheet_reason="profile_resolution_test",
        company_code=profile.company_code,
        company_name=profile.company_name or profile.company_code,
        competence="03/2026",
        source_company_name=profile.company_name or profile.company_code,
        source_title_text="Lancamentos de teste por perfil",
        source_sheet_names=("mar 26",),
        rules_applied=("profile_resolution_test",),
    )
    inspection = InputWorkbookInspection(
        layout_id=MONTHLY_LAYOUT_ID,
        company_code=profile.company_code,
        company_name=profile.company_name or profile.company_code,
        competence="03/2026",
        selected_sheet_name="mar 26",
        source_sheet_names=("mar 26",),
        columns=tuple(columns),
        warnings=(),
        detection=detection,
    )
    return path, inspection


def _sample_value_for_mapping(value_kind: ColumnValueKind) -> object:
    if value_kind == ColumnValueKind.HOURS:
        return "01:00"
    if value_kind == ColumnValueKind.QUANTITY:
        return 1
    return "10,00"


def _column_letter(column_index: int) -> str:
    letters = ""
    quotient = column_index
    while quotient:
        quotient, remainder = divmod(quotient - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
